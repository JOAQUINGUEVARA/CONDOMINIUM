import django_filters
from django.shortcuts import render
from django.views.generic.edit import CreateView, UpdateView, DeleteView,UpdateView
from django.views.generic.list import ListView
from django.views.generic.base import TemplateView
from django.views.generic.detail import DetailView
from django.urls import reverse, reverse_lazy
from django.http import HttpResponseBadRequest,HttpResponse, HttpRequest, JsonResponse,HttpResponseRedirect
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth import authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login as do_login
from django.contrib.auth import logout as do_logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required,permission_required
from django.template.loader import render_to_string
from django.core.paginator import Paginator
import datetime 
from datetime import datetime, timedelta,date
from time import gmtime, strftime
#from bootstrap_modal_forms.generic import BSModalCreateView
from django.db.models import Subquery,F, Count, Value,Q
from django.views.decorators.csrf import csrf_exempt
from django.db.models.signals import post_save,post_delete
from django.dispatch import receiver
from django import forms
#from django.contrib.auth.models import User
#from django.contrib.auth import get_user_model
#User = get_user_model()
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.core.mail import EmailMessage
from django.contrib.auth.mixins import LoginRequiredMixin
#from django.db.models.signals import pre_save
from django.template.loader import render_to_string
from django.contrib import messages
from django.db.models import Sum
import io
from io import StringIO
import xlsxwriter
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch, cm
from reportlab.platypus import Table,TableStyle
from reportlab.lib import colors
from io import BytesIO
from reportlab.lib.pagesizes import letter
from .pdf import draw_pdf,pdf_response
from bootstrap_datepicker_plus.widgets import DateTimePickerInput
from django.core.mail import send_mail, BadHeaderError
from django.views import View
from .forms import NewUserForm
from django.contrib.auth import login
from django.contrib import messages
import time

from .models import IngresoPeatonal, ProponenteProyecto,Visitante,Correspondencia,ZonaComun,UserPerfil,Residente,Parqueadero,Mascota,Deposito,Contrato, MiembroConsejo,Reservas,Vigilante
from .models import Autorizado,Vehiculo,Interior,Apartamento,UserPerfil,Parametros,TipoAutoriza,PlataformaWebVehiculo,PlataformaWebPeatonal,TipoPlataformaWeb,AutorizacionVehiculo,ReunionConsejo,DecisionConsejo,CompromisoConsejo
from .models import IngresoVehiculo, VisitanteVehiculo,Mantenimiento,ActivoFijo,Mantenimiento,Proveedor,ServicioProveedor,Obra,Reparacion,AvanceObra,ReunionConsejo,DecisionConsejo,TipoActivo,Asamblea
from .models import MiembroStaff,InformeRevisor,RecomendacionRevisor,ProcesoJuridico,GestionProcesoJuridico,TipoProceso,TipoIngreso,ResidenteTemp,Conjunto,Proponente,Proyecto,TipoProyecto,TipoContrato,PrestamoActivoFijo
from .models import DecisionAsamblea,TipoAsamblea,MiembroComiteConvivencia,BajaActivoFijo
from upload.models import AnexoProveedor,AnexoProponente,AnexoProponenteProyecto,AnexoProyecto,AnexoAsamblea,AnexoBajaActivoFijo
from upload.models import AnexoContrato,AnexoMantenimiento,AnexoObra,AnexoAvanceObra,AnexoReparacion,AnexoReunionConsejo,AnexoInformeRevisor,AnexoProcesoJuridico,AnexoGestionProcesoJuridico

from .forms import VisitanteForm,DatosCorrespondenciaForm,DatosIngresoPeatonalForm,ResidenteForm,CreaCorrespondenciaForm,UserPerfilForm,MiembroStaffForm,TipoContratoForm,VigilanteForm,GestionProcesoJuridicoForm
from .forms import AutorizadoForm,AutorizadoPlataformaWebPeatonalForm,AutorizadoPlataformaWebVehiculoForm,DatosIngresoVehicularForm,VisitanteVehiculoForm,ResidenteLoginForm,VehiculoForm,MantenimientoForm,ActivoFijoForm
from .forms import ContratoForm,ProveedorForm,ServicioProveedorForm,ObrasForm,ReparacionesForm,AvanceObrasForm,ReunionConsejoForm,DecisionConsejoForm,CompromisoConsejoForm,MiembroConsejoForm,CargaFotoMiembroConsejoForm
from .forms import CargaFotoActivoFijoForm,CargaFotoMiembroStaffForm,InformeRevisorForm,RecomendacionRevisorForm,ProcesoJuridicoForm,TipoProcesoForm,ParqueaderoForm,MascotaForm,TipoActivoFijoForm,PrestamoActivoForm
from .forms import ZonaComunForm,CargaFotoZonaComunForm,SolicitudTokenForm,ResidenteTempForm,DatosReservaZonasComunesForm,ProponenteForm,ProyectosForm,ProponenteProyectoForm,TipoProyectoForm,AutorizadoVehicularForm
from .forms import AsambleaForm,DecisionAsambleaForm,TipoAsambleaForm,ContactForm,ComiteConvivenciaForm,CargaFotoMiembroComiteForm,BajaActivoForm,EmailForm,EmailAnexoForm

from registration.forms import SignupForm

from .filters import CorrespondenciaFilter,ProyectoFilter,ProponenteFilter,ObraFilter,PrestamosActivosFijosFilter,IngresoPeatonalFilter,IngresoVehiculoFilter,ProcesosJuridicosFilter,ReunionesConsejoFilter
from .filters import ContratosFilter,ResidentesFilter,ProveedorFilter,CorrespondenciaFilter,MantenimientosFilter,ReparacionesFilter,ActivosFijosFilter,InformesRevisorFilter,AutorizadoFilter,AutorizacionVehiculoFilter
from .filters import AsambleaFilter,BajaActivosFijosFilter,ResidentesTokenFilter
from .tables import CorrespondenciaTable,ZonasComunesTable,ResidenteTable,ResidenteTable1,MantenimientosTable,ContratosTable,ProveedoresTable,ResidenteTempTable,TokensResidentesTable
from .tables import ResidenteOtroTable,UserPerfilTable,AutorizacionesPeatonalTable,AutorizacionesPlataformaWebPeatonalTable,AutorizacionesPlataformaWebVehiculoTable,ActivoFijo1Table,ActivoFijo2Table,ActivoFijoTable
from .tables import Proveedores1Table,Proveedores2Table,PrestamosActivosFijosTable,BajaActivosFijosTable,CorreosResidentesTable
from .tables import ObrasTable,ReparacionesTable,MiembrosConsejoTable,MiembrosStaffTable,ReunionConsejoTable,InformeRevisorTable,ProcesosJuridicosTable,ApartamentosTable,ComiteConvivenciaTable
from .tables import Reservas,ProponentesTable,ProyectosTable,ProyectosDetalleTable,ProponentesProyectoTable,VigilantesTable,AutorizacionVehiculoTable,IngresoPeatonalTable,IngresoVehicularTable

from pages.models import Panel,Comunicado

import django_tables2
from django_tables2.config import RequestConfig
from django_tables2 import SingleTableMixin
from django_filters.views import FilterView
from django_tables2 import SingleTableView
from django_tables2.export.export import TableExport

from django_filters.views import FilterView
from django_tables2.views import SingleTableMixin

from django.conf import settings

from cv2 import *
#from SimpleCV import Camera, Display, Image
from time import sleep

import cv2

import pygame #importa pygame
import pygame.camera #importa camera desde video capture
from pygame.locals import* #importa comandos locales

from django.core import serializers
from django.core.serializers import serialize
import json
from openpyxl import Workbook
import openpyxl


import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

from django.utils.crypto import get_random_string
import hashlib
import random


def formatNumber(number, decimals, espanol=True):
    if type(number) != int and type(number) != float:
        return number
 
    d={'.':',', ',':'.'}
    return ''.join(d.get(s, s) for s in f"{number:,.{decimals}f}") \
        if espanol \
        else f"{number:,.{decimals}f}"

############################# Home ###################################

class BaseHomePageView(TemplateView):
    template_name = "core/base_home.html"

def UserHomeView(request,tipo_usuario):
    return render(request, "core/base_home.html", {'tipo_usuario': tipo_usuario})

class HomePageView(TemplateView):
    template_name = "core/base_home.html"

################################### Registro ##############################

def login_request(request):
    # Creamos el formulario de autenticación vacío
    form = AuthenticationForm()
    if request.method == "POST":
        # Añadimos los datos recibidos al formulario
        form = AuthenticationForm(data=request.POST)
        # Si el formulario es válido...
        if form.is_valid():
            # Recuperamos las credenciales validadas
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # Verificamos las credenciales del usuario
            user = authenticate(username=username, password=password)
            tipo_usuario = 0
            # Si existe un usuario con ese nombre y contraseña
            if user is not None:
                # Hacemos el login manualmente
                do_login(request, user)
                # Y le redireccionamos a la portada
                if request.user.is_superuser :
                    messages.success(request, "Ud. Ha ingresado exitosamente como administrador!")
                    request.session['ingreso_administrador'] = 1
                    request.session['ingreso_revisor'] = 0
                    request.session['ingreso_consejo'] = 0
                    request.session['ingreso_vigilante'] = 0
                    request.session['ingreso_residente'] = 0
                    tipo_usuario = 1
                    request.session['tipo_usuario'] = tipo_usuario
                    return redirect('pagina_administrador')
                else:
                    if MiembroConsejo.objects.filter(email=request.user.email).exists():
                        messages.success(request, "Ud. Ha ingresado exitosamente como miembro del consejo!")
                        request.session['ingreso_consejo'] = 1
                        request.session['ingreso_administrador'] = 0
                        request.session['ingreso_revisor'] = 0
                        request.session['ingreso_vigilante'] = 0
                        request.session['ingreso_residente'] = 0
                        tipo_usuario = 2
                        request.session['tipo_usuario'] = tipo_usuario
                        return redirect('user_home',tipo_usuario)
                        
                    else:
                        if MiembroStaff.objects.filter(email=request.user.email).exists():
                            messages.success(request, "Ud. Ha ingresado exitosamente como revisor!")
                            request.session['ingreso_revisor'] = 1
                            request.session['ingreso_administrador'] = 0
                            request.sessions['ingreso_consejo'] = 0
                            request.session['ingreso_vigilante'] = 0
                            request.session['ingreso_residente'] = 1
                            tipo_usuario = 3
                            request.session['tipo_usuario'] = tipo_usuario
                            return redirect('pagina_revisor')
                            
                        else:
                            if Vigilante.objects.filter(email=request.user.email).exists():
                                messages.success(request, "Ud. Ha ingresado exitosamente como vigilante!")
                                request.session['ingreso_vigilante'] = 1
                                request.session['ingreso_revisor'] = 0
                                request.session['ingreso_administrador'] = 0
                                request.session['ingreso_consejo'] = 0
                                request.session['ingreso_residente'] = 1
                                tipo_usuario = 4
                                request.session['tipo_usuario'] = tipo_usuario
                                return redirect('pagina_vigilancia')

                            else:
                               if Residente.objects.filter(email=request.user.email).exists():
                                    messages.success(request, "Ud. Ha ingresado exitosamente como vigilante!")
                                    request.session['ingreso_vigilante'] = 0
                                    request.session['ingreso_revisor'] = 0
                                    request.session['ingreso_administrador'] = 0
                                    request.session['ingreso_consejo'] = 0
                                    request.session['ingreso_residente'] = 1
                                    tipo_usuario = 5
                                    request.session['tipo_usuario'] = tipo_usuario
                                    return redirect('user_home',tipo_usuario)
                   
        else:
            mensaje = 'Usuario No Registrado'
            mensaje1 ='Debe entrar por la opción <¿No tiene Clave? e iniciar todo el proceso de registro.'
            mensaje2 = ''
            mensaje3 = ''
            tipo_usuario = 0
            return render(request, "core/mensaje_error_login.html", {'mensaje': mensaje,'mensaje1': mensaje1,'mensaje2': mensaje2,'mensaje3': mensaje3,'tipo_usuario':tipo_usuario})
    # Si llegamos al final renderizamos el formulario
    
    return render(request, "users/login.html", {'form': form})

def logout(request):
    # Finalizamos la sesión
    do_logout(request)
    messages.success(request, "Sesion cerrada exitosamente!")
    # Redireccionamos a la portada
    tipo_usuario = 0
    return redirect('user_home',tipo_usuario)

from .forms import UserCreationFormWithEmail

# Create your views here.

class SignUpResidenteView(CreateView):
    form_class = UserCreationFormWithEmail
    template_name = 'registration/signup.html'

    def get_success_url(self):
        return reverse_lazy('user_home',args=[5])   

    def get_form(self, form_class=None):
        form = super(SignUpResidenteView, self).get_form()
        # Modificar en tiempo real
        form.fields['username'].widget = forms.TextInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Nombre de usuario'})
        form.fields['email'].widget = forms.EmailInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Dirección email'})
        form.fields['password1'].widget = forms.PasswordInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Contraseña'})
        form.fields['password2'].widget = forms.PasswordInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Repite la contraseña'})
        return form
    
    def form_valid(self, form):
        if form.is_valid():
            user = form.save(commit=False)
            if MiembroConsejo.objects.filter(email=email).exists():
                User.is_staff=True
            else:
                User.is_staff=False
            user.save()
        tipo_usuario=request.session['tipo_Usuario']
        return redirect('user_home',tipo_usuario)    
        #return redirect('home')

class SignUpVigilanteView(CreateView):
    form_class = UserCreationFormWithEmail
    template_name = 'registration/signup.html'

    def get_success_url(self):
        return reverse_lazy('user_home',args=[4])   

    def get_form(self, form_class=None):
        form = super(SignUpVigilanteView, self).get_form()
        # Modificar en tiempo real
        form.fields['username'].widget = forms.TextInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Nombre de usuario'})
        form.fields['email'].widget = forms.EmailInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Dirección email'})
        form.fields['password1'].widget = forms.PasswordInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Contraseña'})
        form.fields['password2'].widget = forms.PasswordInput(
            attrs={'class':'form-control mb-2', 'placeholder':'Repite la contraseña'})
        return form
    
    def form_valid(self, form):
        if form.is_valid():
            email = form.cleaned_data['email']
            if Vigilante.objects.filter(email=email).exists():
                user = form.save(commit=False)
                user.save()
            else:
                if MiembroStaff.objects.filter(email=email).exists():
                    user = form.save(commit=False)
                    user.save()
                    messages.warning(self.request, 'No está registrado como vigilante o staff del conjunto')    
        tipo_usuario=request.session['tipo_Usuario']
        return redirect('user_home',tipo_usuario)    
        #return redirect('home')

def RegistroResidenteView(request):
    instance=Residente()
    if request.method=='POST':
        residente_form = ResidenteLoginForm(request.POST,instance)
        if residente_form.is_valid():
            apartamento_sol = residente_form.cleaned_data['apartamento']
            interior_sol = residente_form.cleaned_data['interior']
            email_sol = residente_form.cleaned_data['email']
            token_sol = residente_form.cleaned_data['token']
            if Residente.objects.filter(interior_id = interior_sol,apartamento_id = apartamento_sol, email = email_sol,token=token_sol).exists():
                residente = Residente.objects.get(interior_id = interior_sol,apartamento_id = apartamento_sol, email = email_sol,token=token_sol)
                sapartamento = residente.apartamento.numero
                nombre_res = residente.nombre
                sinterior = residente.interior.numero
                if User.objects.filter(email=email_sol).exists():
                    residente = Residente.objects.get(email=email_sol)
                    url = 'user_home'
                    mensaje = 'Ya existe un usuario registrado con éste mail: '+email_sol
                    mensaje1 ='        Interior: '+sinterior+' Apartamento: '+sapartamento+'.'
                    mensaje2 = 'Si la información anterior de Interior y Apartamento coincide con la suya, significa que ya se ha logeado en el sistema.'
                    mensaje3 = 'Ingrese con el usuario y clave con que se registró inicialmente, si olvidó la clave pida cambio de clave en <Login> por la opción <Olvidó su clave?>'
                    return render(request, "core/mensaje_error_login.html", {'emali':email_sol,'mensaje': mensaje,'mensaje1': mensaje1,'mensaje2': mensaje2,'mensaje3': mensaje3})
                else:
                    return redirect('sign_up',residente.id)
            else:
                url = 'crear_residente_uno'
                parametro1 = interior_sol
                parametro2 = apartamento_sol
                parametro3 = email_sol
                mensaje = 'Residente con ésta información no existe en la base de datos, debe enviar una solicitud a la administración por <Login> opcion <Actualizar Datos> para solicitarle la inclusión de sus datos'
                return render(request, "core/mensaje_error_tres_parametros.html", {'mensaje': mensaje,'url':url,'parametro1':parametro1,'parametro2':parametro2,'parametro3':parametro3})
    else:
        form = ResidenteLoginForm(request.POST,instance)
        return render(request,'core/registro_residente.html',{'form':form})        

@transaction.atomic
def update_profile(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, _('Your profile was successfully updated!'))
            return redirect('settings:profile')
        else:
            messages.error(request, _('Please correct the error below.'))
    else:
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=request.user.profile)
    return render(request, 'profiles/profile.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

def CogeVarIdView(request):
    id = request.GET.get('id', None)
    parametros = Parametros()
    parametros.valor_parametro_uno = id
    parametros.user = request.user.id
    parametros.save()
    data = {'estado':True}    
    return JsonResponse(data)

#################################### Ingreso Peatonal #############################

def IngresoPeatonalBusquedaView(request):
    mensaje1=''
    request.session['tipoautorizacion'] = 1
    if request.method == 'GET':
        query= request.GET.get('q')
        identificacion=query
        submitbutton= request.GET.get('submit')
        request.session['identificacion']=identificacion
        tipo_autorizacion = 0
        desc_autoriza='No tiene'
        fecha_inicial=None
        fecha_final=None
        mensaje=''
        nom_interior = ''
        nom_apartamento = ''
        tipoingreso = TipoIngreso.objects.all()
        apartamentos = Apartamento.objects.all()
        interior = Interior.objects.all()
        tipo_ingreso = 0
        if Autorizado.objects.filter(identificacion=identificacion).exists():
            results= Autorizado.objects.filter(identificacion=identificacion)
            autorizado = Autorizado.objects.get(identificacion=identificacion)
            fecha_final = autorizado.fecha_final
            tipo_autorizacion = autorizado.tipo_autoriza_id
            tipoautoriza = TipoAutoriza.objects.get(id=tipo_autorizacion)
            desc_autoriza = tipoautoriza.descripcion
            idinterior = autorizado.interior_id
            idapartamento = autorizado.apartamento_id
            permanente = autorizado.permanente
            activo = autorizado.activo
            interior = Interior.objects.get(id=idinterior)
            apartamento = Apartamento.objects.get(id=idapartamento)
            nom_interior = interior.numero
            nom_apartamento = apartamento.numero
            tipo_autorizacion = 1
            request.session['tipoautorizacion'] = 2
            request.session['interior'] = interior.numero
            request.session['apartamento'] = apartamento.numero
            if permanente == True:
                mensaje = 'AUTORIZADO! HACER INGRESO : '+autorizado.apartamento.numero  
            elif fecha_final < date.today():
                mensaje = 'AUTORIZADO! HACER INGRESO : '+autorizado.apartamento.numero  
            else:
                mensaje = 'AUTORIZACION HA VENCIDO'  

            if Visitante.objects.filter(identificacion=identificacion).exists():
                visitante=Visitante.objects.get(identificacion=identificacion)
                ingresos = IngresoPeatonal.objects.filter(identificacion=identificacion).order_by('-id')[:5]
            else:
                autorizado= Autorizado.objects.get(identificacion=identificacion)
                a = Visitante( identificacion = identificacion, nombre = autorizado.nombre)
                a.save()    
                ingresos = None
        elif PlataformaWebPeatonal.objects.filter(identificacion=identificacion).exists():
            request.session['tipoautorizacion'] = 3
            results= PlataformaWebPeatonal.objects.filter(identificacion=identificacion)
            ABautorizado = PlataformaWebPeatonal.objects.get(identificacion=identificacion)
            fecha_inicial = ABautorizado.fecha_inicial
            fecha_final = ABautorizado.fecha_final
            idinterior = ABautorizado.interior_id
            idapartamento = ABautorizado.apartamento_id
            interior = Interior.objects.get(id=idinterior)
            apartamento = Apartamento.objects.get(id=idapartamento)
            request.session['interior'] = interior
            request.session['apartamento'] = apartamento
            nom_interior = interior.numero
            nom_apartamento = apartamento.numero
            tipo_autorizacion = 3
            if fecha_final < date.today():
                mensaje = 'AUTORIZACION HA VENCIDO'
            else:
                mensaje = 'AUTORIZADO PLATAFORMA WEB REVISAR DOCUMENTACION Y CREAR INGRESO'    
            desc_autoriza = 'Autorizacion Plataforma Web Peatonal'
            
            if Visitante.objects.filter(identificacion=identificacion).exists():
                request.session['tipoautorizacion'] = 1
                visitante=Visitante.objects.get(identificacion=identificacion)
                ingresos = IngresoPeatonal.objects.filter(identificacion=identificacion).order_by('-id')[:5]
            else:    
                ingresos = None
        elif Visitante.objects.filter(identificacion=identificacion).exists():
            visitante = Visitante.objects.get(identificacion=identificacion)
            ingresos =IngresoPeatonal.objects.filter(identificacion=identificacion).order_by('-id')[:5]
            results= Visitante.objects.filter(identificacion=identificacion).distinct()
            
            if IngresoPeatonal.objects.filter(identificacion=identificacion):
                ultimo_ingreso = IngresoPeatonal.objects.filter(identificacion=identificacion).latest('id')
                request.session['interior'] = ultimo_ingreso.interior.numero
                request.session['apartamento'] = ultimo_ingreso.apartamento.numero
                mensaje = 'VISITANTE FRECUENTE, SOLICITAR AUTORIZACION AL RESIDENTE Y CREAR INGRESO'
            else:
                mensaje = 'VISITANTE PRIMERA VEZ, SOLICITAR AUTORIZACION AL RESIDENTE Y CREAR INGRESO'    
            tipo_autorizacion = 1
            request.session['tipoautorizacion'] = 1
            
        else:
            visitantes = None
            mensaje1 ='VISITANTE NUEVO, INGRESAR DATOS' 
            ingresos = IngresoPeatonal.objects.all().order_by('-id')[:5]
            results=''
            results= Visitante.objects.filter(identificacion=identificacion).distinct()
            request.session['tipoautorizacion'] = 1
        return render(request,'core/ingreso_peatonal.html', {'results':results,'submitbutton': submitbutton,'ingresos':ingresos,
        'query':query,'tipo_autorizacion':tipo_autorizacion,'fecha_inicial':fecha_inicial,'fecha_final':fecha_final,'identificacion':id,
        'desc_autoriza':desc_autoriza,'mensaje':mensaje,'nom_interior':nom_interior,'nom_apartamento':nom_apartamento,'tipo_autirizacion':tipo_autorizacion,'mensaje1':mensaje1,'tipoingreso':tipoingreso,'apartamentos':apartamentos,'interior':interior})
    else:
        #ingresos =IngresoPeatonal.objects.all().order_by('-id')[:5]
        ingresos = None
        query=''
        results=''
        submitbutton=''
        mensaje1 ='VISITANTE NO ENCONTRADO, INGRESAR DATOS'
        return render(request, 'core/ingreso_peatonal.html',{'results':results,'submitbutton': submitbutton,'ingresos':ingresos,'identificacion:':id,'mensaje1':mensaje1,'tipoingreso':tipoingreso,'apartamentos':apartamentos,'interior':interior})

class CreaIngresoPeatonalView(LoginRequiredMixin,CreateView):
    model = IngresoPeatonal
    template_name = 'core/ingreso_peatonal_form.html'
    form_class = DatosIngresoPeatonalForm
    success_url = reverse_lazy('acceso_peatonal') 

    def form_valid(self, form):
        if form.is_valid():
            self.object = form.save(commit=False)
            identificacion = self.request.session['identificacion']
            idinterior = form.cleaned_data['interior']
            interior = Interior.objects.get(numero=idinterior)
            idapartamento = form.cleaned_data['apartamento']
            apartamento = Apartamento.objects.get(numero=idapartamento)
            idtipoingreso = form.cleaned_data['tipoingreso']
            tipoingreso = TipoIngreso.objects.get(descripcion=idtipoingreso)
            if Visitante.objects.filter(identificacion=identificacion).exists():
                visitante = Visitante.objects.get(identificacion=identificacion)
                self.object.nombre = visitante.nombre
            self.object.interior_id = interior.id
            self.object.apartamento_id = apartamento.id
            self.object.identificacion = identificacion
            self.object.vigilante_id = self.request.user.id
            self.object.tipo_autoriza_id = self.request.session['tipoautorizacion']
            self.object.tipoingreso_id = tipoingreso.id
            self.object.save()
            visitante = Visitante.objects.get(identificacion=identificacion)
            apartamento = Apartamento.objects.get(id=apartamento.id)
            mensaje =" El Sr./Sra. "+visitante.nombre+" solicita ingreso a su apartamento <"+apartamento.numero+">"
            asunto = "Solicitud Ingreso Peatonal "
            email = ''
            EnvioMailResidente(apartamento.id,mensaje,asunto,email)
            return redirect('acceso_peatonal')
    

@csrf_exempt
def GuardaIdentificacionVisitanteView(request):
    identifica = request.GET.get('identifica', None)
    request.session['identifica'] = identifica
    id = 0
    data = {'id':id}
    return JsonResponse(data) 

@csrf_exempt
def GuardaDatosIngresoPeatonalView(request):
    identificacion = request.GET.get('identificacion', None)
    nombre = request.GET.get('nombre', None)
    tipoingreso = request.GET.get('tipoingreso', None)
    interior = request.GET.get('interior', None)
    apartamento = request.GET.get('apartamento', None)
    tipoautorizacion = request.session['tipoautorizacion']
    ingreso_peatonal = IngresoPeatonal()
    ingreso_peatonal.identificacion = identificacion
    ingreso_peatonal.tipoingreso_id = tipoingreso
    ingreso_peatonal.interior_id = interior
    ingreso_peatonal.apartamento_id = apartamento
    ingreso_peatonal.vigilante_id=request.user.id
    ingreso_peatonal.tipo_autoriza = tipoautorizacion 
    ingreso_peatonal.hora_salida=None
    ingreso_peatonal.save()
    data = {'id':id}
    return JsonResponse(data) 

@csrf_exempt
def GuardaDatosIngresoPeatonaAutorizadoView(request):
    id = request.GET.get('id', None)
    tipo_autorizacion = request.GET.get('tipo_autorizacion', None)
    tipo_autorizacion = int(tipo_autorizacion)
    if tipo_autorizacion == 1:
        autorizado = Autorizado.objects.get(identificacion=id)
        interior = autorizado.interior_id 
        apartamento = autorizado.apartamento_id
        nombre = autorizado.nombre
    
    if tipo_autorizacion == 2:
        autorizado = PlataformaWebPeatonal.objects.get(identificacion=id)
        interior = autorizado.interior_id 
        apartamento = autorizado.apartamento_id
        nombre = autorizado.nombre

    if Visitante.objects.filter(identificacion=id).exists():
        visitante = Visitante.objects.get(identificacion=id)
        idvisitante =visitante.id
    else:
        visitante = Visitante()           
        visitante.identificacion = id
        visitante.nombre = nombre
        visitante.save()
        visitante = Visitante.objects.get(identificacion=id)
        idvisitante =visitante.id 

    ingreso_peatonal = IngresoPeatonal()
    ingreso_peatonal.identificacion_id = idvisitante
    ingreso_peatonal.tipoingreso_id = 4
    ingreso_peatonal.interior_id = interior
    ingreso_peatonal.apartamento_id = apartamento
    ingreso_peatonal.vigilante_id=request.user.id
    ingreso_peatonal.hora_salida=None
    ingreso_peatonal.save()
    data = {'id':id}
    return JsonResponse(data) 

def SalidaPeatonalView(request):
    id = request.GET.get('id', None)
    IngresoPeatonal.objects.filter(id=id).update(hora_salida=datetime.now())
    fecha = datetime.today().strftime('%A, %B %d, %Y %H:%M:%S')
    data = {'id':id,'fecha':fecha}
    return JsonResponse(data) 

def DireccionaSalidaPaginaAccesosView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
       return redirect('pagina_administrador')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5:
            return redirect('user_home',tipo_usuario)    
        else:
            return redirect('user_home',tipo_usuario)    
        
def DatosVisitanteView(request):
    form =  VisitanteForm()
    context = {'form': form}
    html_form = render_to_string('core/crear_visitante.html',
    context,
    request=request,
    )
    return JsonResponse({'html_form': html_form})

@csrf_exempt
def GuardaDatosVisitanteView(request):
    id = request.GET.get('id', None)
    nombre = request.GET.get('nombre', None)
    visitante = Visitante()
    visitante.nombre = nombre.title()
    visitante.identificacion = id
    visitante.save()
    request.session['identificacion'] = id
    data = {'id':id}
    return JsonResponse(data) 

class AccesoPeatonalView(TemplateView):
    template_name = 'core/ingreso_peatonal.html'

@login_required
def AccesoPeatonalListaView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
        queryset = IngresoPeatonal.objects.all().order_by('-id')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5:
            sw = 2
            email = request.user.email
            residente = Residente.objects.get(email=email)
            idapartamento = residente.apartamento_id
            queryset = IngresoPeatonal.objects.filter(apartamento_id=idapartamento).order_by('-id')
    f = IngresoPeatonalFilter(request.GET, queryset=queryset)
    ingresos = IngresoPeatonalTable(f.qs)
    lista = []
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 15, "page": 1}).configure(ingresos)
    return render(request, "core/ingreso_peatonal_lista.html", {"ingresos":ingresos,'filter': f,'tipo_usuario':tipo_usuario})
    

def ReporteIngresoPeatonalXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada'] 
    ingresos_imprimir = IngresoPeatonal.objects.filter(id__in=lista)
    worksheet.write('A1','Hora Ingreso' )
    worksheet.write('B1','Hora Salida' )
    worksheet.write('C1','Tipo Ingreso')
    worksheet.write('D1','Identificación')
    worksheet.write('E1','Nombre')
    worksheet.write('F1','Apartamento')
    worksheet.write('G1','Vigilante')
    n = 2
    for j in ingresos_imprimir :
        nn = str(n)
        if j.hora_ingreso != None:
            singreso= datetime.strftime(j.hora_ingreso,'%A, %B %d, %Y %H:%M:%S')
            exec("worksheet.write('A"+nn+"','"+singreso+"' )")
        if j.hora_salida != None:
            ssalida= datetime.strftime(j.hora_salida,'%A, %B %d, %Y %H:%M:%S')
            exec("worksheet.write('B"+nn+"','"+ssalida+"' )")
        exec("worksheet.write('C"+nn+"','"+str(j.tipo_autoriza.descripcion)+"' )")
        exec("worksheet.write('D"+nn+"','"+j.identificacion+"' )")
        exec("worksheet.write('E"+nn+"','"+j.nombre+"' )")
        exec("worksheet.write('F"+nn+"','"+j.apartamento.numero+"' )")
        exec("worksheet.write('G"+nn+"','"+str(j.vigilante.username)+"' )")
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='ingreso_peatonal.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

def DireccionaSalidaAutorizacionesPeatonalView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
       return redirect('pagina_autorizaciones')
    else:
        if tipo_usuario == 4:
            return redirect('pagina_vigilancia')
        else:
            if tipo_usuario == 2 or tipo_usuario == 5:
                return redirect('pagina_autorizaciones')    
            
def DireccionaSalidaPaginaAutorizacionesView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
       return redirect('pagina_administrador')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5:
            return redirect('user_home',tipo_usuario)    
        else:
            return redirect('user_home',tipo_usuario)    
    
################################## Paginas Generales ##########################

class PaginaZonasComunesView(TemplateView):
    template_name = 'core\pagina_zonas_comunes.html'

class PaginaAutorizacionesView(TemplateView):
    template_name = 'core\pagina_autorizaciones.html'

class PaginaAccesosView(TemplateView):
    template_name = 'core\pagina_accesos.html'

class InformacionGeneralView(TemplateView):
    template_name = 'core\pagina_informacion_general.html'

class MantenimientoView(TemplateView):
    template_name = 'core\pagina_mantenimiento.html'

class PaginaRevisorView(TemplateView):
    template_name = 'core\pagina_revisor.html'

class AsambleaView(TemplateView):
    template_name = 'core\pagina_asamblea.html'

class PaginaCorrespondenciaView(TemplateView):
    template_name = 'core\pagina_correspondencia.html'

class PqrView(TemplateView):
    template_name = 'core\pagina_pqr.html'

class JuridicoView(TemplateView):
    template_name = 'core\juridico.html'

class SistemaView(TemplateView):
    template_name = 'core\pagina_sistema.html'

def MainView(request):
    request.session['idzona'] = ''
    request.session['fecha'] = ''
    request.session['UltimaPaginaVisitada']=''
    model = Comunicado
    comunicado = Comunicado.objects.all()
    return render(request, "core/main.html", {'comunicado':comunicado})
    
def MensajePaginaEnConstruccionView(request,id):
    mensaje = "Página en Construccion"
    parametro = id
    return render(request, "core/mensaje_pagina_en_construccion.html", {'mensaje': mensaje,'parametro':parametro})

class LegislacionView(TemplateView):
    template_name = 'core\legislacion.html'        

class AboutView(TemplateView):
    template_name = 'core\empresa.html'

class PaginaAdministradorView(TemplateView):
    template_name = 'core\pagina_administrador.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.request.session['UltimaPaginaVisitada']='administrador'
        return context

class VigilanciaView(TemplateView):
    template_name = 'core\pagina_vigilancia.html'

class PaginaAlertasView(TemplateView):
    template_name = 'core\pagina_alertas.html'

##################################### Alertas ################################

def AlertaReciboEnergiaView(request):
    asunto = 'El recibo de energía ha llegado'
    mensaje = " ha llegado el recibo de energía, favor pasar a recogerlo a la portería"
    EnvioMailResidentes(mensaje,asunto)
    return redirect('pagina_alertas')

def AlertaReciboAguaView(request):
    asunto = 'El recibo del agua ha llegado'
    mensaje = " ha llegado el recibo del agua, favor pasar a recogerlo a la portería"
    EnvioMailResidentes(mensaje,asunto)
    return redirect('pagina_alertas')

def AlertaReciboGasView(request):
    asunto = 'El recibo de gas ha llegado'
    mensaje = " ha llegado el recibo de gas, favor pasar a recogerlo a la portería"
    EnvioMailResidentes(mensaje,asunto)
    return redirect('pagina_alertas')    

###################################### Correspondencia ############################

class IngresoCorrespondenciaView(TemplateView):
    template_name = 'core\ingreso_correspondencia.html' 

class MensajeCorrespondenciaView(TemplateView):
    template_name = 'core\mensaje_correspondencia.html'    

class CalendarioView(TemplateView):
    template_name = 'core\calendario.html'    

class PaginaInformacionResidentesView(TemplateView):
    template_name = 'core\pagina_residentes.html'

class CreaCorrespondenciaView(LoginRequiredMixin,CreateView):
    model = Correspondencia
    template_name = 'core/correspondencia_form.html'
    form_class = CreaCorrespondenciaForm
    success_url=reverse_lazy('lista_correspondencia')

    def form_valid(self, form):
        if form.is_valid():
            correspondencia = form.save(commit=False)
            clase_correspondencia = form.cleaned_data['clase_correspondencia']
            correspondencia.clase_correspondencia = clase_correspondencia
            tipo_correspondencia = form.cleaned_data['tipo_correspondencia']
            correspondencia.tipo_correspondencia = tipo_correspondencia
            sinterior = form.cleaned_data['interior']
            interior = Interior.objects.get(numero=sinterior)
            sapartamento = form.cleaned_data['apartamento']
            apartamento = Apartamento.objects.get(numero=sapartamento)
            correspondencia.interior_id = interior.id
            correspondencia.apartamento_id = apartamento.id
            correspondencia.vigilante_id = self.request.user.id
            correspondencia.save()
            #ultima = Correspondencia.objects.latest('id')
            #residente = Residente.objects.filter(interior_id=interior.id,apartamento_id=apartamento.id).exclude(email=None)
            apartamento = Apartamento.objects.get(id=apartamento.id)
            mensaje =" Le ha llegado un "+correspondencia.clase_correspondencia+" de carácter :"+correspondencia.tipo_correspondencia+" a su apartamento <"+apartamento.numero+">"
            asunto = "Correspondencia recibida "
            email = ''
            EnvioMailResidente(apartamento.id,mensaje,asunto,email)
        return redirect('lista_correspondencia')


def DireccionaCorrespondenciaView(request):
    tipo_usuario = request.session['tipo_usuario'] 
    if tipo_usuario == 4 :
        return redirect('pagina_vigilancia')
    else:
        if tipo_usuario == 1 :
            return redirect('pagina_administrador')        
            #messages.info(request, f"Debe registrarse como Vigilante")
        else:
            if tipo_usuario == 2 or tipo_usuario == 5 :
               return redirect('user_home',tipo_usuario)             
               

def CorrespondenciaListView(request):
    tipo_usuario = request.session['tipo_usuario']   
    if tipo_usuario == 1 or  tipo_usuario == 4 :
        queryset = Correspondencia.objects.all().order_by('-id')
    else:
        if tipo_usuario == 2 or tipo_usuario== 5 :
            residente = Residente.objects.get(email=request.user.email)
            idapartamento = residente.apartamento_id
            queryset = Correspondencia.objects.filter(apartamento_id=idapartamento).order_by('-id')
             
    f = CorrespondenciaFilter(request.GET, queryset=queryset)
    correspondencia = CorrespondenciaTable(f.qs)
    lista = []
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 15, "page": 1}).configure(correspondencia)
    return render(request, "core/correspondencia_lista.html", {'correspondencia':correspondencia,'tipo_usuario':tipo_usuario,'filter':f})
    
def ReporteCorrespondenciaXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada'] 
    correspondencia = Correspondencia.objects.filter(id__in=lista)
    worksheet.write('A1','Clase' )
    worksheet.write('B1','Tipo' )
    worksheet.write('C1','Detalle')
    worksheet.write('D1','Fecha Recibo')
    worksheet.write('E1','Fecha Entregado')
    worksheet.write('F1','Apartamento')
    worksheet.write('G1','Entregado')
    worksheet.write('H1','Vigilante')
    n = 2
    for j in correspondencia :
        nn = str(n)
        exec("worksheet.write('A"+nn+"','"+str(j.clase_correspondencia)+"' )")
        exec("worksheet.write('B"+nn+"','"+str(j.tipo_correspondencia)+"' )")
        exec("worksheet.write('C"+nn+"','"+str(j.detalle)+"' )")
        if j.fechahora_recibo != None:
            srecibo= datetime.strftime(j.fechahora_recibo,'%A, %B %d, %Y %H:%M:%S')
            exec("worksheet.write('D"+nn+"','"+srecibo+"' )")
        if j.fechahora_entrega != None:
            sentrega= datetime.strftime(j.fechahora_entrega,'%A, %B %d, %Y %H:%M:%S')
            exec("worksheet.write('E"+nn+"','"+sentrega+"' )")
        if j.entregado == 'True':
            sentregado ='Si'
        else:
            sentregado ='No'        
        exec("worksheet.write('F"+nn+"','"+j.apartamento.numero+"' )")
        exec("worksheet.write('G"+nn+"','"+sentregado+"' )")
        exec("worksheet.write('H"+nn+"','"+str(j.vigilante.username)+"' )")
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='Correspondencia.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

def DatosCorrespondenciaView(request):
    data = dict()
    form = DatosCorrespondenciaForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_correspondencia.html',
        context,
        request=request
    )
    return JsonResponse(data)


def EntregasCorrespondenciaView(request):
    tipo_usuario = request.session['tipo_usuario']
    id = request.GET.get('id', None)
    id=int(id)
    if tipo_usuario == 4:
        id = request.GET.get('id', None)
        id=int(id)
        data= dict()
        correspondencia = Correspondencia.objects.get(id=id)
        if correspondencia.entregado == False :
            Correspondencia.objects.filter(id=id).update(entregado=True,fechahora_entrega=datetime.now())
            correspondencia = Correspondencia.objects.get(id=id)
            #residente = Residente.objects.filter(interior_id=correspondencia.interior_id,apartamento_id=correspondencia.apartamento_id)
            #for i in residente:
            #if i.email != 'None':
            #EnvioMailCorrespondencia(i.nombre,i.email,i.apartamento.numero,correspondencia.tipo_correspondencia,correspondencia.clase_correspondencia,id,2)
            apartamento = Apartamento.objects.get(id=correspondencia.apartamento_id)
            mensaje =" Le ha sido entregado un "+correspondencia.clase_correspondencia+" de carácter :"+correspondencia.tipo_correspondencia+" a su apartamento <"+apartamento.numero+">"
            asunto = "Correspondencia entregada "
            email = ''
            EnvioMailResidente(apartamento.id,mensaje,asunto,email)
            data = {'id':id,'entregado':False}
            #else:
            #data = {'id':id,'entregado':False,'fechahora_entrega':datetime.now()}
        else:
           data = {'id':id,'entregado':True}     

    return JsonResponse(data)

@csrf_exempt
def GuardaDatosCorrespondenciaView(request):
    data = dict()
    clase_correspondencia = request.GET.get('clase_correspondencia', None)
    tipo_correspondencia = request.GET.get('tipo_correspondencia', None)
    interior = request.GET.get('interior', None)
    apartamento = request.GET.get('apartamento', None)
    correspondencia = Correspondencia()
    correspondencia.clase_correspondencia = clase_correspondencia
    correspondencia.tipo_correspondencia = tipo_correspondencia
    correspondencia.apartamento_id = apartamento
    correspondencia.interior_id = interior 
    correspondencia.detalle = ''
    correspondencia.destinatario =''
    correspondencia.remitente = ''
    correspondencia.entregado = False
    correspondencia.vigilante_id=request.user.id
    correspondencia.save()
    data = {'estado':True} 
    return JsonResponse(data)

def CorrespondenciaConsultaView(request):
    f = CorrespondenciaFilter(request.GET, queryset=Correspondencia.objects.all())
    return render(request, 'core/consulta_correspondencia.html', {'filter': f})    

def DireccionaCorrespondenciaView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 4:
        return redirect('pagina_vigilancia')
    else:
        if tipo_usuario == 1:
            return redirect('pagina_administrador')
        if tipo_usuario == 2 or tipo_usuario == 5:
           return redirect('user_home',tipo_usuario)         


################################# ZONAS COMUNES  #######################################

@login_required
def ZonasComunesListaView(request):
    zonas_comunes = ZonasComunesTable(ZonaComun.objects.all())
    #apartamentos = ApartamentosTable(Apartamento.objects.all())
    return render(request, 'core/zonas_comunes_lista.html',{'zonas_comunes':zonas_comunes})    

class CreaZonaComunView(LoginRequiredMixin,CreateView):
    model = ZonaComun
    template_name = 'core/zona_comun_form.html'
    form_class = ZonaComunForm
    success_url = reverse_lazy('lista_zonas_comunes')

class EditaZonaComunView(LoginRequiredMixin,UpdateView):
    model = ZonaComun
    fields = ['descripcion','observaciones','tarifa','arrienda','foto','orden']
    template_name = 'core/zona_comun_form.html'
    success_url = reverse_lazy('lista_zonas_comunes')

class BorraZonaComunView(LoginRequiredMixin,DeleteView):
    model = ZonaComun
    success_url = reverse_lazy('lista_zonas_comunes')
    template_name = 'core/confirmar_borrado_registro.html'

def CargaFotoZonaComunView(request,id):
    if request.method == 'POST':
        activo=ZonaComun.objects.get(id=id)
        form = CargaFotoZonaComunForm(request.POST, request.FILES,instance=activo)
        form.save()
        return redirect("lista_zonas_comunes")
    else:
        form = CargaFotoZonaComunForm()

    return render(request,'core/foto_zona_comun_form.html', locals())

""" def ajaxSwitchArriendaZonasComunesView(request):
    id = request.GET.get('id', None)
    miembro_staff = ZonaComun.objects.get(id=id)
    if miembro_staff.arrienda == True:
        envio = False
    else:
        envio = True
    zona_comun = ZonaComun.objects.filter(id=id).update(arrienda=arrienda)    
    data = {'envio':envio}
    return JsonResponse(data)  """

def ReservaZonasComunesView(request,id):
    idzona = id
    request.session['idzona'] = idzona
    tipo_usuario = request.session['tipo_usuario']
    zona = ZonaComun.objects.filter(id=id)
    if request.method == 'GET':
        query= request.GET.get('q')
        if query != None :
            if query == '':
                reservas = Reservas.objects.filter(zona_comun_id=id).filter(fecha__gte=datetime.today()).order_by('-fecha')
                return render(request, "core/reserva_zonas_comunes.html", {"zona":zona,"reservas":reservas,"idzona":idzona,'iduser':request.user.id,"tipo_usuario":tipo_usuario})
            else:
                #fecha_reserva= datetime.strptime( query, "%d/%m/%Y").date()
                fecha_reserva = query
            submitbutton= request.GET.get('submit')
            reservas = Reservas.objects.filter(zona_comun_id=id).filter(fecha__gte=datetime.today()).order_by('-fecha')
            return render(request, "core/reserva_zonas_comunes.html", {"zona":zona,"reservas":reservas,"idzona":idzona,'iduser':request.user.id,"tipo_usuario":tipo_usuario})
        else:
            if query != None :
                reservas = Reservas.objects.filter(zona_comun_id=id).filter(fecha__gte=datetime.today()).order_by('-fecha')
            else:
                reservas = Reservas.objects.filter(zona_comun_id=id).filter(fecha__gte=datetime.today()).order_by('-fecha')      
                return render(request, "core/reserva_zonas_comunes.html", {"zona":zona,"reservas":reservas,"idzona":idzona,'iduser':request.user.id,"tipo_usuario":tipo_usuario})
    else:
        if query != None :    
            reservas = Reservas.objects.filter(zona_comun_id=id,hora_inicio__gte=datetime.today()).order_by('-fecha')
        else:
            reservas = Reservas.objects.filter(zona_comun_id=id,fecha__gte=datetime.today()).order_by('-fecha')
        return render(request, "core/reserva_zonas_comunes.html", {"zona":zona,"reservas":reservas,"idzona":idzona,'iduser':request.user.id,"tipo_usuario":tipo_usuario})

def DireccionaReservaZonasComunesView(request):
    id = request.session['idzona']
    return redirect('reserva_zonas_comunes',id)

def FiltrarFechaReservaZonasComunes(request):
    fecha = request.GET.get('fecha', None)
    request.session['fecha'] = fecha
    data= dict()
    numres =  Reservas.objects.filter(fecha_inicial=fecha).count()
    data = {'estado':numres}
    return JsonResponse(data)

def PagoReservaZonasComunesView(request,id):
    idzona = request.session['idzona']
    request.session['idzonapagada'] = id
    zona = ZonaComun.objects.get(id=idzona)
    nombre_zona = zona.descripcion
    return render(request, 'core/boton_pagos_pse.html',{'nombre_zona':nombre_zona,'idzona':idzona})

def DireccionaPagoReservaZonasComunes(request):
    idzona = request.session['idzona']
    return redirect('reserva_zonas_comunes',idzona)

def ajaxFiltrarApartamentoView(request):
    id = request.GET.get('idinterior', None)
    apartamentos = Apartamento.objects.filter(interior_id=id).values('numero','id')
    return HttpResponse(json.dumps(list(apartamentos)), content_type='application/json')

class CrearReservaZonasComunesView(LoginRequiredMixin,CreateView):
    template_name = 'core/reserva_zonas_comunes_form.html'
    form_class = DatosReservaZonasComunesForm
    success_message = 'Success: Reserva fué creada'
    
    def form_valid(self, form):
        sw = 0
        idzona = self.request.session['idzona']
        reserva = form.save(commit=False)
        fecha_reserva = reserva.fecha
        hoy = datetime.today()
        hora_inicio = reserva.hora_inicio
        hora_final = reserva.hora_final
        if fecha_reserva != None:
            if  fecha_reserva < datetime.date(hoy):
                mensaje ='No puede reservar en fechas pasadas, revise la fecha a reservar'
                url = 'reserva_zonas_comunes'
                parametro = idzona
                return render(self.request, "core/mensaje_error.html", {'mensaje': mensaje,'idzona':idzona})
            else:
                reservas = Reservas.objects.filter(zona_comun_id=idzona,fecha=fecha_reserva)
                if reservas:
                    for i in reservas:
                        if  hora_inicio == i.hora_inicio:
                            mensaje ='Hora ya reservada'
                            url = 'reserva_zonas_comunes'
                            parametro = idzona
                            return render(self.request, "core/mensaje_error.html", {'mensaje': mensaje,'idzona':idzona})
                        else:
                            if hora_inicio > i.hora_inicio and hora_inicio < i.hora_final :
                                mensaje ='Hora ya reservada'
                                url = 'reserva_zonas_comunes'
                                parametro = idzona
                                return render(self.request, "core/mensaje_error.html", {'mensaje': mensaje,'idzona':idzona})
                            else:
                                reserva.user_id = self.request.user.id
                                reserva.zona_comun_id = idzona
                                reserva.save()
                                user_email = self.request.user.email
                                zona_comun = ZonaComun.objects.get(id=reserva.zona_comun_id)
                                residente = Residente.objects.get(email=user_email)
                                apartamento = Apartamento.objects.get(id=residente.apartamento_id)
                                idapartamento = apartamento.id
                                mensage =" Se ha creado una reserva temporal de la zona común "+zona_comun.descripcion+" a su nombre,el día "+fecha_reserva.strftime('%d/%m/%Y')+" Hora inicial: "+str(hora_inicio)+" hora Final: "+str(hora_final)+" no quedará en firma hasta hasta que efectúe el pago y lo confirme la Administración "
                                asunto = "Reserva Zona Común Temporal"
                                from_email = settings.EMAIL_HOST_USER
                                EnvioMailResidente(idapartamento,mensaje,asunto,residente.email)
                                mensage =" ha hecho una reserva de la zona común "+zona_comun.descripcion+" a su nombre,el día "+fecha_reserva.strftime('%d/%m/%Y')+" Hora inicial: "+str(hora_inicio)+" hora Final: "+str(hora_final)+" no quedará en firma hasta hasta que efectúe el pago y lo confirme la Administración "
                                asunto = "Reserva Zona Común"
                                from_email = settings.EMAIL_HOST_USER
                                conjunto = conjunto.objects.latest('id')
                                EnvioMailResidente(idapartamento,mensaje,asunto,conjunto.email)
                                #EnvioMailReserva(residente.nombre,residente.email,zona_comun.descripcion,fecha_reserva,hora_inicio,hora_final,residente.apartamento.numero)
                                return redirect('reserva_zonas_comunes',idzona)
                else:
                    reserva.user_id = self.request.user.id
                    reserva.zona_comun_id = idzona
                    reserva.save()
                    user_email = self.request.user.email
                    zona_comun = ZonaComun.objects.get(id=reserva.zona_comun_id)
                    residente = Residente.objects.get(email=user_email)
                    apartamento = Apartamento.objects.get(id=residente.apartamento_id)
                    idapartamento = apartamento.id
                    mensaje =" Se ha creado una reserva temporal de la zona común "+zona_comun.descripcion+" el día "+fecha_reserva.strftime('%d/%m/%Y')+" Hora inicial: "+str(hora_inicio)+" hora Final: "+str(hora_final)+" no quedará en firma hasta hasta que efectúe el pago y lo confirme la Administración "
                    asunto = "Reserva Zona Común Temporal"
                    from_email = settings.EMAIL_HOST_USER
                    EnvioMailResidente(idapartamento,mensaje,asunto,residente.email)
                    mensaje =" ha hecho una reserva de la zona común "+zona_comun.descripcion+" a su nombre,el día "+fecha_reserva.strftime('%d/%m/%Y')+" Hora inicial: "+str(hora_inicio)+" hora Final: "+str(hora_final)+" no quedará en firma hasta hasta que efectúe el pago y lo confirme la Administración "
                    asunto = "Reserva Zona Común"
                    from_email = settings.EMAIL_HOST_USER
                    conjunto = Conjunto.objects.latest('id')
                    EnvioMailResidente(idapartamento,mensaje,asunto,conjunto.email)
                    return redirect('reserva_zonas_comunes',idzona)
                       
        else:
            return redirect('direcciona_reserva_zonas_comunes')
    def get_success_url(self):
        idzona = self.request.session['idzona']
        return reverse_lazy('reserva_zonas_comunes',args=[idzona])

def ConsultaReservaZonasComunesAdministradorView(request,id):
    idreserva = id
    reserva = Reservas.objects.get(id=id)
    reservas = Reservas.objects.filter(id=id)
    idzona = reserva.zona_comun_id
    zona = ZonaComun.objects.filter(id=idzona)
    tipo_usuario = 1
    #return render(request, "core/reserva_zonas_comunes.html", {'iduser':request.user.id})
    return render(request, "core/reserva_zonas_comunes.html", {"zona":zona,"reservas":reservas,"idzona":idzona,'iduser':request.user.id,"tipo_usuario":tipo_usuario})

class EditaReservaZonasComunesView(LoginRequiredMixin,UpdateView):
    model = Reservas
    template_name = 'core/reserva_zonas_comunes_form.html'
    #success_url = reverse_lazy('')
    fields = ['fecha','hora_inicio','hora_final']  

    def get_success_url(self):
        idreserva = self.object.id
        idzona = self.request.session['idzona']
        return reverse_lazy('reserva_zonas_comunes',args=[idzona])

@login_required
def ConfirmaReservaZonasComunesView(request,id):
    reserva = Reservas.objects.get(id=id)
    if reserva.confirmada == 0:
        Reservas.objects.filter(id=id).update(confirmada=1)
        reserva = Reservas.objects.get(id=id)
        user = reserva.user
        email = user.email
        if Residente.objects.filter(email=email).exists():
            residente = Residente.objects.get(email=email)
            idapartamento = residente.apartamento_id
        else:
            idapartamento = 0    
                
        mensaje = " la reserva del "+reserva.zona_comun.descripcion+" el día: "+reserva.fecha.strftime('%d/%m/%Y')+" hora: "+reserva.hora_inicio.strftime("%I:%M %p")+" ha sido confirmada por la administración "
        asunto = "Confirmación Reserva "+reserva.zona_comun.descripcion
        
        EnvioMailResidente(idapartamento,mensaje,asunto,email)
    else:
        Reservas.objects.filter(id=id).update(confirmada=0)
    return redirect('reserva_zonas_comunes_administrador',id)


class BorraReservaZonasComunesView(LoginRequiredMixin,DeleteView):
    model = Reservas
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idreserva = self.object.id
        reserva = Reservas.objects.get(id=idreserva)
        idzona = reserva.zona_comun_id
        return reverse_lazy('reserva_zonas_comunes',args=[idzona])

@login_required
def AutorizacionesPeatonalListaView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
        queryset = Autorizado.objects.all().order_by('-id')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5:
            email_user = request.user.email
            residente = Residente.objects.get(email=email_user)
            apartamento = residente.apartamento
            queryset = Autorizado.objects.filter(apartamento=apartamento).order_by('-id')
    request.session['iduser'] = request.user.id
    f = AutorizadoFilter(request.GET, queryset=queryset)
    autorizaciones = AutorizacionesPeatonalTable(f.qs)
    autorizaciones.paginate(page=request.GET.get("page", 1), per_page=15)
    return render(request, "core/autorizaciones_peatonal_lista.html", {"autorizaciones":autorizaciones,'filter':f,'tipo_usuario':tipo_usuario})

@login_required
def AutorizacionesVehicularListaView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1 :
        queryset = AutorizacionVehiculo.objects.all().order_by('-id')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5 :
            sw = 2
            email_user = request.user.email
            residente = Residente.objects.get(email=email_user)
            apartamento = residente.apartamento
            queryset = AutorizacionVehiculo.objects.filter(apartamento=apartamento).order_by('-id')
    request.session['iduser'] = request.user.id
    email_user = request.user.email
    f = AutorizacionVehiculoFilter(request.GET, queryset=queryset)
    autorizaciones = AutorizacionVehiculoTable(f.qs)
    autorizaciones.paginate(page=request.GET.get("page", 1), per_page=15)
    return render(request, "core/autorizaciones_vehicular_lista.html", {"autorizaciones":autorizaciones,'filter':f,'tipo_usuario':tipo_usuario})

@login_required
def AccesoVehicularListaView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
        sw = 1
        queryset = IngresoVehiculo.objects.all().order_by('-id')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5 :
            email = request.user.email
            residente = Residente.objects.get(email=email)
            idapartamento = residente.apartamento
            queryset = IngresoVehiculo.objects.filter(apartamento_id=idapartamento).order_by('-id')
    f = IngresoVehiculoFilter(request.GET, queryset=queryset)
    ingresos = IngresoVehicularTable(f.qs)
    lista = []
    for i in f.qs:
        lista.append(i.id)
        request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 15, "page": 1}).configure(ingresos)
    return render(request, "core/ingreso_vehicular_lista.html", {"ingresos":ingresos,'filter':f,'tipo_usuario':tipo_usuario})
            
def ReporteIngresoVehicularXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada'] 
    ingresos_vehiculo = IngresoVehiculo.objects.filter(id__in=lista)
    worksheet.write('A1','Hora Ingreso' )
    worksheet.write('B1','Hora Salida' )
    worksheet.write('C1','Tipo Autorización')
    worksheet.write('D1','Placa')
    worksheet.write('E1','Apartamento')
    worksheet.write('F1','Vigilante')
    n = 2
    for j in ingresos_vehiculo :
        nn = str(n)
        if j.hora_ingreso != None:
            singreso= datetime.strftime(j.hora_ingreso,'%A, %B %d, %Y %H:%M:%S')
            exec("worksheet.write('A"+nn+"','"+singreso+"' )")
        if j.hora_salida != None:
            ssalida= datetime.strftime(j.hora_salida,'%A, %B %d, %Y %H:%M:%S')
            exec("worksheet.write('B"+nn+"','"+ssalida+"' )")
        exec("worksheet.write('C"+nn+"','"+str(j.tipo_autoriza.descripcion)+"' )")
        exec("worksheet.write('D"+nn+"','"+j.placa+"' )")
        exec("worksheet.write('E"+nn+"','"+j.apartamento.numero+"' )")
        exec("worksheet.write('F"+nn+"','"+str(j.vigilante.username)+"' )")
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='ingreso_vehicular.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

class CreaAutorizadoPeatonalView(LoginRequiredMixin,CreateView):
    model = Autorizado
    template_name = 'core/autorizado_peatonal_form.html'
    form_class = AutorizadoForm
    
    def form_valid(self, form):
        autorizacion = form.save(commit=False)
        autorizacion.user_id = self.request.user.id
        autorizacion.save()
        autorizado = Autorizado.objects.latest('id')
        conjunto = Conjunto.objects.latest('id')
        email_sol = conjunto.email
        apartamento = Apartamento.objects.get(id=autorizado.apartamento_id)
        message ="Sr. Administrador, se ha ingresado una autorización de ingreso peatonal del Apartamento :"+apartamento.numero+" Identificacion: "+autorizado.identificacion+" Nombre: "+autorizado.nombre
        subject = "Autorización Ingreso Peatonal "
        from_email = settings.EMAIL_HOST_USER
        mail = EmailMessage(
        subject,
        message,
        from_email,
        to=[email_sol],
        bcc=None,
        )
        mail.send()
        email_usuario =  self.request.user.email
        email_sol = email_usuario
        message ="Sr. Residente, acaba de hacer una autorización de ingreso peatonal a su Apartamento :"+apartamento.numero+"< Autorizado Identificacion: "+autorizado.identificacion+" Autorizado Nombre: "+autorizado.nombre+">"
        mail = EmailMessage(
        subject,
        message,
        from_email,
        to=[email_sol],
        bcc=None,
        )
        mail.send()
        iduser = self.request.session['iduser']
        user = User.objects.get(id=iduser)
        tipo_usuario = self.request.session['tipo_usuario'] 
        return redirect('autorizaciones_peatonal_lista')
    
class CreaAutorizadoVehicularView(LoginRequiredMixin,CreateView):
    model = AutorizacionVehiculo
    template_name = 'core/autorizado_vehicular_form.html'
    form_class = AutorizadoVehicularForm
    
    def form_valid(self, form):
        autorizacion = form.save(commit=False)
        autorizacion.user_id = self.request.user.id
        autorizacion.save()
        autorizado = AutorizacionVehiculo.objects.latest('id')
        conjunto = Conjunto.objects.latest('id')
        email_sol = conjunto.email
        apartamento = Apartamento.objects.get(id=autorizado.apartamento_id)
        xxx
        message ="Sr. Administrador, se ha ingresado una autorización de ingreso vehicular del Apartamento :"+apartamento.numero+" Identificacion: "+autorizado.identificacion+" Nombre: "+autorizado.nombre+" Placa: "+autorizado.placa
        subject = "Autorización Ingreso Vehicular "
        from_email = settings.EMAIL_HOST_USER
        mail = EmailMessage(
        subject,
        message,
        from_email,
        to=[email_sol],
        bcc=None,
        )
        mail.send()
        email_usuario =  self.request.user.email
        email_sol = email_usuario
        message ="Sr. Residente, acaba de hacer una autorización de ingreso vehicular a su Apartamento :"+apartamento.numero+"< Autorizado Identificacion: "+autorizado.identificacion+" Autorizado Nombre: "+autorizado.nombre+" Placa: "+autorizado.placa+" >"
        mail = EmailMessage(
        subject,
        message,
        from_email,
        to=[email_sol],
        bcc=None,
        )
        mail.send()
        iduser = self.request.session['iduser']
        user = User.objects.get(id=iduser)
        tipo_usuario = self.request.session['tipo_usuario'] 
        return redirect('autorizaciones_vehicular_lista')
        
class EditaAutorizadoPeatonalView(LoginRequiredMixin,UpdateView):
    model = Autorizado
    template_name = 'core/autorizaciones_peatonal_form.html'
    fields = ['identificacion','nombre','interior','apartamento','tipo_autoriza']
    
    def get_success_url(self):
        tipo_usuario = self.request.session['tipo_usuario'] 
        if tipo_usuario == 1 or tipo_usuario == 2 or tipo_usuario == 5 :
            return redirect('user_home',tipo_usuario)
        else:
            return redirect('user_home',0)
    
class EditaAutorizadoVehicularView(LoginRequiredMixin,UpdateView):
    model = AutorizacionVehiculo
    template_name = 'core/autorizado_vehicular_form.html'
    fields = ['placa','identificacion','nombre','fecha_inicial','fecha_final','interior','apartamento','tipo_autoriza']

    def get_success_url(self):
        tipo_usuario = self.request.session['tipo_usuario'] 
        if tipo_usuario == 1 or tipo_usuario == 2 or tipo_usuario == 5 :
            return redirect('user_home',tipo_usuario)
        else:
            tipo_usuario = 0
            return redirect('user_home',tipo_usuario)

class BorraAutorizadoPeatonalView(LoginRequiredMixin,DeleteView):
    model = Autorizado
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        tipo_usuario = self.request.session['tipo_usuario'] 
        if tipo_usuario == 1 or tipo_usuario == 2 or tipo_usuario == 5 :
            return reverse_lazy('user_home',args=[tipo_usuario])
        else:
            return reverse_lazy('user_home',args=[tipo_usuario])
        

class BorraAutorizadoVehicularView(LoginRequiredMixin,DeleteView):
    model = AutorizacionVehiculo
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        tipo_usuario = self.request.session['tipo_usuario'] 
        if tipo_usuario == 1 or tipo_usuario == 2 or tipo_usuario == 5 :
            return reverse_lazy('user_home',args=[tipo_usuario])
        else:
            return reverse_lazy('user_home',args=[tipo_usuario])

def ValidarAutorizadoPeatonalView(request):
    identificacion = request.GET.get('identificacion', None)
    request.session['idtercero'] = identificacion
    data = {
        'is_taken': Autorizado.objects.filter(identificacion=identificacion).exists()
    }
    if data['is_taken']:
        data['error_message'] = 'El tercero ya fué ingresado'
        return JsonResponse(data)
    else:
        return JsonResponse(data)

def ValidarAutorizadoVehicularView(request):
    placa = request.GET.get('placa', None)
    request.session['placa'] = placa
    data = {
        'is_taken': AutorizacionVehiculo.objects.filter(placa=placa).exists()
    }
    if data['is_taken']:
        data['error_message'] = 'La placa ya fué ingresada'
        return JsonResponse(data)
    else:
        return JsonResponse(data)

def ajaxSwitchActivoAutorizadoPeatonalView(request):
    id = request.GET.get('id', None)
    autorizado = Autorizado.objects.get(id=id)
    if autorizado.activo == True:
        activo = False
    else:
        activo = True
    autorizado = Autorizado.objects.filter(id=id).update(activo=activo)
    data = {'activo':activo}
    return JsonResponse(data)    

def ajaxSwitchPermanenteAutorizadoPeatonalView(request):
    id = request.GET.get('id', None)
    autorizado = Autorizado.objects.get(id=id)
    if autorizado.permanente == True:
        permanente = False
    else:
        permanente = True
    autorizado = Autorizado.objects.filter(id=id).update(permanente=permanente)
    data = {'permanente':permanente}
    return JsonResponse(data)    

def ajaxSwitchActivoAutorizadoVehicularView(request):
    id = request.GET.get('id', None)
    autorizado = AutorizacionVehiculo .objects.get(id=id)
    if autorizado.activo == True:
        activo = False
    else:
        activo = True
    autorizado = Autorizado.objects.filter(id=id).update(activo=activo)
    data = {'activo':activo}
    return JsonResponse(data) 

class CreaAutorizadoPlataformaWebPeatonalView(LoginRequiredMixin,CreateView):
    model = PlataformaWebPeatonal
    template_name = 'core/autorizado_plataforma_web_peatonal_form.html'
    form_class = AutorizadoPlataformaWebPeatonalForm
    success_url = reverse_lazy('autorizacion_plataforma_web_peatonal')

class EditaAutorizadoPlataformaWebPeatonalView(LoginRequiredMixin,UpdateView):
    model = PlataformaWebPeatonal
    template_name = 'core/autorizado_plataforma_web_peatonal_form.html'
    form_class = AutorizadoPlataformaWebPeatonalForm
    success_url = reverse_lazy('autorizacion_plataforma_web_peatonal')
    #fields = ['identificacion','nombre','fecha_inicial','fecha_final','interior','apartamento']

class BorraAutorizadoPlataformaWebPeatonalView(LoginRequiredMixin,DeleteView):
    model = PlataformaWebPeatonal
    success_url = reverse_lazy('autorizacion_plataforma_web_peatonal')
    template_name = 'core/confirmar_borrado_registro.html'

def ValidarAutorizadoPlataformaWebPeatonalView(request):
    identificacion = request.GET.get('identificacion', None)
    request.session['idtercero'] = identificacion
    data = {
        'is_taken': PlataformaWebPeatonal.objects.filter(identificacion=identificacion).exists()
    }
    if data['is_taken']:
        data['error_message'] = 'El tercero ya fué ingresado'
        return JsonResponse(data)
    else:
        return JsonResponse(data)


class CreaAutorizadoPlataformaWebVehiculoView(LoginRequiredMixin,CreateView):
    model = PlataformaWebVehiculo
    template_name = 'core/autorizado_plataforma_web_vehiculo_form.html'
    form_class = AutorizadoPlataformaWebVehiculoForm
    success_url = reverse_lazy('autorizacion_plataforma_web_vehiculo')

class EditaAutorizadoPlataformaWebVehiculoView(LoginRequiredMixin,UpdateView):
    model = PlataformaWebVehiculo
    template_name = 'core/autorizado_plataforma_web_vehiculo_form.html'
    form_class = AutorizadoPlataformaWebVehiculoForm
    success_url = reverse_lazy('autorizacion_plataforma_web_vehiculo')
    #fields = ['identificacion','nombre','fecha_inicial','fecha_final','interior','apartamento']

class BorraAutorizadoPlataformaWebVehiculoView(LoginRequiredMixin,DeleteView):
    model = PlataformaWebPeatonal
    success_url = reverse_lazy('autorizacion_plataforma_web_vehiculo')
    template_name = 'core/confirmar_borrado_registro.html'

def ValidarAutorizadoPlataformaWebVehiculoView(request):
    placa = request.GET.get('placa', None)
    request.session['placa'] = placa
    data = {
        'is_taken': PlataformaWebVehiculo.objects.filter(placa=placa).exists()
    }
    if data['is_taken']:
        data['error_message'] = 'El vehículo ya fué ingresado'
        return JsonResponse(data)
    else:
        return JsonResponse(data)

def AutorizacionIngresoPlataformaWebPeatonalView(request):
    autorizaciones = AutorizacionesPlataformaWebPeatonalTable(PlataformaWebPeatonal.objects.all().order_by('-id'))
    autorizaciones.paginate(page=request.GET.get("page", 1), per_page=15)
    return render(request, "core/autorizaciones_plataforma_web_peatonal_lista.html", {"autorizaciones":autorizaciones})

def AutorizacionIngresoPlataformaWebVehiculoView(request):
    autorizaciones = AutorizacionesPlataformaWebVehiculoTable(PlataformaWebVehiculo.objects.all().order_by('-id'))
    autorizaciones.paginate(page=request.GET.get("page", 1), per_page=15)
    return render(request, "core/autorizaciones_plataforma_web_vehiculo_lista.html", {"autorizaciones":autorizaciones})

def tomar_foto(request):
    id = request.GET.get('identidicacion', None)
    foto(id)
    data = {'id':id}
    return JsonResponse(data) 

def ValidarAutorizadoPlataformaWebPeatonalView(request):
    identificacion = request.GET.get('identificacion', None)
    request.session['identificacion'] = identificacion
    data = {}
    data = PlataformaWebPeatonal.objects.filter(identificacion=identificacion)
    if data:
        autorizado_fields = (
        'identificacion','nombre','interior','apartamento','fecha_inicial','fecha_final')
        data = serialize('json', data,fields=autorizado_fields)
        return HttpResponse(json.dumps(data), content_type='application/json' )
    else:
        data=''
        data = serialize('json', data)
        return HttpResponse(json.dumps(data), content_type='application/json' )


@csrf_exempt
def save_image1(request):
    if request.method== 'POST':        
        path = request.POST["src"]
        image = NamedTemporaryFile()
        image.write(urlib.request.urlopen(path).read())
        image.flush()
        image = File(image)
        name = str(image.name).split('\\')[-1]
        name += '.jpg'
        image.name = name
        obj = Image.objects.create(image=image)
        obj.save()
        return render(request, 'some.html')


def save_filewc(file, path=''):

  filename = "uploadedPicWC.jpg"
  fd = open('%s/%s' % (MEDIA_ROOT1, str(path) + str(filename)), 'wb')
  for chunk in file.chunks():
      fd.write(chunk)
  fd.close()

def DatosAutorizadoPlataformaWebPeatonalView(request):
    data = dict()
    form = AutorizadoPlataformaWebPeatonalForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_autorizacion_plataforma_web_peatonal.html',
        context,
        request=request
    )
    return JsonResponse(data)

def DatosAutorizadoPlataformaWebVehiculoView(request):
    data = dict()
    form = AutorizadoPlataformaWebVehiculoForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_autorizacion_plataforma_web_vehiculo.html',
        context,
        request=request
    )
    return JsonResponse(data)

@csrf_exempt
def GuardaDatosAutorizadoPlataformaWebPeatonalView(request):
    identificacion = request.GET.get('identificacion', None)
    nombre = request.GET.get('nombre', None)
    interior = request.GET.get('interior', None)
    apartamento = request.GET.get('apartamento', None)
    fecha_inicial = request.GET.get('fecha_inicial', None)
    fecha_final = request.GET.get('fecha_final', None)
    autorizado = PlataformaWebPeatonal()
    autorizado.identificacion = identificacion
    autorizado.nombre = nombre
    autorizado.interior_id = interior
    autorizado.apartamento_id = apartamento
    autorizado.fecha_inicial =  datetime.strptime(fecha_inicial,"%Y-%m-%d").date()
    autorizado.fecha_final = datetime.strptime(fecha_final,"%Y-%m-%d").date()
    autorizado.vigilante_id=request.user.id
    autorizado.save()
    data = {'id':id}
    return JsonResponse(data) 

@csrf_exempt
def GuardaDatosAutorizadoPlataformaWebVehiculoView(request):
    placa = request.GET.get('placa', None)
    nombre = request.GET.get('nombre', None)
    interior = request.GET.get('interior', None)
    apartamento = request.GET.get('apartamento', None)
    fecha_inicial = request.GET.get('fecha_inicial', None)
    fecha_final = request.GET.get('fecha_final', None)
    autorizado = PlataformaWebVehiculo()
    autorizado.placa = placa
    autorizado.nombre = nombre
    autorizado.interior_id = interior
    autorizado.apartamento_id = apartamento
    autorizado.fecha_inicial =  datetime.strptime(fecha_inicial,"%Y-%m-%d").date()
    autorizado.fecha_final = datetime.strptime(fecha_final,"%Y-%m-%d").date()
    autorizado.vigilante_id=request.user.id
    autorizado.save()
    data = {'id':id}
    return JsonResponse(data)

def DatosAutorizadoPlataformaWebPeatonalEditarView(request):
    id = request.GET.get('id', None)
    request.session['idAutorizado'] = id
    data = dict()
    form = AutorizadoPlataformaWebPeatonalForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_autorizacion_plataforma_web_peatonal_editar.html',
        context,
        request=request
    )
    data['html_form'] = data['html_form']
    return JsonResponse(data)

def DatosAutorizadoPlataformaWebVehiculoEditarView(request):
    id = request.GET.get('id', None)
    request.session['idAutorizado'] = id
    data = dict()
    form = AutorizadoPlataformaWebVehiculoForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_autorizacion_plataforma_web_vehiculo_editar.html',
        context,
        request=request
    )
    data['html_form'] = data['html_form']
    return JsonResponse(data)

def CargaDatosAutorizadoPlatafromaWebPeatonalView(request):
    data = dict()
    id = request.session['idAutorizado']
    autorizado = PlataformaWebPeatonal.objects.get(id=id)
    data = {'id':id,'identificacion':autorizado.identificacion,'nombre':autorizado.nombre,'interior':autorizado.interior_id,
    'apartamento':autorizado.apartamento_id,'fecha_inicial':autorizado.fecha_inicial,'fecha_final':autorizado.fecha_final,'tipo_plataforma':tipo_plataforma}
    form = AutorizadoPlataformaWebPeatonalForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_autorizacion_plataforma_web_peatonal_editar.html',
        context,
        request=request
    )
    return JsonResponse(data)

def CargaDatosAutorizadoPlatafromaWebVehiculoView(request):
    data = dict()
    id = request.session['idAutorizado']
    autorizado = PlataformaWebVehiculo.objects.get(id=id)
    data = {'id':id,'placa':autorizado.placa,'nombre':autorizado.nombre,'interior':autorizado.interior_id,
    'apartamento':autorizado.apartamento_id,'fecha_inicial':autorizado.fecha_inicial,'fecha_final':autorizado.fecha_final,'tipo_plataforma':tipo_plataforma}
    form = AutorizadoPlataformaWebVehiculoForm()
    context = {'form': form}
    data['html_form'] = render_to_string('core/datos_autorizacion_plataforma_web_peatonal_editar.html',
        context,
        request=request
    )
    return JsonResponse(data)
@csrf_exempt

def BorrarParqueaderoView(request,id):
    parqueadero = Parqueadero.objects.get(id=id)
    idinterior = parqueadero.interior_id
    idapartamento = parqueadero.apartamento_id
    Parqueadero.objects.filter(id=id).delete()
    residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
    idresidente =residente.id
    return redirect('informacion_residente_detalle',id=idresidente)

def BorrarVehiculoView(request,id):
    parqueadero = Parqueadero.objects.get(id=id)
    idinterior = parqueadero.interior_id
    idapartamento = parqueadero.apartamento_id
    Vehiculo.objects.filter(id=id).delete()
    residente = Residente.objects.get(interior_id=idinterior,apartamento_id=idapartamento)
    idresidente =residente.id
    return HttpResponseRedirect(reverse('informacion_residente_detalle',idresidente))

class CrearVehiculoView(LoginRequiredMixin,CreateView):
    model = Vehiculo
    template_name = 'core/vehiculo_form.html'
    form_class = VehiculoForm
    #success_url = reverse_lazy('informacion_residentes')

    def get_initial(self,*args,**kwargs):
        idinterior =self.kwargs['idinterior']
        idapartamento =self.kwargs['idapartamento'] 
        initial=super(CrearVehiculoView,self).get_initial(**kwargs)
        initial['interior']= idinterior
        initial['apartamento']= idapartamento
        return initial

    def form_valid(self, form):
        if form.is_valid():
            object = form.save(commit=False)
            object.interior_id = self.kwargs['idinterior']
            object.apartamento_id = self.kwargs['idapartamento']
            idinterior = self.kwargs['idinterior']
            idapartamento = self.kwargs['idapartamento']
            residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
            object.save()
            return redirect('informacion_residente_detalle',id=residente.id)
        else:
            messageform='Información Inválida'
            return render(self.request, "core/mensaje_error.html", {'form': messageform})
        return HttpResponseRedirect('informacion_residentes')


class CrearParqueaderoView(LoginRequiredMixin,CreateView):
    model = Parqueadero
    template_name = 'core/parqueadero_form.html'
    form_class = ParqueaderoForm
    #success_url = reverse_lazy('informacion_residentes')

    def get_initial(self,*args,**kwargs):
        idinterior =self.kwargs['idinterior']
        idapartamento =self.kwargs['idapartamento'] 
        initial=super(CrearParqueaderoView,self).get_initial(**kwargs)
        initial['interior']= idinterior
        initial['apartamento']= idapartamento
        return initial
    
    def form_valid(self, form):
        if form.is_valid():
            object = form.save(commit=False)
            object.interior_id = self.kwargs['idinterior']
            object.apartamento_id = self.kwargs['idapartamento']
            idinterior = self.kwargs['idinterior']
            idapartamento = self.kwargs['idapartamento']
            residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
            object.save()
            return redirect('informacion_residente_detalle',id=residente.id)
        else:
            messageform='Información Inválida'
            return render(self.request, "core/mensaje_error.html", {'form': messageform})
        return HttpResponseRedirect('informacion_residentes')

""" class ResidenteView(TemplateView):
    template_name = 'core\pagina_residentes.html' """

class BlogView(TemplateView):
    template_name = 'core\blog.html'

class ConsejeroView(TemplateView):
    template_name = 'core\pagina_consejero.html'

def InformacionResidentesView(request):
    queryset = Residente.objects.select_related().all()
    f = ResidentesFilter(request.GET, queryset=queryset)
    lista = []
    residentes = ResidenteTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 25, "page": 1}).configure(residentes)
    return render(request, "core/residentes_lista.html", {'residentes':residentes,'filter': f})
   
    
def SolicitudesResidentesView(request):
    residentes = ResidenteTempTable(ResidenteTemp.objects.all().order_by('interior','apartamento'))
    apartamentos = ApartamentosTable(Apartamento.objects.all())
    return render(request, "core/residentes_temp_lista.html", {"residentes":residentes,'apartamentos':apartamentos})

def SolicitudesResidentesUnoView(request,id):
    residentes = ResidenteTempTable(ResidenteTemp.objects.filter(id=id).order_by('interior','apartamento'))
    apartamentos = ApartamentosTable(Apartamento.objects.all())
    return render(request, "core/residentes_temp_lista.html", {"residentes":residentes,'apartamentos':apartamentos})
    
def SeleccionParametroApartamentoResidentesView(request):
    id = request.GET.get('id', None)
    request.session['idapto'] = id
    data = {'id':id}
    return JsonResponse(data)

def InformacionResidentesFiltroApartamentoView(request):
    idapto = request.session['idapto']
    residentes = ResidenteTable(Residente.objects.filter(apartamento_id=idapto))
    apartamentos = ApartamentosTable(Apartamento.objects.all())
    return render(request, "core/residentes_lista.html", {"residentes":residentes,'apartamentos':apartamentos})

@login_required
def InformacionResidenteDetalleView(request,id):
    residente = Residente.objects.get(id=id)
    idinterior = residente.interior_id
    idapartamento = residente.apartamento_id
    residentes = Residente.objects.filter(interior_id=idinterior,apartamento=idapartamento)
    vehiculos = Vehiculo.objects.filter(interior_id=idinterior,apartamento_id=idapartamento)
    mascotas = Mascota.objects.filter(interior_id=idinterior,apartamento_id=idapartamento)
    parqueaderos = Parqueadero.objects.filter(interior_id=idinterior,apartamento_id=idapartamento)
    depositos = Deposito.objects.filter(interior_id=idinterior,apartamento_id=idapartamento)
    return render(request, "core/residentes_detalle.html", {"residentes":residentes,'vehiculos':vehiculos,'parqueaderos':parqueaderos,'mascotas':mascotas,'depositos':depositos})

def BorrarResidenteView(request,id):
    Residente.objects.filter(id=id).delete()
    id = request.user.id
    return HttpResponseRedirect(reverse('informacion_residentes'))

class CrearMascotaView(LoginRequiredMixin,CreateView):
    model = Mascota
    template_name = 'core/mascota_form.html'
    form_class = MascotaForm
    #success_url = reverse_lazy('informacion_residentes')

    def get_initial(self,*args,**kwargs):
        idinterior =self.kwargs['idinterior']
        idapartamento =self.kwargs['idapartamento'] 
        initial=super(CrearMascotaView,self).get_initial(**kwargs)
        initial['interior']= idinterior
        initial['apartamento']= idapartamento

        return initial

    def form_valid(self, form):
        if form.is_valid():
            object = form.save(commit=False)
            object.interior_id = self.kwargs['idinterior']
            object.apartamento_id = self.kwargs['idapartamento']
            idinterior = self.kwargs['idinterior']
            idapartamento = self.kwargs['idapartamento']
            residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
            object.save()
            return redirect('informacion_residente_detalle',id=residente.id)
        else:
            messageform='Información Inválida'
            return render(self.request, "core/mensaje_error.html", {'form': messageform})
        return HttpResponseRedirect('informacion_residentes')

class EditarMascotaView(LoginRequiredMixin,UpdateView):
    model = Mascota
    form_class = MascotaForm
    template_name = 'core/mascota_form.html'
    success_url = reverse_lazy('informacion_residentes')
    #fields = ['identificacion','nombre','interior','apartamento','tipo_autoriza']

    def get_success_url(self):
        idmascota = self.object.id
        mascota = Mascota.objects.get(id=idmascota)
        idinterior = mascota.interior_id
        idapartamento = mascota.apartamento_id
        residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
        return reverse_lazy('informacion_residente_detalle',args=[residente.id])

def BorraMascotaView(request,id):
    mascota = Mascota.objects.get(id=id)
    mascota = Mascota.objects.filter(id=id).delete()
    idinterior = mascota.interior_id
    idapartamento = mascota.apartamento_id
    # Pueden ser varios, tomo el ultimo para poder tomar un id de un residente de ese interior/apartamento
    residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
    return redirect('informacion_residentes_detalle',residente.id) 

class CrearResidenteView(LoginRequiredMixin,CreateView):
    model = Residente
    template_name = 'core/residente_form.html'
    form_class = ResidenteForm
    success_url = reverse_lazy('informacion_residentes')

class CrearResidenteTempView(CreateView):
    model = ResidenteTemp
    template_name = 'core/residente_form.html'
    form_class = ResidenteTempForm
    success_url = reverse_lazy('user_home',args=[5])

    def get_success_url(self):
        id = self.object.id
        residente_temp = ResidenteTemp.objects.get(id=id)
        email_sol = residente_temp.email
        nombre_sol = residente_temp.nombre
        idapartamento = residente_temp.apartamento_id
        apartamento = Apartamento.objects.get(id=idapartamento)
        mensaje =" "+nombre_sol+" su solicitud ha sido enviada al administrador, una vez verificada,via mail le será enviado un token para que se registre en el sistema"
        asunto = "Solicitud registro de datos residente"
        EnvioMailResidente(idapartamento,mensaje,asunto,email_sol)
        return reverse('user_home',0)

def AdicionarResidenteDesdeTemp(self,id):
    res_temp = ResidenteTemp.objects.get(id=id)
    identificacion = res_temp.identificacion
    if Residente.objects.filter(identificacion=identificacion).exists():
        token = get_random_string(length=12)
        Residente.objects.filter(identificacion=identificacion).update(nombre =res_temp.nombre,tipo_residente=res_temp.tipo_residente ,telefono = res_temp.telefono,
        celular = res_temp.celular,email= res_temp.email,edad = res_temp.edad,genero = res_temp.genero,persona_discapacitada = res_temp.persona_discapacitada,
        interior = res_temp.interior,apartamento = res_temp.apartamento,token = token)
        ResidenteTemp.objects.filter(id=id).delete()
        mensaje =" "+res_temp.nombre+" su solicitud ha sido respondida por el administrador, su token es: "+token+ ", ingrese por la opcion <usuario no registrado> ingrese los datos y haga login ingresado un nombre y clave propios para que se registre cada vez que quiera interactuar con ésta página"
        asunto = "Respuesta solicitud registro de datos residente"
        apartamento = Apartamento.objects.get(numero=res_temp.apartamento)
        email = res_temp.email
        EnvioMailResidente(apartamento.id,mensaje,asunto,email)
        return redirect('pagina_informacion_residentes')
    else:
        residente_temp = ResidenteTemp.objects.filter(id=id)
        for i in residente_temp:
            token = get_random_string(length=12)    
            residente = Residente()
            residente.identificacion = i.identificacion
            residente.nombre = i.nombre
            residente.tipo_residente = i.tipo_residente
            residente.telefono = i.telefono
            residente.celular = i.celular
            residente.email= i.email
            residente.edad = i.edad
            residente.genero = i.genero
            residente.persona_discapacitada = i.persona_discapacitada
            residente.interior_id =i.interior_id
            residente.apartamento_id = i.apartamento_id
            residente.token = token
            residente.save()
            ResidenteTemp.objects.filter(id=id).delete()
            mensaje =" "+i.nombre+" su solicitud ha sido respondida por el administrador, su token es: "+token+ ", ingrese por la opcion <usuario no registrado> ingrese los datos y haga login ingresado un nombre y clave propios para que se registre cada vez que quiera interactuar con ésta página"
            asunto = "Respuesta solicitud registro de datos residente"
            apartamento = Apartamento.objects.get(id=i.apartamento_id)
            email = i.email
            EnvioMailResidente(apartamento.id,mensaje,asunto,email)
            return redirect('pagina_informacion_residentes')

class EditarResidenteView(LoginRequiredMixin,UpdateView):
    model = Residente
    form_class = ResidenteForm
    template_name = 'core/residente_form.html'
    success_url = reverse_lazy('informacion_residentes')
    #fields = ['identificacion','nombre','interior','apartamento','tipo_autoriza']

    def get_success_url(self):
        idresidente = self.object.id
        return reverse_lazy('informacion_residente_detalle',args=[idresidente])


class EditarParqueaderoView(UpdateView):
    model = Parqueadero
    form_class = ParqueaderoForm
    template_name = 'core/parqueadero_form.html'
    success_url = reverse_lazy('informacion_residentes')
    #fields = ['identificacion','nombre','interior','apartamento','tipo_autoriza']

    def get_success_url(self):
        idparqueadero = self.object.id
        parqueadero = Parqueadero.objects.get(id=idparqueadero)
        idinterior = parqueadero.interior_id
        idapartamento = parqueadero.apartamento_id
        residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
        return reverse_lazy('informacion_residente_detalle',args=[residente.id])

class EditarVehiculoView(LoginRequiredMixin,UpdateView):
    model = Vehiculo
    form_class = VehiculoForm
    template_name = 'core/vehiculo_form.html'
    success_url = reverse_lazy('informacion_residentes')
    #fields = ['identificacion','nombre','interior','apartamento','tipo_autoriza']

    def get_success_url(self):
        idvehiculo = self.object.id
        vehiculo = Vehiculo.objects.get(id=idvehiculo)
        idinterior = vehiculo.interior_id
        idapartamento = vehiculo.apartamento_id
        residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
        return reverse_lazy('informacion_residente_detalle',args=[residente.id])

def PoneVehiculoView(request,idinterior,idapartamento):
    vehiculo = Vehiculo()
    vehiculo.interior_id = idinterior
    vehiculo.apartamento_id = idapartamento
    vehiculo.save()
    return redirect('vehiculo_form',id)

def ResidenteVehiculoView(request,id):
    Vehiculo.objects.filter(id=id).delete()
    return render(request, "core/informacion_residente_detalle.html")

def BorraVehiculoView(request,id):
    vehiculo = Vehiculo.objects.get(id=id)
    Vehiculo.objects.filter(id=id).delete()
    idinterior = vehiculo.interior_id
    idapartamento = vehiculo.apartamento_id
    # Pueden ser varios, tomo el ultimo para poder tomar un id de un residente de ese interior/apartamento
    residente = Residente.objects.filter(interior_id=idinterior,apartamento_id=idapartamento).latest('id')
    return redirect('informacion_residente_detalle',residente.id) 

def EditaVehiculoView(request):
    idinterior = request.GET.get('idinterior', None)
    idapartamento = request.GET.get('idapartamento', None)
    form = VehiculoForm()
    context = {'form': form}
    html_form = render_to_string('core/vehiculo_form.html',
        context,
        request=request,
    )
    return JsonResponse({'html_form': html_form})

def SeleccionEmailAnexoView(request,id,tipomail):
    request.session['idresidente'] = id
    return render(request, "core/seleccion_con_sin_anexo.html", {'id':id,'tipomail':tipomail})
    #template_name = "core/seleccion_con_sin_anexo.html"

def EnvioCorreosMasivosResidentesView(request):
    Residente.objects.update(envio_email=False)
    queryset = Residente.objects.all()
    f = ResidentesFilter(request.GET, queryset=queryset)
    lista = []
    residentes = CorreosResidentesTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    return render(request, "core/pagina_correo_residentes.html", {'residentes':residentes,'filter': f})

def ajaxSwitchMarcaResidentesCorreosView(request):
    id = request.GET.get('id', None)
    residente = Residente.objects.get(id=id)
    if residente.envio_email == True:
        envio_email = False
    else:
        envio_email = True
    Residente.objects.filter(id=id).update(aprobado=envio_email)    
    data = {'envio':envio_mail}
    return JsonResponse(data) 

def MarcaEnvioCorreoResidentesView(request):
    #Residente.objects.all().update(envio_email=False)
    lista = request.session['lista_filtrada']
    residentes = Residente.objects.filter(id__in=lista)
    for i in residentes:
        print(i.id)
        if i.envio_email == True:
            Residente.objects.filter(id=i.id).update(envio_email=False)
        else:
            Residente.objects.filter(id=i.id).update(envio_email=True)
    residentes = Residente.objects.filter(envio_email=True)        
    f = ResidentesFilter(request.GET, queryset=residentes)
    residentes = CorreosResidentesTable(f.qs)
    return render(request, "core/pagina_correo_residentes_result.html", {'residentes':residentes})                    

 
def crea_user_perfil(request):
    interior = Interior.objects.all()
    apartamento = Apartamento.objects.all()
    user = User()
    user_perfil = UserPerfil()
    residente = Residente()
    for i in interior:
        for l in apartamento:
            password = get_random_string(length=12)
            user.password = password
            user.is_superuser = False
            user.username = 'Interior'+i.numero.strip() + l.numero.strip()
            user.lastname = 'Interior'+i.numero.strip()
            user.email = 'user'+i.numero.strip() + l.numero.strip()+'@gmail.com'
            user.is_staff = False
            user.is_active = True
            user.date_joined = date.today()
            user.first_name = l.numero.strip()
            user.save()
            user.id +=1
            user_last = User.objects.latest('id')
            user_perfil.user_id = user_last.id
            xinterior = Interior.objects.get(numero=i.numero)
            xapartamento = Apartamento.objects.get(numero=l.numero)
            user_perfil.interior_id = xinterior.id
            user_perfil.apartamento_id = xapartamento.id
            user_perfil.save()
            user_perfil.id +=1
    mensaje = "Usuarios Creados"
    return render(request, "core/mensaje_usuarios_creados.html", {'mensaje': mensaje})           

def ajax_validar_email(request):
    email = request.GET.get('email', None)
    data = {
        'is_taken': User.objects.filter(email=email).exists()
    }
    if data['is_taken']:
        data['error_message'] = 'Este email ya existe en nuestros registros'
        return JsonResponse(data)
    else:
        return JsonResponse(data)

def EnvioMailUsuario(iduser,nombre_res,nombre_sol,email_sol,password):
    clave = password
    email = email_sol
    message ="Sr/a. "+nombre_res+" su usuario inicial y clave para accesar al sistema son: usuario: "+nombre_sol+" y su clave es: "+clave+", se sugiere cambiarla con frecuencia"
    subject = "Credenciales de acceso al porgrama CONDOMINIUM"
    from_email = settings.EMAIL_HOST_USER
    destinatarios = {'email':email}
    mail = EmailMessage(
            subject,
            message,
            from_email,
            to=[email],
            bcc=None,
    )
    mail.send()
    return True


def NuevoUsuario(request):
    if request.method == 'POST':
        formulario = UserCreationForm(request.POST)
        if formulario.is_valid:
            formulario.save
            return HttpResponseRedirect('main')
    else:
        formulario = UserCreationForm()
    return render(request, "core/nuevousuario.html", {'formulario':formulario},context_instance=RequestContext(request))     

class VerUsuariosView(ListView):
    template_name = 'registration/listado_usuarios.html'
    model = UserPerfil
    paginate_by = 4

    def get_context_data(self, **kwargs):
        context = super( VerUsuariosView, self).get_context_data(**kwargs)
        
        table1 = UserPerfilTable(UserPerfil.objects.all().order_by('-created' ))
        table1.paginate(page=self.request.GET.get("page", 1), per_page=10)
        usuario = UserPerfil.objects.all()
                
        context = {'table1':table1}
        return context

class EditarUsuariosView(LoginRequiredMixin,UpdateView):
    model = UserPerfil
    success_url = reverse_lazy('core:ver_usuario')
    fields = ['nombre','email','usuario','interior','apartamento','clave']    


def NuevoUsuarioView(request):
    if request.method == "POST":
        form = NewUserForm(request.POST)
        form_perfil = UserPerfilForm(request.POST)
        if form.is_valid() and form_perfil.is_valid():
            user = form.save()
            perfil = form_perfil.save(commit=False)
            user = User.objects.latest('id')
            perfil.user_id = user.id
            perfil.save()
            do_login(request, user)
            messages.success(request, "Registro efectuado" )
            return redirect("main")
            messages.error(request, "Registro no efectuado, Información Inválida.")
        else:
           mensaje = "Error"
           return render(request, "core/mensaje_error_usuario.html", {'mensaje':mensaje})
    form = NewUserForm()
    form_perfil = UserPerfilForm(request.POST)
    return render (request=request, template_name="registration/registro_usuario.html", context={"register_form":form,"register_perfil":form_perfil})

##################################### INGRESO VEHICULAR ############################

def VerFotoVehiculo(request):
    cap=cv2.VideoCapture(0)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out = cv2.VideoWriter('outpu.avi',fourcc,20.0,(640,480))
    while(cap.isOpened()):
        ret,frame = cap.read()
        if ret == True:
            out.write(frame)
            cv2.imshow('frame',frame)

            if cv2.waitKey(1):
                break
        else:
            break
    cap.release()
    out.release()        
    cv2.destroyAllWindows()        
    return redirect('/ingreso_vehicular_busqueda_parametro')

def DireccionaSalidaAccesoPeatonalView(request):
    tipo_usuario = request.session['tipo_usuario'] 
    if tipo_usuario == 1:
       return redirect('pagina_administrador')
    else:
        if tipo_usuario == 4:
            return redirect('pagina_vigilancia')
        else:
            if tipo_usuario == 2 or tipo_usuario == 5 :
                return redirect('pagina_accesos')
            

def DireccionaSalidaAccesoVehicularView(request):
    tipo_usuario = request.session['tipo_usuario'] 
    if tipo_usuario == 1:
       return redirect('pagina_administrador')
    else:
        if tipo_usuario == 4:
            return redirect('pagina_vigilancia')
        else:
            if tipo_usuario == 2 or tipo_usuario == 5 :
                return redirect('pagina_accesos')
               

def DireccionaPaginaVigilanciaVehicularlView(request):
    tipo_usuario = request.session['tipo_usuario'] 
    if tipo_usuario == 4:
        return redirect('pagina_vigilancia')
    else:
        if tipo_usuario == 1:
            return redirect('pagina_administrador')        
        else:    
            return redirect('user_home',tipo_usuario)

def DireccionaPaginaVigilanciaPeatonalView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 4:
        return redirect('pagina_vigilancia')
    else:
        messages.info(request, f"Debe registrarse como Vigilante")
        tipo_usuario=request.session['tipo_Usuario']
        return redirect('user_home',tipo_usuario)    
            

def DireccionaSalidaAutorizacionesVehicularView(request):
    tipo_usuario = request.session['tipo_usuario']
    print('Tipo Usuario: ',tipo_usuario)
    if tipo_usuario == 1:
       return redirect('pagina_accesos')
    else:
        if tipo_usuario == 4:
            return redirect('pagina_vigilancia')
        else:
            if tipo_usuario == 2 or tipo_usuario ==5:
                #return redirect('user_home',tipo_usuario)
                return redirect('pagina_autorizaciones')
            
def IngresoVehicularBusquedaView(request):
    if request.method == 'GET':
        query= request.GET.get('q')
        placa=query
        request.session['placa'] = placa
        submitbutton= request.GET.get('submit')
        request.session['placa']=placa
        tipo_autorizacion = 0
        desc_autoriza='No tiene'
        fecha_inicial=None
        fecha_final=None
        mensaje=''
        nom_interior = ''
        nom_apartamento = ''
        tipo_ingreso = 0
        if  AutorizacionVehiculo.objects.filter(placa=placa).exists():
            results= AutorizacionVehiculo.objects.filter(placa=placa)
            autorizado = AutorizacionVehiculo.objects.get(placa=placa)
            tipo_autorizacion = autorizado.tipo_autoriza_id
            tipoautoriza = TipoAutoriza.objects.get(id=tipo_autorizacion)
            desc_autoriza = tipoautoriza.descripcion
            idinterior = autorizado.interior_id
            idapartamento = autorizado.apartamento_id
            interior = Interior.objects.get(id=idinterior)
            apartamento = Apartamento.objects.get(id=idapartamento)
            nom_interior = interior.numero
            nom_apartamento = apartamento.numero
            tipo_autorizacion = 1
            request.session['tipoautorizacion'] = 2
            mensaje = 'AUTORIZADO! HACER INGRESO'  
            if VisitanteVehiculo.objects.filter(placa=placa).exists():
                visitante=VisitanteVehiculo.objects.get(placa=placa)
                ingresos = IngresoVehiculo.objects.filter(placa=placa).order_by('-id')[:5]
            else:
                autorizacion_vehiculo = AutorizacionVehiculo.objects.get(placa=placa)    
                a = VisitanteVehiculo(placa=placa,identificacion=autorizacion_vehiculo.identificacion,nombre=autorizacion_vehiculo.nombre)
                a.save()
        elif PlataformaWebVehiculo.objects.filter(placa=placa).exists():
            request.session['tipoautorizacion'] = 3
            results= PlataformaWebVehiculo.objects.filter(placa=placa)
            ABautorizado = PlataformaWebVehiculo.objects.get(placa=placa)
            fecha_inicial = ABautorizado.fecha_inicial
            fecha_final = ABautorizado.fecha_final
            idinterior = ABautorizado.interior_id
            idapartamento = ABautorizado.apartamento_id
            interior = Interior.objects.get(id=idinterior)
            apartamento = Apartamento.objects.get(id=idapartamento)
            nom_interior = interior.numero
            nom_apartamento = apartamento.numero
            tipo_autorizacion = 3
            if fecha_final < date.today():
                mensaje = 'AUTORIZACION HA VENCIDO'
            else:
                mensaje = 'AUTORIZADO PLATAFORMA WEB REVISAR DOCUMENTACION Y CREAR INGRESO'    
            desc_autoriza = 'Autorizacion Plataforma Web '
            
            if VisitanteVehiculo.objects.filter(placa=placa).exists():
                request.session['tipoautorizacion'] = 1
                visitante=Visitante.objects.get(Placa=placa)
                ingresos = IngresoPeatonal.objects.filter(placa=placa).order_by('-id')[:5]
            else:    
                ingresos = None
                visitante = ''
        elif VisitanteVehiculo.objects.filter(placa=placa).exists():
            
            visitante = VisitanteVehiculo.objects.get(placa=placa)
            ingresos =IngresoVehiculo.objects.filter(placa=placa).order_by('-id')[:5]
            results= VisitanteVehiculo.objects.filter(placa=placa).distinct()
            mensaje = 'VISITANTE, SOLICITAR AUTORIZACION AL RESIDENTE'
            tipo_autorizacion = 1
            request.session['tipoautorizacion'] = 1
        else:
            visitante = ''    
            ingresos = IngresoVehiculo.objects.all().order_by('-id')[:5]
            #results= VisitanteVehiculo.objects.filter(placa=placa).distinct()
            results= VisitanteVehiculo.objects.filter(placa=placa).distinct()
            request.session['tipoautorizacion'] = 1 
            mensaje = 'VISITANTE, SOLICITAR AUTORIZACION AL RESIDENTE'   
        return render(request,'core/ingreso_vehicular.html', {'results':results,'submitbutton': submitbutton,'ingresos':ingresos,
        'query':query,'tipo_autorizacion':tipo_autorizacion,'fecha_inicial':fecha_inicial,'fecha_final':fecha_final,'visitante':visitante,
        'desc_autoriza':desc_autoriza,'mensaje':mensaje,'nom_interior':nom_interior,'nom_apartamento':nom_apartamento,'tipo_autirizacion':tipo_autorizacion})
    else:
        ingresos = None
        query=''
        results=''
        submitbutton=''
        mensaje1 ='VISITANTE NO ENCONTRADO, INGRESAR DATOS'
        return render(request, 'core/ingreso_vehicular.html',{'results':results,'submitbutton': submitbutton,'ingresos':ingresos,'identificacion:':id,'mensaje1':mensaje1})
            
def IngresoVehicularBusquedaParametroView(request):
    placa = request.session['placa']
    submitbutton= request.GET.get('submit')
    results= VisitanteVehiculo.objects.filter(placa=placa)
    if VisitanteVehiculo.objects.filter(placa=placa).exists():
        ingresos =IngresoVehiculo.objects.filter(placa=placa).order_by('-id')[:5]
    else:
        visitante = None    
        ingresos =None
    return render(request,'core/ingreso_vehicular.html', {'results':results,'submitbutton': submitbutton,'ingresos':ingresos,'placa':placa})
    
def AccesoVehicularView(request):
    return render(request,'core\ingreso_vehicular.html')

class CreaIngresoVehicularView(LoginRequiredMixin,CreateView):
    model = IngresoVehiculo
    template_name = 'core/ingreso_vehicular_form.html'
    form_class = DatosIngresoVehicularForm
    success_url = reverse_lazy('acceso_vehicular') 

    def form_valid(self, form):
        if form.is_valid():
            self.object = form.save(commit=False)
            placa = self.request.session['placa']
            idinterior = form.cleaned_data['interior']
            interior = Interior.objects.get(numero=idinterior)
            numero_apartamento = form.cleaned_data['apartamento']
            apartamento = Apartamento.objects.get(numero=numero_apartamento)
            self.object.interior_id = interior.id
            self.object.apartamento_id = apartamento.id
            self.object.placa = placa
            self.object.vigilante_id = self.request.user.id
            self.object.tipo_autoriza_id = self.request.session['tipoautorizacion']
            self.object.save()
            if VisitanteVehiculo.objects.filter(placa=placa):
                visitante_vehiculo = VisitanteVehiculo.objects.get(placa=placa)
            else:
                nombre = self.request.session['nombre_visitante_vehicular']
                identificacion = self.request.session['cc_visitante_vehicular']
                a = Visitante(identificacion=identificacion,nombre=nombre)
                a.save()
            visitante_vehiculo = VisitanteVehiculo.objects.get(placa=placa)        
            #apartamento = Apartamento.objects.get(id=apartamento.id)
            mensaje =" El Vehículo con placa "+visitante_vehiculo.placa+" nombre: "+visitante_vehiculo.nombre+" solicita ingreso a su apartamento "
            asunto = "Solicitud Ingreso vehicular "
            email = ''
            EnvioMailResidente(apartamento.id,mensaje,asunto,email)
            return redirect('acceso_vehicular')

@csrf_exempt
def GuardaDatosIngresoVehicularView(request):
    placa = request.GET.get('placa', None)
    interior = request.GET.get('interior', None)
    apartamento = request.GET.get('apartamento', None)
    visitante = VisitanteVehiculo.objects.get(placa=placa)
    ingreso_vehicular = IngresoVehiculo()
    ingreso_vehicular.placa_id = visitante.id
    ingreso_vehicular.interior_id = interior
    ingreso_vehicular.apartamento_id = apartamento
    ingreso_vehicular.vigilante_id=request.user.id
    ingreso_vehicular.hora_salida=None
    ingreso_vehicular.save()
    data = {'id':placa.value}
    return JsonResponse(data) 


def DatosVisitanteVehicularView(request):
    form =  VisitanteVehiculoForm()
    context = {'form': form}
    html_form = render_to_string('core/crear_visitante_vehicular.html',
    context,
    request=request,
    )
    return JsonResponse({'html_form': html_form})

def GuardaDatosVisitanteVehicularView(request):
    identificacion = request.GET.get('identificacion', None)
    placa = request.GET.get('placa', None)
    nombre = request.GET.get('nombre', None)
    request.session['nombre_visitante_vehicular']=nombre
    request.session['cc_visitante_vehicular']=identificacion
    visitante = VisitanteVehiculo()
    visitante.nombre = nombre
    visitante.identificacion = identificacion
    visitante.placa = placa
    visitante.save()
    data = {'id':identificacion}
    return JsonResponse(data) 

class CrearVisitanteView(LoginRequiredMixin,CreateView):
    model = Visitante
    template_name = 'core/crear_visitante.html'
    form_class = VisitanteForm
        
    def form_valid(self, form):
        if Parametros.objects.filter(user=self.request.user.id).exists():
            parametros = Parametros.objects.filter(user=self.request.user.id).latest('id')
            cc = parametros.valor_parametro_uno
            self.object = form.save(commit=False)
            cc = self.kwargs.get('cc')
            self.object.identificacion = cc
            self.object.save()
            Parametros.objects.filter(user=self.request.user.id).delete()
            return HttpResponseRedirect(reverse('ingreso_peatonal_acceso',cc))    

def SalidaVehicularView(request):
    id = request.GET.get('id', None)
    IngresoVehiculo.objects.filter(id=id).update(hora_salida=datetime.now())
    fecha = datetime.today().strftime('%A, %B %d, %Y %H:%M:%S')
    data = {'id':id,'fecha':fecha}
    return JsonResponse(data) 


########################################### Activos Fijos ################################

class CreaActivoFijoView(LoginRequiredMixin,CreateView):
    model = ActivoFijo
    template_name = 'core/activo_fijo_form.html'
    form_class = ActivoFijoForm
    success_url = reverse_lazy('activos_fijos_lista')

    def get_initial(self,*args,**kwargs):
        initial=super(CreaActivoFijoView,self).get_initial(**kwargs)
        initial['tipo_activo']=self.request.session['tipo_activo']
        initial['nombre']=self.request.session['nombre_activo']
        return initial

class EditaActivoFijoView(LoginRequiredMixin,UpdateView):
    model = ActivoFijo
    fields = ['nombre','tipo_activo','descripcion','marca','serial','valor_libros','cantidad','mantenimiento','frecuencia_mantenimiento','ultimo_mantenimiento','placa_numero','prestado','estado']
    success_url = reverse_lazy('activos_fijos_lista')
    template_name = 'core/activo_fijo_form.html'

    def get_success_url(self):
        idactivo = self.object.id
        return reverse_lazy('activos_fijos_detalle',args=[idactivo])


class BorraActivoFijoView(LoginRequiredMixin,DeleteView):
    model = ActivoFijo
    success_url = reverse_lazy('activos_fijos_lista')
    template_name = 'core/confirmar_borrado_registro.html'

@login_required
def ActivosFijosDetalleView(request,id):
    request.session['UltimaPaginaVisitada']='Activos'
    activo1 =ActivoFijoTable(ActivoFijo.objects.filter(id=id))
    activo2 =ActivoFijo2Table(ActivoFijo.objects.filter(id=id))
    prestamos = PrestamosActivosFijosTable(PrestamoActivoFijo.objects.filter(activo_fijo_id=id))
    bajas = BajaActivoFijo.objects.filter(activo_fijo_id=id)
    anexos = AnexoBajaActivoFijo.objects.filter(baja_activo_id__in=bajas)
    bajas = BajaActivosFijosTable(bajas)
    request.session['idactivo'] = id
    mantenimientos = MantenimientosTable(Mantenimiento.objects.filter(activo_fijo_id=id))
    return render(request, "core/activos_fijos_detalle.html", {"activo1":activo1,"activo2":activo2,"mantenimientos":mantenimientos,"prestamos":prestamos,"bajas":bajas,"anexos":anexos})

@login_required
def ActivosFijosListaView(request):
    queryset = ActivoFijo.objects.select_related().all().order_by('nombre')
    f = ActivosFijosFilter(request.GET, queryset=queryset)
    activo1 =ActivoFijo1Table(f.qs)
    lista = []
    activos_fijos = ProponentesTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    request.session['tipo_activo'] = ''
    request.session['nombre_activo'] = ''
    request.session['UltimaPaginaVisitada']='Activos'
    RequestConfig(request, paginate={"per_page": 10, "page": 1}).configure(activo1)
    return render(request, "core/activos_fijos_lista.html", {"activo1":activo1,'filter':f})

class CreaTipoActivoFijoView(LoginRequiredMixin,CreateView):
    model = TipoActivo
    template_name = 'core/tipo_activo_form.html'
    form_class = TipoActivoFijoForm
    #success_url = reverse_lazy('lista_proveedores')

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        self.request.session['tipo_activo'] = self.object.id
        return HttpResponseRedirect(reverse('crea_activo_fijo'))

@csrf_exempt
def GuardaNombreActivoView(request):    
    nombre_activo = request.GET.get('nombre_activo', None)
    request.session['nombre_activo'] = nombre_activo
    id = 0
    data = {'id':id}
    return JsonResponse(data)  

def CargaFotoActivoFijoView(request,id):
    if request.method == 'POST':
        activo=ActivoFijo.objects.get(id=id)
        form = CargaFotoActivoFijoForm(request.POST, request.FILES,instance=activo)
        form.save()
        return redirect("activos_fijos_lista")
    else:
        form = CargaFotoActivoFijoForm()

    return render(request,'core/foto_activo_fijo_form.html', locals())

class CreaPrestamoActivoView(LoginRequiredMixin,CreateView):
    model = PrestamoActivoFijo
    template_name = 'core/prestamo_activo_form.html'
    form_class = PrestamoActivoForm
    
    def get_initial(self,*args,**kwargs):
        initial=super(CreaPrestamoActivoView,self).get_initial(**kwargs)
        idactivo = self.request.session['idactivo']
        initial['activo_fijo'] = idactivo
        return initial
    
    def get_success_url(self):
        idactivo = self.object.id
        ActivoFijo.objects.filter(id=idactivo).update(prestado=True)
        return reverse_lazy('activos_fijos_detalle',args=[idactivo])
    
class BorraPrestamoActivoFijoView(LoginRequiredMixin,DeleteView):
    model = PrestamoActivoFijo
    template_name = 'core/confirmar_borrado_registro.html'
    
    def get_success_url(self):
        idprestamo = self.object.id
        prestamo = PrestamoActivoFijo.objects.get(id=idprestamo)
        idactivo = prestamo.activo_fijo_id
        return reverse_lazy('activos_fijos_detalle',args=[idactivo])

def cabecera_activos_fijos_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))
                
def tabla_activos_fijos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = ActivoFijo.objects.filter(id__in=lista)[y1:y2]
    for i in cuerpo:
        if i.prestado == True:
            i.prestado='✔'
        else:
            i.prestado='✘'    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 650 - registros_cuerpo*16
    y = c
    cuerpo_total = ActivoFijo.objects.filter(id__in=lista)
    total = cuerpo_total.aggregate(Sum('cantidad'))['cantidad__sum']
    #Creamos una tupla de encabezados para neustra tabla
    encabezados = ['Nombre','Tipo Activo','Cantidad','Prestado','Marca','Serial','Estado']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.nombre,cuerpo.tipo_activo.descripcion,cuerpo.cantidad,cuerpo.prestado,cuerpo.marca,cuerpo.serial,cuerpo.estado) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[6 * cm, 5 * cm, 2 * cm, 2. * cm,2. * cm,2. * cm,5 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 30,y)
    y -= 10
    if pagina == ultima_pagina: 
        pdf.drawString(430, y, u"Total Activos  : "+('{:,}'.format(total)+'.00'))

def ReporteActivosFijosPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = ActivoFijo.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "ActivoFijo"    
    titulo = 'REPORTE DE ACTIVOS FIJOS'
    request.session["titulo"] = titulo
    retorno = 'reporte_activos_fijos'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def ReportePaginasPdfActivosFijosView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_activos_fijos_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_activos_fijos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response
    
def ReporteActivosFijosXlsView(request):
    activos = ActivoFijo.objects.all()
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    activos_imprimir = ActivoFijo.objects.all().order_by('tipo_activo','nombre')
    worksheet.write('A1','Nombre' )
    worksheet.write('B1','Tipo Activo' )
    worksheet.write('C1','Cantidad')
    worksheet.write('D1','Descripcion')
    worksheet.write('E1','Marca')
    worksheet.write('F1','Serial')
    worksheet.write('G1','Valor Libros')
    worksheet.write('H1','Cantidad')
    worksheet.write('I1','Prestado')
    worksheet.write('J1','Estado')
    n = 2
    for j in activos_imprimir :
        nn = str(n)
        exec("worksheet.write('A"+nn+"','"+j.nombre+"' )")
        exec("worksheet.write('B"+nn+"','"+j.tipo_activo.descripcion+"' )")
        exec("worksheet.write('C"+nn+"','"+str(j.cantidad)+"' )")
        if j.descripcion != None:
            exec("worksheet.write('D"+nn+"','"+j.descripcion+"' )")
        if j.marca != None:
            exec("worksheet.write('E"+nn+"','"+j.marca+"' )")
        if j.serial != None:
            exec("worksheet.write('F"+nn+"','"+j.serial+"' )")
        exec("worksheet.write('G"+nn+"','"+str(j.valor_libros)+"' )")
        exec("worksheet.write('H"+nn+"','"+str(j.cantidad)+"' )")
        if j.prestado == True:
            prestado = 'Si'
            exec("worksheet.write('I"+nn+"','"+prestado+"' )")
        else:
            prestado = 'No'
            exec("worksheet.write('I"+nn+"','"+prestado+"' )")
        exec("worksheet.write('J"+nn+"','"+j.estado+"' )")
        n = n + 1 

    workbook.close()
    output.seek(0)
    filename='activos_fijos.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

######################## PRESTAMO ACTIVOS FIJOS ##########################

@login_required
def PrestamosActivosListaView(request):
    queryset = PrestamoActivoFijo.objects.select_related().all().order_by('fecha')
    f = PrestamosActivosFijosFilter(request.GET, queryset=queryset)
    lista = []
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] == lista    
    prestamos =PrestamosActivosFijosTable(f.qs)
    request.session['UltimaPaginaVisitada']='Prestamos Activos'
    RequestConfig(request, paginate={"per_page": 10, "page": 1}).configure(prestamos)
    return render(request, "core/prestamos_activos_fijos_lista.html", {"prestamos":prestamos,'filter':f})

def cabecera_activos_fijos_prestamo_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))

def tabla_activos_fijos_prestamo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,idprestamo,tipo_rep):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if tipo_rep == 1:
        cuerpo = PrestamoActivoFijo.objects.filter(id__in=lista)[y1:y2]
        cuerpo_total = PrestamoActivoFijo.objects.filter(id__in=lista)
    if tipo_rep == 2:
        cuerpo = PrestamoActivoFijo.objects.filter(id=idprestamo)[y1:y2]
        cuerpo_total = PrestamoActivoFijo.objects.filter(id=idprestamo)
    for i in cuerpo:
        if i.devuelto == True:
            i.devuelto='✔'
        else:
            i.devuelto='✘'    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto para calcular el grid
    c = 650 - registros_cuerpo*16
    y = c
    total = cuerpo_total.aggregate(Sum('cantidad'))['cantidad__sum']
    #Creamos una tupla de encabezados para neustra tabla
    encabezados = ['Fecha','Activo Fijo','Responsable','Cantidad','Devuelto']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.fecha,cuerpo.activo_fijo.nombre,cuerpo.responsable,cuerpo.cantidad,cuerpo.devuelto) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[3 * cm, 5 * cm, 5 * cm, 2 * cm,2 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 30,y)
    y -= 10
    if pagina == ultima_pagina: 
        pdf.drawString(430, y, u"Total : "+('{:,}'.format(total)+'.00'))

def ReportePrestamosActivosFijosPdfView(request,id,tipo_rep):
    idprestamo = id
    request.session['idprestamo'] = idprestamo
    if tipo_rep == 1:
        request.session['tipo_rep'] = 1
        lista = request.session['lista_filtrada']
        cantidad_reg = PrestamoActivoFijo.objects.filter(id__in=lista).count()
    if tipo_rep == 2:
        request.session['tipo_rep'] = 2
        cantidad_reg = PrestamoActivoFijo.objects.filter(id=idprestamo).count()
    filas_hoja = 35
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    if n>1:
        for i in range(total_paginas):
            n = n+1
            paginas.append(n)
    else:
        n=1
        paginas.append(n)
    titulo = 'PRESTAMO ACTIVO FIJO'
    request.session["titulo"] = titulo
    retorno = 'documento_activos_fijos_prestamo'
    if n>1:
        return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno,"tipo_rep":tipo_rep})
    else:    
        return redirect('paginas_reporte_prestamo_pdf_activos_fijos',n)
    
def PaginasReportePrestamosActivosFijosPdfView(request,pagina):
    tipo_rep = request.session['tipo_rep']
    idprestamo = request.session['idprestamo']
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
    pdf = Canvas(buffer)
    cabecera_activos_fijos_prestamo_pdf(pdf,pagina,titulo)
    tabla_activos_fijos_prestamo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,idprestamo,tipo_rep)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

######################## BAJA ACTIVOS FIJOS ##############################
 
def cabecera_activos_fijos_baja_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))

def tabla_activos_fijos_baja_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,idprestamo,tipo_rep):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if tipo_rep == 1:
        cuerpo = BajaActivoFijo.objects.filter(id__in=lista)[y1:y2]
        cuerpo_total = BajaActivoFijo.objects.filter(id__in=lista)
    if tipo_rep == 2:
        cuerpo = BajaActivoFijo.objects.filter(id=idprestamo)[y1:y2]
        cuerpo_total = BajaActivoFijo.objects.filter(id=idprestamo)
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto para calcular el grid
    c = 650 - registros_cuerpo*16
    y = c
    total = cuerpo_total.aggregate(Sum('cantidad'))['cantidad__sum']
    #Creamos una tupla de encabezados para neustra tabla
    encabezados = ['Fecha','Activo Fijo','Detalle','Cantidad','Causa Baja']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.fecha,cuerpo.activo_fijo.nombre,cuerpo.detalle[0:60],cuerpo.cantidad,cuerpo.causa_baja) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[3 * cm, 5 * cm, 7 * cm, 2 * cm,2 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 30,y)
    y -= 10
    if pagina == ultima_pagina: 
        pdf.drawString(430, y, u"Total : "+('{:,}'.format(total)+'.00'))

@login_required
def BajasActivosListaView(request):
    queryset = BajaActivoFijo.objects.select_related().all().order_by('fecha')
    f = BajaActivosFijosFilter(request.GET, queryset=queryset)
    lista = []
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] == lista    
    bajas = BajaActivosFijosTable(f.qs)
    anexos = AnexoBajaActivoFijo.objects.filter(activo_fijo_id__in=lista )
    request.session['UltimaPaginaVisitada']='Baja Activos'
    RequestConfig(request, paginate={"per_page": 10, "page": 1}).configure(bajas)
    return render(request, "core/baja_activos_fijos_lista.html", {'bajas':bajas,'filter':f,'anexos':anexos})

class CreaBajaActivoView(LoginRequiredMixin,CreateView):
    model = PrestamoActivoFijo
    template_name = 'core/baja_activo_form.html'
    form_class = BajaActivoForm
    
    def get_initial(self,*args,**kwargs):
        initial=super(CreaBajaActivoView,self).get_initial(**kwargs)
        idactivo = self.request.session['idactivo']
        initial['activo_fijo'] = idactivo
        return initial
    
    def get_success_url(self):
        idactivo = self.object.id
        ActivoFijo.objects.filter(id=idactivo).update(estado='Baja')
        return reverse_lazy('activos_fijos_detalle',args=[idactivo])
    
class BorraBajaActivoFijoView(LoginRequiredMixin,DeleteView):
    model = BajaActivoFijo
    template_name = 'core/confirmar_borrado_registro.html'
    
    def get_success_url(self):
        idbaja = self.object.id
        baja = BajaActivoFijo.objects.get(id=idbaja)
        idactivo = baja.activo_fijo_id
        return reverse_lazy('activos_fijos_detalle',args=[idactivo])
    
def ReporteBajasActivosFijosPdfView(request,id,tipo_rep):
    idbaja = id
    request.session['idbaja'] = idbaja
    if tipo_rep == 1:
        request.session['tipo_rep'] = 1
        lista = request.session['lista_filtrada']
        cantidad_reg = BajaActivoFijo.objects.filter(id__in=lista).count()
    if tipo_rep == 2:
        request.session['tipo_rep'] = 2
        cantidad_reg = BajaActivoFijo.objects.filter(id=idbaja).count()
    filas_hoja = 35
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    if n>1:
        for i in range(total_paginas):
            n = n+1
            paginas.append(n)
    else:
        n=1
        paginas.append(n)
    titulo = 'BAJAS ACTIVO FIJO'
    request.session["titulo"] = titulo
    retorno = 'documento_activos_fijos_baja'
    if n>1:
        return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno,"tipo_rep":tipo_rep})
    else:    
        return redirect('paginas_reporte_bajas_pdf_activos_fijos',n)
    
def PaginasReporteBajasActivosFijosPdfView(request,pagina):
    tipo_rep = request.session['tipo_rep']
    idbaja = request.session['idbaja']
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
    pdf = Canvas(buffer)
    cabecera_activos_fijos_baja_pdf(pdf,pagina,titulo)
    tabla_activos_fijos_baja_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,idbaja,tipo_rep)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response
     
###################################### MANTENIMIENTOS ##############################

@login_required
def MantenimientosListaView(request):
    request.session['UltimaPaginaVisitada']='Mantenimientos'
    queryset = Mantenimiento.objects.select_related().all().order_by('fecha')
    f = MantenimientosFilter(request.GET, queryset=queryset)
    lista = []
    mantenimientos = MantenimientosTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 25, "page": 1}).configure(mantenimientos)
    return render(request, "core/mantenimientos_lista.html", {'mantenimientos':mantenimientos,'filter': f})    

@login_required
def MantenimientosDetalleView(request,id):
    request.session['UltimaPaginaVisitada']='Mantenimientos'
    mantenimientos = Mantenimiento.objects.filter(id=id).order_by('-fecha')
    anexos = AnexoMantenimiento.objects.filter(mantenimiento_id=id)
    return render(request, "core/mantenimientos_detalle.html", {"mantenimientos":mantenimientos,'anexos':anexos})

def MantenimientosRedireccionarView(request):
    if request.session['UltimaPaginaVisitada'] == 'Activos':
        return redirect('activos_fijos_lista')
    else:
        if request.session['UltimaPaginaVisitada'] == 'Mantenimientos':
            return redirect('lista_mantenimientos')
                    
def ActivoMantenimientos(request):
    idactivo = request.session['idactivo']
    actualiza_mantenimiento_activo_entra(idactivo)
    return redirect('activo_fijo_uno',idactivo) 
    
class CreaMantenimientoView(LoginRequiredMixin,CreateView):
    model = Mantenimiento
    template_name = 'core/mantenimiento_form.html'
    form_class = MantenimientoForm
    success_url = reverse_lazy('lista_mantenimientos')
    
    def get_success_url(self):
        if self.request.session['UltimaPaginaVisitada'] == 'Activos':
            return reverse_lazy('activos_fijos_lista')
        else:
            if self.request.session['UltimaPaginaVisitada'] == 'Mantenimientos':    
                return reverse_lazy('lista_mantenimientos')

class EditaMantenimientoView(LoginRequiredMixin,UpdateView):
    model = Mantenimiento
    fields = ['fecha','activo_fijo','proveedor','descripcion','contrato','calificacion','terminado']
    template_name = 'core/mantenimiento_form.html'
    #success_url = reverse_lazy('activo_mantenimientos')
    
    def get_success_url(self):
        idmantenimiento = self.object.id
        mantenimiento = Mantenimiento.objects.get(id=idmantenimiento)
        idmantenimiento = mantenimiento.id
        actualiza_mantenimiento_activo_entra(idmantenimiento)
        return reverse_lazy('mantenimientos_detalle',args=[idmantenimiento])

class BorraMantenimientoView(LoginRequiredMixin,DeleteView):
    model = Mantenimiento
    template_name = 'core/confirmar_borrado_registro.html'
    #success_url = reverse_lazy('activo_mantenimientos')

    def get_success_url(self):
        idmantenimiento = self.object.id
        mantenimiento = Mantenimiento.objects.get(id=idmantenimiento)
        activo = ActivoFijo.objects.get(id=mantenimiento.activo_fijo_id)
        idactivo = activo.id
        return reverse_lazy('activos_fijos_uno',args=[idactivo])

def cabecera_mantenimientos_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))
                
def tabla_mantenimientos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    #lista = request.session['lista_filtrada'] 
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Mantenimiento.objects.filter(id__in=lista)[y1:y2]
    for i in cuerpo:
        if i.terminado == True:
            i.terminado ='✔'
        else:
            i.terminado ='✘'    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    cuerpo_total = Mantenimiento.objects.filter(id__in=lista)
    #total = cuerpo_total.aggregate(Sum('cantidad'))['cantidad__sum']
    #Creamos una tupla de encabezados para neustra tabla
    encabezados = ['Fecha','Activo Fijo','Proveedor','Descripción','Contrato','Calificación','Terminado']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.fecha,cuerpo.activo_fijo.nombre,cuerpo.proveedor,cuerpo.descripcion[0:50],cuerpo.contrato,cuerpo.calificacion,cuerpo.terminado) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,4.5 * cm,3 * cm,5.5 * cm,3.5 * cm,1.2* cm,1.2 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 6),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 8,y)
    y -= 10
    
def ReporteMantenimientosPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Mantenimiento.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Mantenimiento"    
    titulo = 'REPORTE DE MANTENIMIENTOS'
    request.session["titulo"] = titulo
    retorno = 'reporte_mantenimientos'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteMantenimientosPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_mantenimientos_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_mantenimientos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

class BorraAnexoMantenimientoView(LoginRequiredMixin,DeleteView):
    model = AnexoMantenimiento
    success_url = reverse_lazy('lista_mantenimientos')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idmantenimiento = self.object.mantenimiento_id
        mantenimiento = Mantenimiento.objects.get(id=idmantenimiento)
        return reverse_lazy('mantenimientos_detalle',args=[mantenimiento.id])

def ReporteMantenimientosXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    mantenimientos_imprimir = Mantenimiento.objects.all().order_by('fecha')
    worksheet.write('A1','Fecha' )
    worksheet.write('B1','Activo Fijo' )
    worksheet.write('C1','Proveedor')
    worksheet.write('D1','Descripción')
    worksheet.write('E1','Contrato')
    worksheet.write('F1','Calificación')
    worksheet.write('G1','Terminado')
    n = 2
    
    for j in mantenimientos_imprimir :
        nn = str(n)
        sfecha= j.fecha.strftime("%m/%d/%Y")
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        exec("worksheet.write('B"+nn+"','"+j.activo_fijo.nombre+"' )")
        exec("worksheet.write('C"+nn+"','"+j.proveedor.nombre+"' )")
        exec("worksheet.write('D"+nn+"','"+j.descripcion+"' )")
        exec("worksheet.write('E"+nn+"','"+j.contrato.numero+"' )")
        exec("worksheet.write('F"+nn+"','"+str(j.calificacion)+"' )")
        if j.terminado == True:
            exec("worksheet.write('G"+nn+"','Si' )")
        else:
            exec("worksheet.write('G"+nn+"','No' )")                    
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='mantenimientos.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response

def actualiza_mantenimiento_activo(sender, instance, **kwargs):
    idactivo = instance.activo_fijo_id
    fecha = instance.fecha
    activo = ActivoFijo.objects.get(id=idactivo)
    mantenimiento = activo.mantenimiento
    if activo.mantenimiento == 'SI':
        fecha_ultimo_mantenimiento = activo.ultimo_mantenimiento
        frecuencia = activo.frecuencia_mantenimiento
        fecha_proximo_mantenimiento = fecha + timedelta(frecuencia*30.4167)
        print(fecha_proximo_mantenimiento)
        dias = date.today()-fecha_proximo_mantenimiento
        print('dias :'+str(dias))
        dias_dif = dias.days
        if  dias_dif >0:
            dias_sin_mant = dias_dif
            meses_sin_mant = dias_dif/30
        else:
            dias_sin_mant = 0
            meses_sin_mant = 0
        if Mantenimiento.objects.filter(activo_fijo_id=idactivo).count()>0:
            mantenimiento_ultimo = Mantenimiento.objects.filter(activo_fijo_id=idactivo).latest('id')
            ActivoFijo.objects.filter(id=idactivo).update(ultimo_mantenimiento=mantenimiento_ultimo.fecha,dias_sin_mantenimiento=dias_sin_mant,meses_sin_mantenimiento=meses_sin_mant)
        else:
            ActivoFijo.objects.filter(id=idactivo).update(ultimo_mantenimiento=None) 
    
post_save.connect(actualiza_mantenimiento_activo, sender=Mantenimiento)
post_delete.connect(actualiza_mantenimiento_activo, sender=Mantenimiento)


def actualiza_mantenimiento_activo_entra(idactivo):
    if Mantenimiento.objects.filter(activo_fijo_id=idactivo).count()>0:
        mantenimiento = Mantenimiento.objects.filter(activo_fijo_id=idactivo).latest('fecha') 
        fecha = mantenimiento.fecha
        activo = ActivoFijo.objects.get(id=idactivo)
        if activo.mantenimiento == 'SI':
            frecuencia = activo.frecuencia_mantenimiento
            fecha_proximo_mantenimiento = fecha + timedelta(frecuencia*30.4167)
            dias = date.today()-fecha_proximo_mantenimiento
            dias_dif = dias.days
            if  dias_dif >0:
                dias_sin_mant = dias_dif
                meses_sin_mant = dias_dif/30
            else:
                dias_sin_mant = 0
                meses_sin_mant = 0
            if Mantenimiento.objects.filter(activo_fijo_id=idactivo).count()>0:
                mantenimiento_ultimo = Mantenimiento.objects.filter(activo_fijo_id=idactivo).latest('fecha')
                ActivoFijo.objects.filter(id=idactivo).update(ultimo_mantenimiento=mantenimiento_ultimo.fecha,dias_sin_mantenimiento=dias_sin_mant,meses_sin_mantenimiento=meses_sin_mant)
            else:
                ActivoFijo.objects.filter(id=idactivo).update(ultimo_mantenimiento=None) 


####################################CONTRATOS#################################################

@login_required
def ContratosListaView(request):
    request.session['numero'] = 0
    request.session['fecha_contrato'] = 0
    queryset = Contrato.objects.select_related().all().order_by('numero')
    f = ContratosFilter(request.GET, queryset=queryset)
    lista = []
    contratos = ContratosTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 25, "page": 1}).configure(contratos)
    return render(request, "core/contratos_lista.html", {'contratos':contratos,'filter': f})

@login_required
def ContratosDetalleView(request,id):
    request.session['idcontrato'] = id
    contratos = Contrato.objects.filter(id=id)
    anexos = AnexoContrato.objects.filter(contrato_id=id)
    return render(request, "core/contratos_detalle.html", {"contratos":contratos,"anexos":anexos})


class CreaContratoView(LoginRequiredMixin,CreateView):
    model = Contrato
    template_name = 'core/contrato_form.html'
    form_class = ContratoForm
    success_url = reverse_lazy('lista_contratos')

    def get_initial(self):
        initial = super(CreaContratoView, self).get_initial()
        initial['numero'] = self.request.session['numero']
        initial['fecha_contrato'] = self.request.session['fecha_contrato']
        return initial
    
        
class EditaContratoView(LoginRequiredMixin,UpdateView):
    model = Contrato
    fields = ['numero','proveedor','tipo_contrato','objeto','fecha_contrato','valor','valor_anticipo','descripcion','activo','vigencia']
    success_url = reverse_lazy('contratos')
    template_name = 'core/contrato_form.html'

    def get_success_url(self):
        idcontrato = self.object.id
        return reverse_lazy('contratos_detalle',args=[idcontrato])

class BorraContratoView(LoginRequiredMixin,DeleteView):
    model = Contrato
    success_url = reverse_lazy('contratos')
    template_name = 'core/confirmar_borrado_registro.html'

class PaginaProveedoresView(TemplateView):
    template_name = 'core\pagina_proveedores.html'

def cabecera_contratos_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))

def tabla_contratos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Contrato.objects.filter(id__in=lista)[y1:y2]
    for i in cuerpo:
        if i.activo == True:
            i.activo ='✔'
        else:
            i.activo ='✘'
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    cuerpo_total = Contrato.objects.filter(id__in=lista)            
    encabezados = ['Número','Proveedor','Objeto','Fecha','Descripción','valor','Valor Ant.','activo']  
    largo = 38
    for j in cuerpo:
        k = 1
        segmentos = int(len(j.descripcion)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        for i in x:
            if i == 0:
                inicial = 0
            else:
                inicial = (final)
            final = (inicial + largo)
            if i == 0:
                detalles = [(j.numero,j.proveedor if j.proveedor.nombre==None else j.proveedor.nombre[0:20],j.objeto,j.fecha_contrato,j.descripcion[inicial:final],formatNumber(j.valor, 2),formatNumber(j.valor_anticipo, 2),j.activo)]
                detalle = Table([encabezados] + detalles, colWidths=[2 * cm,3.5 * cm,3 * cm,2 * cm,5 *cm,2 * cm,1.5 *cm,1 * cm])
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]
                ))
                sw = 1
            else:
                encabezados = []   
                detalles = [('',j.descripcion[inicial:final],'')]
                detalle = Table(detalles, colWidths=[10.5 * cm,5 *cm,4.5 *cm])
                detalle.setStyle(TableStyle(
                [
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #La primera fila(encabezados) va a estar centrada
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                #('BOX',(sw-1,5),(8,-1),2,colors.red),
                ]
                ))
            detalle.wrapOn(pdf, 500, 600)
            detalle.drawOn(pdf,15,y)
            y -=18
        k += 1        

def ReporteContratosPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Contrato.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Proyecto"    
    titulo = 'REPORTE DE CONTRATOS'
    request.session["titulo"] = titulo
    retorno = 'reporte_contratos'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteContratosPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_contratos_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_contratos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

def ReporteContratosXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    contratos_imprimir = Contrato.objects.all().order_by('fecha_contrato')
    worksheet.write('A1','Numero' )
    worksheet.write('B1','Proveedor' )
    worksheet.write('C1','Tipo Contrato')
    worksheet.write('D1','Objeto')
    worksheet.write('E1','Fecha Contrato')
    worksheet.write('F1','Valor Contrato')
    worksheet.write('G1','Valor Anticipo')
    worksheet.write('H1','Descripción')
    worksheet.write('I1','Activo')
    worksheet.write('J1','Vigencia')
    n = 2
    
    for j in contratos_imprimir :
        nn = str(n)
        sfecha= j.fecha_contrato.strftime("%m/%d/%Y")
        exec("worksheet.write('A"+nn+"','"+j.numero+"' )")
        exec("worksheet.write('B"+nn+"','"+j.proveedor.nombre+"' )")
        exec("worksheet.write('C"+nn+"','"+j.tipo_contrato.descripcion+"' )")
        if j.objeto != None:
            exec("worksheet.write('D"+nn+"','"+j.objeto+"' )")
        if sfecha != None:
            exec("worksheet.write('E"+nn+"','"+sfecha+"' )")
        exec("worksheet.write('F"+nn+"','"+str(j.valor)+"' )")
        exec("worksheet.write('G"+nn+"','"+str(j.valor_anticipo)+"' )")
        if j.descripcion != None:
            exec("worksheet.write('H"+nn+"','"+str(j.descripcion)+"' )")
        if j.activo == True:
            exec("worksheet.write('I"+nn+"','Si' )")
        else:
            exec("worksheet.write('I"+nn+"','No' )")
        exec("worksheet.write('J"+nn+"','"+str(j.vigencia)+"' )")                        
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='contratos.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

################################# PROVEEDORES ###############################
""" def ListaProveedoresView(request):
    proveedores = ProveedoresTable(Proveedor.objects.all())
    request.session['cc_nit'] = ''
    request.session['tipo_identificacion'] = ''
    request.session['servicio_provee'] = ''
    return render(request, "core/proveedores_lista.html", {"proveedores":proveedores}) """

@login_required
def ListaProveedoresView(request):
    request.session['cc_nit'] = ''
    request.session['tipo_identificacion'] = ''
    request.session['servicio_provee'] = ''
    queryset = Proveedor.objects.all()
    f = ProveedorFilter(request.GET, queryset=queryset)
    proveedores = ProveedoresTable(f.qs)
    lista = []
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 25, "page": 1}).configure(proveedores)
    return render(request, "core/proveedores_lista.html", {'proveedores':proveedores,'filter': f})

@login_required
def ProveedorDetalleView(request,id):
    proveedor1 = Proveedores1Table(Proveedor.objects.filter(id=id))
    proveedor2 = Proveedores2Table(Proveedor.objects.filter(id=id))
    anexos = AnexoProveedor.objects.filter(proveedor_id=id)
    contratos = Contrato.objects.filter(proveedor_id=id).order_by('fecha_contrato')
    mantenimientos = Mantenimiento.objects.filter(proveedor_id=id).order_by('fecha')
    reparaciones = Reparacion.objects.filter(proveedor_id=id).order_by('fecha')
    obras = Obra.objects.filter(proveedor_id=id).order_by('fecha')
    return render(request, "core/proveedor_detalle.html", {"proveedor1":proveedor1,"proveedor2":proveedor2,"anexos":anexos,"contratos":contratos,"mantenimientos":mantenimientos,"reparaciones":reparaciones,"obras":obras,"idproveedor":id})

class CreaProveedorView(LoginRequiredMixin,CreateView):
    model = Proveedor
    template_name = 'core/proveedor_form.html'
    form_class = ProveedorForm
    #success_url = reverse_lazy('lista_proveedores')

    def get_initial(self):
        initial = super(CreaProveedorView, self).get_initial()
        initial['cc_nit'] = self.request.session['cc_nit']
        initial['tipo_identificacion'] = self.request.session['tipo_identificacion']
        initial['servicio_provee'] = self.request.session['servicio_provee']
        return initial

    def get_success_url(self):
        self.request.session['cc_nit'] = ''
        self.request.session['tipo_identificacion'] = ''
        self.request.session['servicio_provee'] = ''
        return reverse_lazy('lista_proveedores')

class EditaProveedorView(LoginRequiredMixin,UpdateView):
    model = Proveedor
    fields = ['cc_nit','tipo_identificacion','servicio_provee','nombre','telefono','celular','direccion','email','persona_contacto','telefono_contacto','celular_contacto','email_contacto','calificacion']
    success_url = reverse_lazy('lista_proveedores')
    #template_name = 'core/proveedor_form.html'

    def get_success_url(self):
        idproveedor = self.object.id
        return reverse_lazy('proveedor_detalle',args=[idproveedor])

class BorraProveedorView(LoginRequiredMixin,DeleteView):
    model = Proveedor
    success_url = reverse_lazy('lista_proveedores')
    template_name = 'core/confirmar_borrado_registro.html'


def cabecera_proveedores_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))
                
def tabla_proveedores_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Proveedor.objects.filter(id__in=lista)[y1:y2]
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    cuerpo_total = Proveedor.objects.filter(id__in=lista)
    #Creamos una tupla de encabezados para neustra tabla

    encabezados = ['CC/Nit','Servicio','Contacto','Celular','Dirección','Email','Calif.']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.cc_nit,cuerpo.servicio_provee.descripcion[0:30],cuerpo.persona_contacto if cuerpo.persona_contacto==None else cuerpo.persona_contacto[0:30],cuerpo.celular,cuerpo.direccion if cuerpo.direccion==None else cuerpo.direccion[0:35],cuerpo.email if cuerpo.email==None else cuerpo.email[0:35],cuerpo.calificacion) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,3.1 * cm,3.5 * cm,2.5 * cm,4 * cm,4 * cm,1.2 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 5),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 15,y)
    y -= 10
    
def ReporteProveedoresPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Proveedor.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Proveedor"    
    titulo = 'REPORTE DE PROVEEDORES'
    request.session["titulo"] = titulo
    retorno = 'reporte_proveedores'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteProveedoresPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_proveedores_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_proveedores_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

def ReporteProveedoresXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    proveedores_imprimir = Proveedor.objects.all().order_by('servicio_provee','nombre')
    worksheet.write('A1','Nit' )
    worksheet.write('B1','Tipo Identificación' )
    worksheet.write('C1','Servicio Provee')
    worksheet.write('D1','Nombre')
    worksheet.write('E1','Telefono')
    worksheet.write('F1','Celular')
    worksheet.write('G1','Dirección')
    worksheet.write('H1','Email')
    worksheet.write('I1','Persona Contacto')
    worksheet.write('J1','Teléfono Contacto')
    worksheet.write('K1','Celular Contacto')
    worksheet.write('L1','Email Contacto')
    worksheet.write('M1','Calificación')
    n = 2
    
    for j in proveedores_imprimir :
        nn = str(n)
        exec("worksheet.write('A"+nn+"','"+j.cc_nit+"' )")
        exec("worksheet.write('B"+nn+"','"+j.tipo_identificacion.descripcion+"' )")
        exec("worksheet.write('C"+nn+"','"+j.servicio_provee.descripcion+"' )")
        exec("worksheet.write('D"+nn+"','"+j.nombre+"' )")
        if j.telefono != None:
            exec("worksheet.write('E"+nn+"','"+j.telefono+"' )")
        if j.celular != None:
            exec("worksheet.write('F"+nn+"','"+j.celular+"' )")
        if j.direccion != None:
            exec("worksheet.write('G"+nn+"','"+j.direccion+"' )")
        if j.email != None:
            exec("worksheet.write('H"+nn+"','"+j.email+"' )")
        if j.persona_contacto != None:
            exec("worksheet.write('I"+nn+"','"+j.persona_contacto+"' )")    
        if j.telefono_contacto != None:
            exec("worksheet.write('J"+nn+"','"+j.telefono_contacto+"' )")
        if j.celular_contacto != None:
            exec("worksheet.write('K"+nn+"','"+j.celular_contacto+"' )")
        if j.email_contacto != None:
            exec("worksheet.write('L"+nn+"','"+j.email_contacto+"' )")
        exec("worksheet.write('M"+nn+"','"+str(j.calificacion)+"' )")                
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='proveedores.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    
Parametros
############################################## PROPONENTES ################################################

@login_required
def ListaProponentesView(request):
    queryset = Proponente.objects.select_related().all()
    f = ProponenteFilter(request.GET, queryset=queryset)
    lista = []
    proponentes = ProponentesTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    request.session['cc_nit'] = ''
    request.session['tipo_identificacion'] = ''
    request.session['servicio_provee'] = ''
    RequestConfig(request, paginate={"per_page": 25, "page": 1}).configure(proponentes)
    return render(request, "core/proponentes_lista.html", {'proponentes':proponentes,'filter': f})

@login_required    
def ProponentesDetalleView(request,id):
    proponente = Proponente.objects.filter(id=id)
    anexos = AnexoProponente.objects.filter(proponente_id=id)
    request.session['cc_nit'] = ''
    request.session['tipo_identificacion'] = ''
    request.session['servicio_provee'] = ''
    
    return render(request, "core/proponente_detalle.html", {"proponente":proponente,'anexos':anexos})

class CreaProponenteView(LoginRequiredMixin,CreateView):
    model = Proponente
    template_name = 'core/proponente_form.html'
    form_class = ProponenteForm
    
    def get_initial(self):
        initial = super(CreaProponenteView, self).get_initial()
        initial['cc_nit'] = self.request.session['cc_nit']
        initial['tipo_identificacion'] = self.request.session['tipo_identificacion']
        initial['servicio_provee'] = self.request.session['servicio_provee']
        return initial

    def get_success_url(self):
        self.request.session['cc_nit'] = ''
        self.request.session['tipo_identificacion'] = ''
        self.request.session['servicio_provee'] = ''
        return reverse_lazy('lista_proponentes')

class EditaProponenteView(LoginRequiredMixin,UpdateView):
    model = Proponente
    fields = ['cc_nit','tipo_identificacion','servicio_provee','nombre','telefono','celular','direccion','email','persona_contacto','telefono_contacto','celular_contacto','email_contacto','calificacion']
    success_url = reverse_lazy('lista_proponentes')
    template_name = 'core/Proponente_form.html'

    def get_success_url(self):
        idproponente = self.object.id
        return reverse_lazy('proponente_detalle',args=[idproponente])

class BorraProponenteView(LoginRequiredMixin,DeleteView):
    model = Proponente
    success_url = reverse_lazy('Lista_proponentes')
    template_name = 'core/confirmar_borrado_registro.html'


def ProponenteProveedodorView(self,id):
    proponente = Proponente.objects.get(id=id)
    identificacion = proponente.cc_nit
    if Proveedor.objects.filter(cc_nit=identificacion).exists():
        Proveedor.objects.filter(cc_nit=identificacion).update(servicio_provee=proponente.servicio_provee,nombre=proponente.nombre,telefono=proponente.telefono,
        celular=proponente.celular,direccion=proponente.direccion,email=proponente.email,persona_contacto=proponente.persona_contacto,telefono_contacto=proponente.telefono_contacto,
        celular_contacto=proponente.celular_contacto,email_contacto=proponente.email_contacto,calificacion=proponente.calificacion)
        return redirect('lista_proveedores')
    else:
        proponente = Proponente.objects.filter(id=id)
        for i in proponente:
            proveedor = Proveedor()
            proveedor.cc_nit = i.cc_nit
            proveedor.tipo_identificacion = i.tipo_identificacion
            proveedor.nombre = i.nombre
            proveedor.servicio_provee = i.servicio_provee
            proveedor.telefono = i.telefono
            proveedor.celular = i.celular
            proveedor.direccion = i.direccion
            proveedor.email= i.email
            proveedor.persona_contacto = i.persona_contacto
            proveedor.telefono_contacto = i.telefono_contacto
            proveedor.celular_contacto = i.celular_contacto
            proveedor.email_contacto =i.email_contacto
            proveedor.calificacion = i.calificacion
            proveedor.save()
            id = proveedor.id
            return redirect('proveedor_detalle',id)

def ProponenteProyectoProveedodorView(self,id):
    proponente_proyecto = ProponenteProyecto.objects.get(id=id)
    idproponente = proponente_proyecto.proponente_id
    proponente = Proponente.objects.get(id=idproponente)
    identificacion = proponente.cc_nit
    if Proveedor.objects.filter(cc_nit=identificacion).exists():
        Proveedor.objects.filter(cc_nit=identificacion).update(servicio_provee=proponente.servicio_provee,nombre=proponente.nombre,telefono=proponente.telefono,
        celular=proponente.celular,direccion=proponente.direccion,email=proponente.email,persona_contacto=proponente.persona_contacto,telefono_contacto=proponente.telefono_contacto,
        celular_contacto=proponente.celular_contacto,email_contacto=proponente.email_contacto,calificacion=proponente.calificacion)
        proveedor = Proveedor.objects.get(cc_nit=identificacion)
        return redirect('lista_proveedores')
        #return HttpResponseRedirect(reverse('proveedor_detalle',proveedor.id))
    else:
        proponente = Proponente.objects.filter(id=idproponente)
        for i in proponente:
            proveedor = Proveedor()
            proveedor.cc_nit = i.cc_nit
            proveedor.tipo_identificacion = i.tipo_identificacion
            proveedor.nombre = i.nombre
            proveedor.servicio_provee = i.servicio_provee
            proveedor.telefono = i.telefono
            proveedor.celular = i.celular
            proveedor.direccion = i.direccion
            proveedor.email= i.email
            proveedor.persona_contacto = i.persona_contacto
            proveedor.telefono_contacto = i.telefono_contacto
            proveedor.celular_contacto = i.celular_contacto
            proveedor.email_contacto =i.email_contacto
            proveedor.calificacion = i.calificacion
            proveedor.save()
            id = proveedor.id
        return redirect('proveedor_detalle',id)
        

def ImprimirProponentesXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    
    lista=request.session['lista_proponentes']    
    proponentes_imprimir = Proponente.objects.filter(id__in=lista).order_by('servicio_provee','nombre')
    worksheet.write('A1','Nit' )
    worksheet.write('B1','Tipo Identificación' )
    worksheet.write('C1','Servicio Provee')
    worksheet.write('D1','Nombre')
    worksheet.write('E1','Telefono')
    worksheet.write('F1','Celular')
    worksheet.write('G1','Dirección')
    worksheet.write('H1','Email')
    worksheet.write('I1','Persona Contacto')
    worksheet.write('J1','Teléfono Contacto')
    worksheet.write('K1','Celular Contacto')
    worksheet.write('L1','Email Contacto')
    worksheet.write('M1','Calificación')
    n = 2
    for j in proponentes_imprimir :
        nn = str(n)
        exec("worksheet.write('A"+nn+"','"+j.cc_nit+"' )")
        exec("worksheet.write('B"+nn+"','"+j.tipo_identificacion.descripcion+"' )")
        exec("worksheet.write('C"+nn+"','"+j.servicio_provee.descripcion+"' )")
        exec("worksheet.write('D"+nn+"','"+j.nombre+"' )")
        if j.telefono != None:
            exec("worksheet.write('E"+nn+"','"+j.telefono+"' )")
        if j.celular != None:
            exec("worksheet.write('F"+nn+"','"+j.celular+"' )")
        if j.direccion != None:
            exec("worksheet.write('G"+nn+"','"+j.direccion+"' )")
        if j.email != None:
            exec("worksheet.write('H"+nn+"','"+j.email+"' )")
        if j.persona_contacto != None:
            exec("worksheet.write('I"+nn+"','"+j.persona_contacto+"' )")    
        if j.telefono_contacto != None:
            exec("worksheet.write('J"+nn+"','"+j.telefono_contacto+"' )")
        if j.celular_contacto != None:
            exec("worksheet.write('K"+nn+"','"+j.celular_contacto+"' )")
        if j.email_contacto != None:
            exec("worksheet.write('L"+nn+"','"+j.email_contacto+"' )")
        exec("worksheet.write('M"+nn+"','"+str(j.calificacion)+"' )")                
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='proponentes.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    
    return response    

def cabecera_proponentes_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))
                
def tabla_proponentes_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Proponente.objects.filter(id__in=lista)[y1:y2]
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    cuerpo_total = Proponente.objects.filter(id__in=lista)
    #Creamos una tupla de encabezados para neustra tabla
    encabezados = ['CC/Nit','Servicio','Contacto','Celular','Dirección','Email','Calif.']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.cc_nit,cuerpo.servicio_provee.descripcion[0:20],cuerpo.persona_contacto if cuerpo.persona_contacto == None else cuerpo.persona_contacto[0:25],cuerpo.celular,cuerpo.direccion[0:35],cuerpo.email if cuerpo.email==None else cuerpo.email[0:30] ,cuerpo.calificacion) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[2 * cm,2.8 * cm,4 * cm,2 * cm,4 * cm,4 * cm,1 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 6),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 20,y)
    y -= 10
    
def ReporteProponentesPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Proponente.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Proponente"    
    titulo = 'REPORTE DE PROPONENTES'
    request.session["titulo"] = titulo
    retorno = 'reporte_proponentes'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteProponentesPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_proponentes_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_proponentes_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

class CreaServicioView(LoginRequiredMixin,CreateView):
    model = ServicioProveedor
    template_name = 'core/crear_servicio.html'
    form_class = ServicioProveedorForm
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        self.request.session['servicio_provee'] = self.object.id
        return HttpResponseRedirect(reverse('crea_proveedor'))

class CreaTipoProyectoView(LoginRequiredMixin,CreateView):
    model = TipoProyecto
    template_name = 'core/tipo_proyecto_form.html'
    form_class = TipoProyectoForm
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        self.request.session['servicio_provee'] = self.object.id
        return HttpResponseRedirect(reverse('crear_proyecto'))

class CreaTipoContratoView(LoginRequiredMixin,CreateView):
    model = TipoContrato
    template_name = 'core/tipo_contrato_form.html'
    form_class = TipoContratoForm
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        self.request.session['tipo_contrato'] = self.object.id
        return HttpResponseRedirect(reverse('crea_contrato'))

@csrf_exempt
def GuardaIdentificacionView(request):
    cc_nit = request.GET.get('cc_nit', None)
    request.session['cc_nit'] = cc_nit
    if Proponente.objects.filter(cc_nit=cc_nit).exists():
        exists = True
    else:
        exists = False    
    data = {'exists':exists}
    return JsonResponse(data)

@csrf_exempt
def GuardaTipoIdentificacionView(request):    
    tipo_identificacion = request.GET.get('tipo_identificacion', None)
    request.session['tipo_identificacion'] = tipo_identificacion
    id = 0
    data = {'id':id}
    return JsonResponse(data)   

@csrf_exempt
def GuardaNumeroContratoView(request):
    numero = request.GET.get('numero', None)
    request.session['numero'] = numero
    if Contrato.objects.filter(numero=numero).exists():
        exists = True
    else:
        exists = False    
    data = {'exists':exists}
    return JsonResponse(data)

@csrf_exempt
def GuardaFechaContratoView(request):    
    fecha_contrato = request.GET.get('fecha_contrato', None)
    request.session['fecha_contrato'] = fecha_contrato
    id = 0
    data = {'id':id}
    return JsonResponse(data)  

@login_required
def ListaProyectosView(request):
    request.session['fecha'] = ''
    request.session['tipo_proyecto'] = ''
    queryset = Proyecto.objects.select_related().all()
    f = ProyectoFilter(request.GET, queryset=queryset)
    lista = []
    proyectos = ProyectosTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    RequestConfig(request, paginate={"per_page": 25, "page": 1}).configure(proyectos)
    return render(request, "core/proyectos_lista.html", {'proyectos':proyectos,'filter': f})
    

class ListaProyectosFilterView(SingleTableMixin, FilterView):
    table_class = ProyectosTable
    model = Proyecto
    template_name = "core/proyectos_lista_filtro.html"
    paginate_by = 15
    filterset_class = ProyectoFilter

class AdicionaProponenteProyectoView(LoginRequiredMixin,CreateView):
    model = ProponenteProyecto
    template_name = 'core/proponente_proyecto_form.html'
    form_class = ProponenteProyectoForm
    #success_url = reverse_lazy('lista_proyecto')    

    def form_valid(self, form):
        proponente_proyecto = form.save(commit=False)
        proponente_proyecto.proyecto_id = self.request.session['idproyecto']
        proponente_proyecto.save()
        idproyecto = self.request.session['idproyecto']
        return redirect('proyecto_detalle',idproyecto)

@login_required
def ProyectoDetalleView(request,id):
    proyecto = ProyectosDetalleTable(Proyecto.objects.filter(id=id))
    proponentes = ProponenteProyecto.objects.filter(proyecto_id=id)
    anexos_proyecto = AnexoProyecto.objects.filter(proyecto_id=id)
    lproponente_proyecto = ProponenteProyecto.objects.filter(proyecto_id=id)
    anexos_proponente_proyecto = AnexoProponenteProyecto.objects.filter(proponente_proyecto_id__in=Subquery(lproponente_proyecto.values('id')))
    return render(request, "core/proyecto_detalle.html", {'proyecto':proyecto,'proponentes':proponentes,'anexos_proyecto':anexos_proyecto,'anexos_proponente_proyecto':anexos_proponente_proyecto})

def GuardaIdProyectoView(request):
    id = request.GET.get('id', None)
    request.session['idproyecto'] = id
    data = {'idproyecto':id}
    return JsonResponse(data)

class CreaProyectoView(LoginRequiredMixin,CreateView):
    model = Proyecto
    template_name = 'core/proyecto_form.html'
    form_class = ProyectosForm
    success_url = reverse_lazy('lista_proyecto')

    def get_initial(self):
        initial = super(CreaProyectoView, self).get_initial()
        initial['fecha'] = self.request.session['fecha']
        initial['tipo_proyecto'] = self.request.session['tipo_proyecto']
        return initial

    def get_success_url(self):
        self.request.session['fecha'] = ''
        self.request.session['tipo_proyecto'] = ''
        return reverse_lazy('lista_proyectos')
    
class EditaProyectoView(LoginRequiredMixin,UpdateView):
    model = Proyecto
    fields = ['fecha','tipo_proyecto','descripcion','aprobado','aprobado_por','fecha_aprobacion','presupuesto']
    template_name = 'core/proyecto_form.html'
    #success_url = reverse_lazy('lista_proyectos')

    def get_success_url(self):
        idproyecto = self.object.id
        return reverse_lazy('proyecto_detalle',args=[idproyecto])
    
class BorraProyectoView(LoginRequiredMixin,DeleteView):
    model = Proyecto
    template_name = 'core/confirmar_borrado_registro.html'
    success_url = reverse_lazy('lista_proyectos')

def cabecera_proyectos_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))


def tabla_proyectos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    #lista = request.session['lista_filtrada'] 
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Proyecto.objects.filter(id__in=lista)[y1:y2]
    for i in cuerpo:
        if i.aprobado == True:
            i.aprobado ='✔'
        else:
            i.aprobado ='✘'
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    cuerpo_total = Proyecto.objects.filter(id__in=lista)            
    sw = 0
    encabezados = ['Fecha','Tipo','Descripción','Aprobado','Aprobado Por','Fecha Aprob.','Presupuesto']  
    largo = 85
    for j in cuerpo:
        segmentos = int(len(j.descripcion)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        for i in x:
            if i == 0:
                inicial = 0
            else:
                inicial = (final)
            final = (inicial + largo)
            if i == 0:            
                detalles = [(j.fecha,j.tipo_proyecto.descripcion[0:15],j.descripcion[inicial:final],j.aprobado,j.aprobado_por,j.fecha_aprobacion.date(),formatNumber(j.presupuesto,0))]
                detalle = Table([encabezados] + detalles, colWidths=[2 * cm,2.5 * cm,9 * cm,1 * cm,2 * cm,1.8 * cm,1.5 * cm])
                encabezados = []
            else: 
                encabezados = []   
                detalles = [('','',j.descripcion[inicial:final],'','','','')]
                detalle = Table([encabezados] + detalles, colWidths=[2 * cm,2.5 * cm,9 * cm,1 * cm,2 * cm,1.8 * cm,1.5 * cm])
            if i==0 :
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]
                ))
            else:
                detalle.setStyle(TableStyle(
                [
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]
                ))    
            detalle.wrapOn(pdf, 500, 600)
            detalle.drawOn(pdf, 20,y)
            y -= 18
            sw = 1

  
def ReporteProyectosPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Proyecto.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Proyecto"    
    titulo = 'REPORTE DE PROYECTOS'
    request.session["titulo"] = titulo
    retorno = 'reporte_proyectos'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteProyectosPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    #cantidad_reg = request.session["cantidad_reg"] 
    #total_paginas = request.session["total_paginas"]
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_proyectos_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_proyectos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    #y -= 10
    #if pagina == total_paginas: 
    #    pdf.drawString(430, y, u"Total Activos      : "+('{:,}'.format(total_activos)+'.00'))
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

def ReporteProyectosXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    proyectos_imprimir = Proyecto.objects.all().order_by('fecha')
    worksheet.write('A1','Fecha' )
    worksheet.write('B1','Tipo Proyecto' )
    worksheet.write('C1','Descripción')
    worksheet.write('D1','Aprobado')
    worksheet.write('E1','Fecha Aprobación')
    worksheet.write('F1','Presupuesto')
    n = 2
    
    for j in proyectos_imprimir :
        nn = str(n)
        sfecha = j.fecha.strftime("%m/%d/%Y")
        sfecha_aprobacion = j.fecha_aprobacion.strftime("%m/%d/%Y")
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        exec("worksheet.write('B"+nn+"','"+j.tipo_proyecto.descripcion+"' )")
        exec("worksheet.write('C"+nn+"','"+j.descripcion+"' )")
        if j.aprobado == True:
            exec("worksheet.write('D"+nn+"','Si' )")
        else:
            exec("worksheet.write('D"+nn+"','No' )")          
        if sfecha_aprobacion != '':
            exec("worksheet.write('E"+nn+"','"+sfecha_aprobacion+"' )")
        exec("worksheet.write('F"+nn+"','"+str(j.presupuesto)+"' )")
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='proyectos.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

class BorraAnexoProyectoView(LoginRequiredMixin,DeleteView):
    model = AnexoProyecto
    #success_url = reverse_lazy('lista_proyectos')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idProyecto = self.object.proyecto_id
        proyecto = Proyecto.objects.get(id=idProyecto)
        return reverse_lazy('proyecto_detalle',args=[proyecto.id])

class BorraAnexoProponenteProyectoView(LoginRequiredMixin,DeleteView):
    model = AnexoProponenteProyecto
    #success_url = reverse_lazy('lista_proyectos')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idProponenteProyecto = self.object.id
        proponente_proyecto = ProponenteProyecto.objects.get(id=idProponenteProyecto)
        idproponente = proponente_proyecto.proyecto_id
        return reverse_lazy('proyecto_detalle',args=[idproponente])

def ajaxSwitchAprobadoProyectoView(request):
    id = request.GET.get('id', None)
    proyecto = Proyecto.objects.get(id=id)
    if proyecto.aprobado == True:
        aprobado = False
    else:
        aprobado = True
    Proyecto.objects.filter(id=id).update(aprobado=aprobado,fecha_aprobacion=datetime.now())    
    data = {'aprobado':aprobado}
    return JsonResponse(data) 

class CreaProponenteProyectoView(LoginRequiredMixin,CreateView):
    model = ProponenteProyecto
    template_name = 'core/proponente_proyecto_form.html'
    form_class = ProponenteProyectoForm
    success_url = reverse_lazy('lista_proyectos')
    
    def get_success_url(self):
        idProyecto = self.object.id
        proyecto = Proyecto.objects.get(Proyecto_id=idproyecto)
        idproyecto = proyecto.id
        return reverse_lazy('proyecto_detalle',args=[idproyecto])

class BorraProponenteProyectoView(LoginRequiredMixin,DeleteView):
    model = ProponenteProyecto
    #success_url = reverse_lazy('lista_proyectos')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idProponente_proyecto = self.object.id
        proponente_proyecto = ProponenteProyecto.objects.get(id=idProponente_proyecto)
        proyecto = Proyecto.objects.get(id=proponente_proyecto.proyecto_id)
        idproyecto = proyecto.id
        return reverse_lazy('proyecto_detalle',args=[idproyecto])

class EditaProponenteProyectoView(LoginRequiredMixin,UpdateView):
    model = ProponenteProyecto
    fields = ['proponente','fecha','descripcion','valor','seleccionado','fecha_seleccion','votos_favor','votos_contra']
    template_name = 'core/proponente_proyecto_form.html'
    #success_url = reverse_lazy('lista_proyectos')
    
    def get_success_url(self):
        idProponente_proyecto = self.object.id
        proponente_proyecto = ProponenteProyecto.objects.get(id=idProponente_proyecto)
        proyecto = Proyecto.objects.get(id=proponente_proyecto.proyecto_id)
        idproyecto = proyecto.id
        return reverse_lazy('proyecto_detalle',args=[idproyecto])
    
""" class BorraAnexoProponenteProyectoView(LoginRequiredMixin,DeleteView):
    model = AnexoProponenteProyecto
    success_url = reverse_lazy('lista_proyectos')
    template_name = 'core/confirmar_borrado_registro.html' """

def ajaxSwitchSeleccionadoProponenteView(request):
    id = request.GET.get('id', None)
    proponente = ProponenteProyecto.objects.get(id=id)
    if proponente.seleccionado == True:
        seleccionado = False
    else:
        seleccionado = True
    ProponenteProyecto.objects.filter(id=id).update(seleccionado=seleccionado,fecha_seleccion=datetime.now())    
    data = {'seleccionado':seleccionado}
    return JsonResponse(data) 


######################################### OBRAS ####################################

@login_required
def ListaObrasView(request):
    queryset = Obra.objects.select_related().all().order_by('fecha')
    f = ObraFilter(request.GET, queryset=queryset)
    lista = []
    obras =ObrasTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    return render(request, "core/obras_lista.html", {'obras':obras,'filter': f})        

@login_required
def ObraDetalleView(request,id):
    request.session['idobra'] = 0
    obra = Obra.objects.filter(id=id)
    avance_obra = AvanceObra.objects.filter(obra_id=id)
    anexos_obra = AnexoObra.objects.filter(obra_id__in=Subquery(obra.values('id')))
    anexos_avance_obra = AnexoAvanceObra.objects.filter(avance_obra_id__in=Subquery(avance_obra.values('id')))
    return render(request, "core/obra_detalle.html", {'obra':obra,'avance_obra':avance_obra,'anexos_obra':anexos_obra,'anexos_avance_obra':anexos_avance_obra})

class CreaObrasView(LoginRequiredMixin,CreateView):
    model = Obra
    template_name = 'core/obra_form.html'
    form_class = ObrasForm
    success_url = reverse_lazy('lista_obras')
    
class EditaObrasView(LoginRequiredMixin,UpdateView):
    model = Obra
    fields = ['fecha','proveedor','interventor','descripcion','contrato','valor','valor_anticipo','valor_pagado','saldo_pagar','avance_obra','fecha_terminacion','calificacion','terminada']
    template_name = 'core/obra_form.html'
    success_url = reverse_lazy('lista_obras')
    
    def get_success_url(self):
        idobra = self.object.id
        obra = Obra.objects.get(id=idobra)
        idobra = obra.id
        if AvanceObra.objects.filter(obra_id=idobra).exists():
            avances = AvanceObra.objects.filter(obra_id=idobra)
            suma_porc_avances = 0
            suma_valor_avances = 0
            for avance in avances:
                suma_porc_avances = suma_porc_avances + avance.porcentaje_avance
                suma_valor_avances = suma_valor_avances + avance.valor
            saldo_pagar = obra.valor - obra.valor_anticipo -suma_valor_avances 
            Obra.objects.filter(id=idobra).update(avance_obra=suma_porc_avances, valor_pagado = suma_valor_avances + obra.valor_anticipo, saldo_pagar = saldo_pagar)
        else:
            saldo_pagar = obra.valor - obra.valor_anticipo
            Obra.objects.filter(id=idobra).update(avance_obras=0,saldo_pagar=saldo_pagar) 
        
        return reverse_lazy('obra_detalle',args=[idobra])

    def get_initial(self,*args,**kwargs):
        idobra=self.kwargs['pk']
        obra = Obra.objects.get(id=idobra) 
        if AvanceObra.objects.filter(obra_id=idobra).exists():
            avances = AvanceObra.objects.filter(obra_id=idobra)
            suma_porc_avances = 0
            suma_valor_avances = 0
            for avance in avances:
                suma_porc_avances = suma_porc_avances + avance.porcentaje_avance
                suma_valor_avances = suma_valor_avances + avance.valor
            saldo_pagar = obra.valor - obra.valor_anticipo -suma_valor_avances 
            #Obra.objects.filter(id=idobra).update(avance_obra=suma_porc_avances, valor_pagado = suma_valor_avances + obra.valor_anticipo, saldo_pagar = saldo_pagar)        
        initial=super(EditaObrasView,self).get_initial(**kwargs)
        initial['valor_pagado'] =  suma_valor_avances + obra.valor_anticipo
        initial['avance_obra']  = suma_porc_avances
        initial['saldo_pagar']  = saldo_pagar

        return initial    

class BorraObrasView(LoginRequiredMixin,DeleteView):
    model = Obra
    template_name = 'core/confirmar_borrado_registro.html'
    success_url = reverse_lazy('lista_obras')

def cabecera_obras_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 540, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,580, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,550 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,550 , u"Página No. : "+str(pagina))


def tabla_obras_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    #lista = request.session['lista_filtrada'] 
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Obra.objects.filter(id__in=lista)[y1:y2]
    for i in cuerpo:
        if i.terminada == True:
            i.terminada ='✔'
        else:
            i.terminada ='✘'
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 500 - registros_cuerpo*16
    y = c
    cuerpo_total = Obra.objects.filter(id__in=lista)            
    sw = 0
    encabezados = ['Fecha','Proveedor','Interventor','Descripción','Contrato','Valor','Anticipo','Saldo Pagar','Avance Obra','Fecha Term.','Calificación','Terminada']  
    largo = 59
    for j in cuerpo:
        segmentos = int(len(j.descripcion)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        for i in x:
            if i == 0:
                inicial = 0
            else:
                inicial = (final)
            final = (inicial + largo)
            if i == 0:            
                detalles = [(j.fecha,j.proveedor.nombre[0:15],j.interventor.nombre[0:15],j.descripcion[inicial:final],j.contrato.numero[0:10],formatNumber(j.valor,0),formatNumber(j.valor_anticipo,0),formatNumber(j.saldo_pagar,0),j.avance_obra,j.fecha_terminacion,j.calificacion,j.terminada)]
                detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,2.5 * cm,2.5 * cm,8 * cm,1.5 * cm,1.6 * cm,1.6 * cm,1.6 *cm,1.6 *cm,1.5 *cm,1.5 *cm,1.5 * cm])
            else: 
                encabezados = []   
                detalles = [('','','',j.descripcion[inicial:final],'','','','','','','','')]
                detalle = Table(detalles, colWidths=[1.5 * cm,2.5 * cm,2.5 * cm,8 * cm,1.5 * cm,1.6 * cm,1.6 * cm,1.6 *cm,1.6 *cm,1.5 *cm,1.5 *cm,1.5 * cm])
            if i==0 :
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]
                ))
            else:
                detalle.setStyle(TableStyle(
                [
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                ))    
            detalle.wrapOn(pdf, 500, 600)
            detalle.drawOn(pdf, 30,y)
            y -= 18
            sw = 1

  
def ReporteObrasPdfView(request):
    filas_hoja = 25
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Obra.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Obras"    
    titulo = 'REPORTE DE OBRAS'
    request.session["titulo"] = titulo
    retorno = 'reporte_obras'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteObrasPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    #cantidad_reg = request.session["cantidad_reg"] 
    #total_paginas = request.session["total_paginas"]
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #pdf.setPageSize((28*cm, 30*cm))

    from reportlab.lib.pagesizes import letter, landscape
    pdf.setPageSize(landscape(letter))
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_obras_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_obras_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

class BorraAnexoObraView(LoginRequiredMixin,DeleteView):
    model = AnexoObra
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idobra = self.object.obra_id
        obra = Obra.objects.get(id=idobra)
        return reverse_lazy('obra_detalle',args=[obra.id])

class CreaAvanceObraView(LoginRequiredMixin,CreateView):
    model = AvanceObra
    template_name = 'core/avance_obra_form.html'
    form_class = AvanceObrasForm
    #success_url = reverse_lazy('lista_obras')

    def get_success_url(self):
        idavance = self.object.id
        avance = AvanceObra.objects.get(id=idavance)
        obra = Obra.objects.get(id=avance.obra_id)
        idobra = obra.id
        if AvanceObra.objects.filter(obra_id=avance.obra_id).exists():
            avances = AvanceObra.objects.filter(obra_id=idobra)
            suma_porc_avances = 0
            suma_valor_avances = 0
            for avance in avances:
                suma_porc_avances = suma_porc_avances + avance.porcentaje_avance
                suma_valor_avances = suma_valor_avances + avance.valor
            saldo_pagar = obra.valor - obra.valor_anticipo -suma_valor_avances 
            Obra.objects.filter(id=idobra).update(avance_obra=suma_porc_avances, valor_pagado = suma_valor_avances + obra.valor_anticipo, saldo_pagar = saldo_pagar)
        else:
            por_avance = self.object.porcentaje_avance
            Obra.objects.filter(id=idobra).update(avance_obra=por_avance)
        
        return reverse_lazy('obra_detalle',args=[idobra])

    def get_initial(self,*args,**kwargs):
        idobra=self.kwargs['id'] 
        obra = Obra.objects.get(id=idobra) 
        initial=super(CreaAvanceObraView,self).get_initial(**kwargs)
        initial['obra']= obra.id
        return initial    

def GuardaIdObraView(request):
    id = request.GET.get('id', None)
    request.session['idobra'] = id
    data = {'idobra':id}
    return JsonResponse(data)

class EditaAvanceObraView(LoginRequiredMixin,UpdateView):
    model = AvanceObra
    fields = ['fecha','descripcion','valor','porcentaje_avance']
    template_name = 'core/avance_obra_form.html'
    #success_url = reverse_lazy('obra_detalle')

    def get_success_url(self):
        idavance = self.object.id
        avance = AvanceObra.objects.get(id=idavance)
        obra = Obra.objects.get(id=avance.obra_id)
        idobra = obra.id 
        if AvanceObra.objects.filter(obra_id=idobra).exists():
            avances = AvanceObra.objects.filter(obra_id=idobra)
            suma_porc_avances = 0
            suma_valor_avances = 0
            for avance in avances:
                suma_porc_avances = suma_porc_avances + avance.porcentaje_avance
                suma_valor_avances = suma_valor_avances + avance.valor
            saldo_pagar = obra.valor - obra.valor_anticipo -suma_valor_avances 
            Obra.objects.filter(id=idobra).update(avance_obra=suma_porc_avances, valor_pagado = suma_valor_avances + obra.valor_anticipo, saldo_pagar = saldo_pagar)
        else:
            Obra.objects.filter(id=idobra).update(avance_obra=0)
        
        return reverse_lazy('obra_detalle',args=[idobra])

class BorraAvanceObraView(LoginRequiredMixin,DeleteView):
    model = AvanceObra
    template_name = 'core/confirmar_borrado_registro.html'
    #success_url = reverse_lazy('obra_detalle')

    def get_success_url(self):
        idavance = self.object.id
        avance = AvanceObra.objects.get(id=idavance)
        obra = Obra.objects.get(id=avance.obra_id)
        idobra = obra.id 
        if AvanceObra.objects.filter(obra_id=idobra).exists():
            avances = AvanceObra.objects.filter(obra_id=idobra)
            suma_porc_avances = 0
            suma_valor_avances = 0
            for avance in avances:
                suma_porc_avances = suma_porc_avances + avance.porcentaje_avance
                suma_valor_avances = suma_valor_avances + avance.valor
            saldo_pagar = obra.valor - obra.valor_anticipo -suma_valor_avances 
            Obra.objects.filter(id=idobra).update(avance_obra=suma_porc_avances, valor_pagado = suma_valor_avances + obra.valor_anticipo, saldo_pagar = saldo_pagar)
        else:
            saldo_pagar = obra.valor - obra.valor_anticipo
            Obra.objects.filter(id=idobra).update(avance_obra=0, valor_pagado = obra.valor_anticipo, saldo_pagar = saldo_pagar)
        return reverse_lazy('obra_detalle',args=[idobra])

class BorraAnexoAvanceObraView(LoginRequiredMixin,DeleteView):
    model = AnexoAvanceObra
    success_url = reverse_lazy('obra_detalle')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idanexo = self.object.id
        anexo = AnexoAvanceObra.objects.get(id=idanexo)
        avance = AvanceObra.objects.get(id=anexo.avance_obra_id)
        obra = Obra.objects.get(id=avance.obra_id)
        idobra = obra.id 
        return reverse_lazy('obra_detalle',args=[idobra])

'fecha','proveedor','interventor','descripcion','contrato','valor','valor_anticipo','valor_pagado','saldo_pagar','avance_obra','fecha_terminacion','calificacion','terminada'

def ReporteObrasXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada'] 
    obras_imprimir = Obra.objects.all().filter(id__in=lista).order_by('fecha')
    worksheet.write('A1','Fecha' )
    worksheet.write('B1','Proveedor' )
    worksheet.write('C1','Interventor')
    worksheet.write('D1','Descripción')
    worksheet.write('E1','Contrato')
    worksheet.write('F1','Valor')
    worksheet.write('G1','Valor Anticipo')
    worksheet.write('H1','Valor Pagado')
    worksheet.write('I1','Saldo Pagar')
    worksheet.write('J1','Avance Obra')
    worksheet.write('K1','Fecha Terminación')
    worksheet.write('L1','Calificación')
    worksheet.write('M1','Terminada')
    n = 2
    
    for j in obras_imprimir :
        nn = str(n)
        sfecha= j.fecha.strftime("%m/%d/%Y")
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        exec("worksheet.write('B"+nn+"','"+j.proveedor.nombre+"' )")
        exec("worksheet.write('C"+nn+"','"+j.interventor.nombre+"' )")
        exec("worksheet.write('D"+nn+"','"+j.descripcion+"' )")
        exec("worksheet.write('E"+nn+"','"+j.contrato.numero+"' )")
        exec("worksheet.write('F"+nn+"','"+str(j.valor)+"' )")
        exec("worksheet.write('G"+nn+"','"+str(j.valor_anticipo)+"' )")
        exec("worksheet.write('H"+nn+"','"+str(j.valor_pagado)+"' )")
        exec("worksheet.write('I"+nn+"','"+str(j.saldo_pagar)+"%' )")
        exec("worksheet.write('J"+nn+"','"+str(j.avance_obra)+"%' )")
        if j.fecha_terminacion != None:
            sfecha_terminacion= j.fecha_terminacion.strftime("%m/%d/%Y")    
            exec("worksheet.write('k"+nn+"','"+strftime(sfecha_terminacion)+"' )")
        exec("worksheet.write('L"+nn+"','"+str(j.calificacion)+"' )")
        if j.terminada == True:
            exec("worksheet.write('M"+nn+"','Si' )")
        else:
            exec("worksheet.write('M"+nn+"','No' )")                    
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='obras.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

######################################### REPARACIONES  ######################################

@login_required
def ListaReparacionesView(request):
    queryset = Reparacion.objects.select_related().all().order_by('fecha')
    f = ReparacionesFilter(request.GET, queryset=queryset)
    lista = []
    reparaciones =ReparacionesTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    return render(request, "core/reparaciones_lista.html", {'reparaciones':reparaciones,'filter': f})

@login_required
def ReparacionesDetalleView(request,id):
    reparaciones = Reparacion.objects.filter(id=id)
    anexos = AnexoReparacion.objects.filter(reparacion_id=id)
    return render(request, "core/reparaciones_detalle.html", {"reparaciones":reparaciones,"anexos":anexos})

class CreaReparacionesView(LoginRequiredMixin,CreateView):
    model = Reparacion
    template_name = 'core/reparacion_form.html'
    form_class = ReparacionesForm
    success_url = reverse_lazy('lista_reparaciones')
    
class EditaReparacionesView(LoginRequiredMixin,UpdateView):
    model = Reparacion
    fields = ['fecha','proveedor','descripcion','valor','observacion','calificacion','terminado']
    template_name = 'core/reparacion_form.html'
    #success_url = reverse_lazy('lista_reparaciones')

    def get_success_url(self):
        idreparacion = self.object.id
        return reverse_lazy('reparaciones_detalle',args=[idreparacion])

    
class BorraReparacionesView(LoginRequiredMixin,DeleteView):
    model = Reparacion
    template_name = 'core/confirmar_borrado_registro.html'
    success_url = reverse_lazy('lista_reparaciones')

def cabecera_reparaciones_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))
                
def tabla_reparaciones_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina):
    #lista = request.session['lista_filtrada'] 
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    cuerpo = Reparacion.objects.filter(id__in=lista)[y1:y2]
    for i in cuerpo:
        if i.terminado == True:
            i.terminado ='✔'
        else:
            i.terminado ='✘'    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    cuerpo_total = Reparacion.objects.filter(id__in=lista)
    #total = cuerpo_total.aggregate(Sum('cantidad'))['cantidad__sum']
    #Creamos una tupla de encabezados para neustra tabla
    encabezados = ['Fecha','Proveedor','Descripción','Valor','Observación','Calificación','Terminado']  
    #Creamos una lista de tuplas que van a contener a las personas
    detalles = [(cuerpo.fecha,cuerpo.proveedor,cuerpo.descripcion[0:65],cuerpo.valor,cuerpo.observacion[0:65],cuerpo.calificacion,cuerpo.terminado) for cuerpo in cuerpo]
    #Establecemos el tamaño de cada una de las columnas de la tabla
    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,4.5 * cm,5.5 * cm,1.5 * cm,5.5 * cm,1.2 * cm,1.2 * cm])
    #Aplicamos estilos a las celdas de la tabla
    detalle.setStyle(TableStyle(
        [
            #La primera fila(encabezados) va a estar centrada
            ('ALIGN',(0,0),(3,0),'CENTER'),
            #Los bordes de todas las celdas serán de color negro y con un grosor de 1
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            #El tamaño de las letras de cada una de las celdas será de 10
            ('FONTSIZE', (0, 0), (-1, -1), 5),
        ]
    ))
    # Restamos a y por cada fila de la grid
    #Establecemos el tamaño de la hoja que ocupará la tabla
    detalle.wrapOn(pdf, 500, 600)
    #Definimos la coordenada donde se dibujará la tabla
    detalle.drawOn(pdf, 8,y)
    #stotal = format(total_pedido,',d')
    y -= 10
    
def ReporteReparacionesPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = Reparacion.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Reparacion"    
    titulo = 'REPORTE DE REPARACIONES'
    request.session["titulo"] = titulo
    retorno = 'reporte_reparaciones'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteReparacionesPdfView(request,pagina):
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    #cantidad_reg = request.session["cantidad_reg"] 
    #total_paginas = request.session["total_paginas"]
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    pdf = Canvas(buffer)
    pdf.setPageSize((22*cm, 30*cm))
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_reparaciones_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_reparaciones_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina)
    #y -= 10
    #if pagina == total_paginas: 
    #    pdf.drawString(430, y, u"Total Activos      : "+('{:,}'.format(total_activos)+'.00'))
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

class BorraAnexoReparacionView(LoginRequiredMixin,DeleteView):
    model = AnexoReparacion
    #success_url = reverse_lazy('lista_reparaciones')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idreparacion = self.object.reparacion_id
        reparacion = Reparacion.objects.get(id=idreparacion)
        return reverse_lazy('reparaciones_detalle',args=[reparacion.id])

def ReporteReparacionesXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    reparaciones_imprimir = Reparacion.objects.all().order_by('fecha')
    worksheet.write('A1','Fecha' )
    worksheet.write('B1','Proveedor')
    worksheet.write('C1','Descripción')
    worksheet.write('D1','Valor')
    worksheet.write('E1','Observación')
    worksheet.write('F1','Calificación')
    worksheet.write('G1','Terminado')
    n = 2
    
    for j in reparaciones_imprimir :
        nn = str(n)
        sfecha= j.fecha.strftime("%m/%d/%Y")
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        exec("worksheet.write('B"+nn+"','"+j.proveedor.nombre+"' )")
        if j.descripcion == None:
            exec("worksheet.write('C"+nn+"','"+j.descripcion+"' )")
        exec("worksheet.write('D"+nn+"','"+str(j.valor)+"' )")
        if j.observacion == None:
            exec("worksheet.write('E"+nn+"','"+j.observacion+"' )")
        exec("worksheet.write('F"+nn+"','"+str(j.calificacion)+"' )")
        if j.terminado == True:
            exec("worksheet.write('G"+nn+"','Si' )")
        else:
            exec("worksheet.write('G"+nn+"','No' )")                    
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='reparaciones.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response

################################### CONSEJO ####################################

class PaginaConsejoView(TemplateView):
    template_name = 'core\pagina_consejo.html'

@login_required
def ListaReunionConsejoView(request):
    tipo_usuario = request.session['tipo_usuario']
    #reuniones_consejo = ReunionConsejoTable(ReunionConsejo.objects.all().order_by('-fecha'))
    #return render(request, "core/reuniones_consejo_lista.html", {"reuniones_consejo":reuniones_consejo})
    request.session['user'] = request.user.id
    request.session['detalle_reporte_reunion_consejo'] = 0
    queryset = ReunionConsejo.objects.all().order_by('-fecha')
    f = ReunionesConsejoFilter(request.GET, queryset=queryset)
    lista = []
    reuniones_consejo =f.qs
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    request.session["idreunion"] = 0
    #compromisos = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista)
    return render(request, "core/reuniones_consejo_lista.html", {"reuniones_consejo":reuniones_consejo,'filter':f,'tipo_usuario':tipo_usuario})

class CreaReunionConsejoView(LoginRequiredMixin,CreateView):
    model = ReunionConsejo
    template_name = 'core/reunion_consejo_form.html'
    form_class = ReunionConsejoForm
    success_url = reverse_lazy('lista_reuniones_consejo')
    
    def get_form(self):
        form = super().get_form()
        form.fields["hora_inicio"].widget = DateTimePickerInput()
        form.fields["hora_final"].widget = DateTimePickerInput()
        return form
    
class EditaReunionConsejoView(LoginRequiredMixin,UpdateView):
    model = ReunionConsejo
    template_name = 'core/reunion_consejo_form.html'
    form_class = ReunionConsejoForm
    
    def get_success_url(self):
        idreunion = self.object.id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])
    
class BorraReunionConsejoView(LoginRequiredMixin,DeleteView):
    model = ReunionConsejo
    template_name = 'core/confirmar_borrado_registro.html'
    success_url = reverse_lazy('lista_reuniones_consejo')


class AplicaCompromisoConsejoView(LoginRequiredMixin,UpdateView):
    model = CompromisoConsejo
    fields = ['cumplido','fecha_cumplido']
    template_name = 'core/compromiso_consejo_form.html'
    
    def get_success_url(self):
        tipo_usuario = self.request.session['tipo_usuario']
        idcompromiso = self.object.id
        compromiso = CompromisoConsejo.objects.get(id=idcompromiso) 
        reunion = compromiso.reunion_consejo_id
        consejo = MiembroConsejo.objects.filter(activo=True)
        asunto = "Cumplimiento compromiso consejo"
        for i in consejo:
            mensaje = "Sr. Consejero: "+i.nombre+" se le informa que la administración ha aplicado el compromiso de la reunion de consejo : "+compromiso.compromiso
            EnvioMailOtros(mensaje,asunto,i.email)
        return reverse_lazy('compromiso_detalle',args=[idcompromiso])
    
def cabecera_reunion_consejo_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(240,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))

from html.parser import HTMLParser

def tabla_reunion_consejo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,detalle,id):
    id = int(id)
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if id > 0:
        cuerpo = ReunionConsejo.objects.filter(id=id)[y1:y2]
    else:        
        cuerpo = ReunionConsejo.objects.filter(id__in=lista)[y1:y2]
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    if detalle == 0:
        c = 700 - registros_cuerpo*16
    else:        
        c = 520 - registros_cuerpo*16
    y = c
    if id > 0:
        cuerpo_total = ReunionConsejo.objects.filter(id=id)
    else:    
        cuerpo_total = ReunionConsejo.objects.filter(id__in=lista)
    if detalle == 0:
        encabezados = ['Fecha','Número Acta','Hora Inicial','Hora Final','Total Horas']
    else :
        encabezados = ['Fecha','Contenido','Acta','Hora.Inic.','Hora.Fin.','Total Hor.']    
    largo = 80
    tiempo_total = 0
    sw = 0
    for j in cuerpo:
        linea =j.contenido.strip()
        tiempo_reunion = float((datetime.strptime(str(j.hora_final), '%H:%M:%S')-datetime.strptime(str(j.hora_inicio), '%H:%M:%S')).seconds/3600.0)
        if detalle == 0:
            if sw > 0 :
                encabezados=[]
            detalles = [(j.fecha,j.numero_acta,j.hora_inicio,j.hora_final,tiempo_reunion)]
            sdetalle = Table([encabezados] + detalles, colWidths=[2 * cm,4 * cm,4 * cm,4 *cm,4 *cm])
        else:
            if sw > 0 :
                encabezados=[]
            detalles = [(j.fecha,linea,j.numero_acta,j.hora_inicio,j.hora_final,tiempo_reunion)]    
            sdetalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,14 * cm,1.0 * cm,1.3 * cm,1.3 *cm,1.3 *cm])
        sdetalle.setStyle(TableStyle(
        [
        #La primera fila(encabezados) va a estar centrada
        ('ALIGN',(0,0),(3,0),'CENTER'),
        #Los bordes de todas las celdas serán de color negro y con un grosor de 1
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        #El tamaño de las letras de cada una de las celdas será de 10
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]
        ))
        sdetalle.wrapOn(pdf, 500, 600)
        if detalle == 0:
            sdetalle.drawOn(pdf, 30,y)
            y -= 18
        else:    
            sdetalle.drawOn(pdf, 10,y)
            y -= 18
        tiempo_total  += tiempo_reunion
        sw = 1
    if detalle == 0:
        pdf.drawString(450,y, u"Tiempo Total ==> "+str(tiempo_total))
    else:    
        pdf.drawString(490,y, u"Tiempo Total ==> "+str(tiempo_total))

def tabla_compromiso_reunion_consejo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id,estado):
    id = int(id)
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if id > 0 :
        if estado == 0:    
            cuerpo = CompromisoConsejo.objects.filter(reunion_consejo_id=id,cumplido=False)[y1:y2]
        else:
            cuerpo = CompromisoConsejo.objects.filter(reunion_consejo_id=id,cumplido=True)[y1:y2]    
    else:
        if estado == 0:
            cuerpo = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista,cumplido=False)[y1:y2]
        elif estado == 1:
                cuerpo = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista,cumplido=True)[y1:y2]        
        elif estado == 2:
                   cuerpo = CompromisoConsejo.objects.filter(cumplido=False)[y1:y2]     
    for i in cuerpo:
        if i.cumplido == True:
            i.cumplido ='Si'
        else:
            i.fecha_cumplido = (date.today()-i.reunion_consejo.fecha).days
            i.cumplido ='No'    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    if id > 0 :
        if estado == 0:
            cuerpo_total = CompromisoConsejo.objects.filter(reunion_consejo_id=id,cumplido=False).select_related
        else:
            cuerpo_total = CompromisoConsejo.objects.filter(reunion_consejo_id=id,cumplido=True).select_related    
    else:
        if estado == 0:    
            cuerpo_total = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista,cumplido=False).select_related
        else:
            cuerpo_total = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista,cumplido=True).select_related    
    sw = 0
    #total = cuerpo_total.aggregate(Sum('cantidad'))['cantidad__sum']
    encabezados = ['Fecha','Compromiso','Cumplido','Días']
    largo = 120
    for j in cuerpo:
        segmentos = int(len(j.compromiso)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        if sw >= 1 :
            encabezados = [] 
        for i in x:
            if i == 0:
                inicial = 0
            else:
                inicial = (final)
            final = (inicial + largo)
            if i == 0:            
                detalles = [(j.reunion_consejo.fecha,j.compromiso[inicial:final],j.cumplido,j.fecha_cumplido)]
                detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,15 * cm,1.5 * cm,1.5 * cm])
            else: 
                detalles = [('',j.compromiso[inicial:final],'','')]
                detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,15 * cm,1.5 * cm,1.8 * cm])
            if i==0 :
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                ))
            
            detalle.wrapOn(pdf, 500, 600)
            detalle.drawOn(pdf, 15,y)
            y -= 18
            sw = 1

def tabla_decisiones_reunion_consejo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id):

    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if id > 0:    
        cuerpo = DecisionConsejo.objects.filter(reunion_consejo_id=id)[y1:y2]
    else:        
        cuerpo = DecisionConsejo.objects.filter(reunion_consejo_id__in=lista)[y1:y2]
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    if id > 0:
        cuerpo_total = DecisionConsejo.objects.filter(reunion_consejo_id=id).select_related
    else:
        cuerpo_total = DecisionConsejo.objects.filter(reunion_consejo_id__in=lista).select_related
    sw = 0
    encabezados = ['Fecha','Decision','Favor','Contra','Abstención']
    largo = 125
    for j in cuerpo:
        segmentos = int(len(j.decision)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        for i in x:
            if i == 0:
                inicial = 0
            else:
                inicial = (final)
            final = (inicial + largo)
            if i == 0:            
                detalles = [(j.reunion_consejo.fecha,j.decision[inicial:final],j.numero_votos_favor,j.numero_votos_contra,j.numero_votos_abstencion)]
                detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,14 * cm,1.5 * cm,1.5 * cm,1.5 *cm])
            else: 
                encabezados = []   
                detalles = [('',j.decision[inicial:final],'','','')]
                detalle = Table(detalles, colWidths=[1.5 * cm,14 * cm,1.5 * cm,1.5 * cm,1.5 *cm])
            if i==0 :
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                ))
            else:
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                )) 
            detalle.wrapOn(pdf, 600, 600)
            detalle.drawOn(pdf, 15,y)
            y -= 18
            sw = 1


def ReporteReunionesConsejoPdfView(request,tipo,estado):
    id = request.session["idreunion"]
    lista = request.session['lista_filtrada']
    request.session['estado'] = estado
    filas_hoja = 25
    if  tipo == 1:
        if estado == 1:
            if id > 0 :
                cantidad_reg = CompromisoConsejo.objects.filter(reunion_consejo_id=id,cumplido=True).count()
            else:    
                cantidad_reg = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista,cumplido=True).count()
        if estado == 0:
            cantidad_reg = CompromisoConsejo.objects.filter(reunion_consejo_id__in=lista,cumplido=False).count()
        if estado == 2:
            # Reporte desde la página de consejo
            cantidad_reg = CompromisoConsejo.objects.filter(cumplido=False).count()    
    elif tipo == 2:
        if id > 0 :
            cantidad_reg = DecisionConsejo.objects.filter(reunion_consejo_id=id).count()
        else:    
            cantidad_reg = DecisionConsejo.objects.filter(reunion_consejo_id__in=lista).count()
    elif tipo == 3 :
        if id > 0 :
            cantidad_reg = ReunionConsejo.objects.filter(id=id).count()
        else:        
            cantidad_reg = ReunionConsejo.objects.filter(id__in=lista).count()    
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = ""    
    if  tipo == 1:
        titulo = 'REPORTE DE COMPROMISOS CONSEJO '
        retorno = 'reporte_compromisos_reunion_consejo_pdf'    
    if tipo == 2:
        titulo = 'REPORTE DE DECISIONES CONSEJO '
        retorno = 'reporte_decisiones_reunion_consejo_pdf'
    if tipo == 3:
        titulo = 'REPORTE DE REUNIONES CONSEJO '
        retorno = 'reporte_reunion_consejo_pdf'    
    request.session["titulo"] = titulo
    if estado == 2:
        request.session["idreunion"] = id
    else:
        # Es para el reporte que se saca de pendientes en la pagina del consejo
        request.session["idreunion"] = 0    
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteReunionConsejoPdfView(request,pagina,tipo):
    estado = request.session["estado"]
    id = request.session["idreunion"]
    detalle = request.session['detalle_reporte_reunion_consejo']
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_reunion_consejo_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    if tipo == 1:
        tabla_compromiso_reunion_consejo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id,estado)
    if tipo == 2:
        tabla_decisiones_reunion_consejo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id)    
    if tipo == 3:
        tabla_reunion_consejo_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,detalle,id)        
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

class ReportePendientesConsejoPdfView(TemplateView):
    template_name = 'core/tipo_reporte_pendientes_pdf.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        self.request.session['estado'] = 0
        self.request.session['idreunion'] = 0
        self.request.session['idinforme'] = 0
        self.request.session['lista_filtrada'] = ''
        self.request.session['detalle_reporte_reunion_consejo'] = 0

def ReportePendientesRecomendacionesRevisorPdfView(request):
    return redirect('reportes_informe_revisor_pdf',tipo,estado)
    

class ReportePendientesDecisionesAsambleaPdfView(TemplateView):
    template_name = 'core/tipo_reporte_pendientes_pdf.html'

class ReportesReunionesConsejoXlsView(TemplateView):
    template_name = 'core/reportes_reuniones_consejo_xls.html'

class TipoReporteReunionesConsejoPdfView(TemplateView):
    template_name = 'core/tipo_reporte_reuniones_consejo_pdf.html'

def AjaxGuardaIdReunionView(request):
    id = request.GET.get('id', None)
    request.session["idreunion"] = int(id)
    data = {'id':id}
    return JsonResponse(data)    

def ReporteReunionesConsejoXlsView(request,tipo,estado):
    detalle = request.session['detalle_reporte_reunion_consejo']
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada']
    reunion_imprimir = ReunionConsejo.objects.filter(id__in=lista)
    worksheet.write('A1','Fecha' )
    worksheet.write('B1','Acta No.' )
    if tipo == 1:
        worksheet.write('C1','Compromiso' )
        worksheet.write('D1','Cumplido' )    
        worksheet.write('E1','Fecha Cumplido' )
    if tipo == 2:
        worksheet.write('C1','Decisión' )
        worksheet.write('D1','Votos a Favor' )
        worksheet.write('E1','Votos Contra' )
        worksheet.write('F1','Abstención' )
        worksheet.write('F1','Total Decisiones' )
    if tipo == 3:
        worksheet.write('C1','Hora Inicio' )
        worksheet.write('D1','Hora Final' )
        worksheet.write('E1','Tiempo Reunión' )
    n = 2
    p = 1
    for j in reunion_imprimir :
        nn = str(n)
        if j.fecha != None :
            sfecha= j.fecha.strftime("%m/%d/%Y")
        else:
            sfecha=''    
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        if p == 1:
            exec("worksheet.write('B"+nn+"','"+j.numero_acta+"' )")
        if tipo == 1:
            if estado == 0:
                compromiso_imprimir = CompromisoConsejo.objects.filter(reunion_consejo_id=j.id,cumplido=False)
            else:
                compromiso_imprimir = CompromisoConsejo.objects.filter(reunion_consejo_id=j.id,cumplido=True)    
            for k in compromiso_imprimir :
                exec("worksheet.write('C"+nn+"','"+k.compromiso+"' )")
                if k.cumplido == True:
                    exec("worksheet.write('D"+nn+"','Si' )")
                else:
                    exec("worksheet.write('D"+nn+"','No' )")
                if  k.fecha_cumplido !=None :      
                    sfecha_cumplido= k.fecha_cumplido.strftime("%m/%d/%Y")
                else:
                    sfecha_cumplido=''    
                exec("worksheet.write('E"+nn+"','"+sfecha_cumplido+"' )")
                n += 1
                nn = str(n)
            p += 1    
        if tipo == 2:
            decision_imprimir = DecisionConsejo.objects.filter(reunion_consejo_id=j.id)
            for k in decision_imprimir :
                exec("worksheet.write('C"+nn+"','"+k.decision+"' )")
                exec("worksheet.write('D"+nn+"','"+str(k.numero_votos_favor)+"' )")
                exec("worksheet.write('E"+nn+"','"+str(k.numero_votos_contra)+"' )")
                exec("worksheet.write('F"+nn+"','"+str(k.numero_votos_abstencion)+"' )")
                
        if tipo == 3:
            reunion_imprimir = ReunionConsejo.objects.filter(id=j.id)
            for k in reunion_imprimir :
                tiempo_reunion = float((datetime.strptime(str(k.hora_final), '%H:%M:%S')-datetime.strptime(str(k.hora_inicio), '%H:%M:%S')).seconds/3600.0)
                exec("worksheet.write('C"+nn+"','"+str(k.hora_inicio)+"' )")
                exec("worksheet.write('D"+nn+"','"+str(k.hora_final)+"' )")
                exec("worksheet.write('E"+nn+"','"+str(tiempo_reunion)+"' )")        
            n += 1    
    workbook.close()
    output.seek(0)
    if tipo == 1:
        filename='informe_compromisos_consejo.xlsx'
    else:
        filename='informe_decisiones_consejo.xlsx'    
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    


class BorraAnexoReunionConsejoView(LoginRequiredMixin,DeleteView):
    model = AnexoReunionConsejo
    success_url = reverse_lazy('lista_reuniones_consejo')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idreunion = self.object.reunion_id
        reunion = ReunionConsejo.objects.get(id=idreunion)
        return reverse_lazy('detalle_reunion_consejo',args=[reunion.id])

def ReunionConsejoContenidoView(request,id):
    reunion_consejo = ReunionConsejo.objects.filter(id=id)
    return render(request, "core/reunion_consejo_contenido.html", {'reunion_consejo':reunion_consejo})

@login_required
def DetalleReunionConsejoView(request,id):
    request.session['detalle_reporte_reunion_consejo'] = 1
    reuniones_consejo = ReunionConsejo.objects.filter(id=id)
    anexos = AnexoReunionConsejo.objects.filter(reunion_id=id)
    decisiones = DecisionConsejo.objects.filter(reunion_consejo_id=id)
    compromisos = CompromisoConsejo.objects.filter(reunion_consejo_id=id)
    return render(request, "core/reuniones_consejo_detalle.html", {"reuniones_consejo":reuniones_consejo,'anexos':anexos,'decisiones':decisiones,'compromisos':compromisos})

class CreaDecisionConsejoView(LoginRequiredMixin,CreateView):
    model = DecisionConsejo
    template_name = 'core/decision_reunion_consejo_form.html'
    form_class = DecisionConsejoForm
    #success_url = reverse_lazy('reuniones_consejo_detalle')

    def get_initial(self,*args,**kwargs):
        initial=super(CreaDecisionConsejoView,self).get_initial(**kwargs)
        reunion = ReunionConsejo.objects.get(id=self.kwargs['id']) 
        initial['reunion_consejo']=reunion.id
        return initial

    def get_success_url(self):
        idreunion = self.object.reunion_consejo_id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

class BorraDecisionConsejoView(LoginRequiredMixin,DeleteView):
    model = DecisionConsejo
    #success_url = reverse_lazy('lista_reuniones_consejo')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idreunion = self.object.reunion_consejo_id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

class BorraCompromisoConsejoView(LoginRequiredMixin,DeleteView):
    model = CompromisoConsejo
    #success_url = reverse_lazy('lista_reuniones_consejo')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idreunion = self.object.reunion_consejo_id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

class CreaCompromisoConsejoView(LoginRequiredMixin,CreateView):
    model = CompromisoConsejo
    template_name = 'core/compromiso_reunion_consejo_form.html'
    form_class = CompromisoConsejoForm
    
    def get_success_url(self):
        idreunion = self.object.reunion_consejo_id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

    def form_valid(self, form):
        if form.is_valid():
            idreunion = self.kwargs['id']
            object = form.save(commit=False)
            object.reunion_consejo_id = idreunion
            object.cumplido = False
            object.fecha_cumplido=None
            object.save()
            idreunion = self.kwargs['id']
            #return reverse_lazy('detalle_reunion_consejo',args=[idreunion])
            return super(CreaCompromisoConsejoView, self).form_valid(form)
        else:
            mensaje = "Error"
            return render(self.request, "core/mensaje_error_usuario.html", {'mensaje':mensaje})

@login_required
def CompromisoDetalleView(request,id):
    tipo_usuario = request.session['tipo_usuario']
    compromiso = CompromisoConsejo.objects.get(id=id)
    idreunion = compromiso.reunion_consejo_id
    reunion = ReunionConsejo.objects.filter(id=idreunion)
    #reuniones_consejo = ReunionConsejo.objects.filter(id=compromiso.reunion_consejo_id)
    anexos = AnexoReunionConsejo.objects.filter(reunion_id=idreunion)
    #decisiones = DecisionConsejo.objects.filter(reunion_consejo_id=reunion.id)
    compromisos = CompromisoConsejo.objects.filter(reunion_consejo_id=idreunion)
    return render(request, "core/compromiso_reunion_consejo.html", {"reunion":reunion,'anexos':anexos,'compromisos':compromisos,'tipo_usuario':tipo_usuario})

def PublicaReunionConsejoView(request,id):
    reunion = ReunionConsejo.objects.get(id=id)
    idreunion = id
    anexos = AnexoReunionConsejo.objects.filter(reunion_id=id,puiblicar=True)
    envia = request.user.username
    email = request.user.email
    subject ="De: "+envia+" Email:  "+email+" Hora: "+strftime("%Y-%m-%d %H:%M:%S", gmtime())
    message = "Acta Consejo No."+reunion.numero_acta
    from_email = settings.EMAIL_HOST_USER
    message = "Acta Consejo No."+reunion.numero_acta
    currentDateTime = datetime.datetime.now()
    consejeros = MiembroConsejo.objects.filter(envio_comunicado=True)
    con_lis = list(consejeros.email.items()) 
    listamail= con_lis
    #listamail=['jjguevara_a@hotmail.com']
    destinatarios = listamail
    
    mail = EmailMessage(
            subject,
            message,
            from_email,
            to=None,
            bcc=destinatarios,
    )
    for anexo in anexos:
        ruta_archivo = anexo.archivo
        filename = anexo.descripcion
        mail.attach_file(ruta_archivo)
        #mail.attach('pedido.pdf', pdf, 'application/pdf')
    mail.send()
    return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

@login_required
def ListaMiembrosConsejoView(request):
    miembros_consejo = MiembrosConsejoTable(MiembroConsejo.objects.all().order_by('orden'))
    return render(request, "core/miembros_consejo_lista.html", {"miembros_consejo":miembros_consejo})    

class EditaMiembroConsejoView(UpdateView):
    model = MiembroConsejo
    fields = ['interior','apartamento','nombre','cargo','email','envio','publicar','comunidad','activo','orden']
    template_name = 'core/miembro_consejo_form.html'
    success_url = reverse_lazy('lista_miembros_consejo')

def CargaFotoMiembroConsejoView(request,id):
    if request.method == 'POST':
        miembro=MiembroConsejo.objects.get(id=id)
        form = CargaFotoMiembroConsejoForm(request.POST, request.FILES,instance=miembro)
        form.save()
        return redirect("lista_miembros_consejo")
    else:
        form = CargaFotoMiembroConsejoForm()

    return render(request,'core/foto_miembro_consejo_form.html', locals())

def BorraMiembroConsejoView(request,pk):
    consejero = MiembroConsejo.objects.get(pk=pk)
    email = consejero.email
    User.objects.filter(email=email).update(is_staff=False)
    MiembroConsejo.objects.filter(pk=pk).delete()
    return redirect('lista_miembros_consejo')    

class CreaMiembroConsejoView(LoginRequiredMixin,CreateView):
    model = MiembroConsejo
    template_name = 'core/miembro_consejo_form.html'
    form_class = MiembroConsejoForm
    success_url = reverse_lazy('lista_miembros_consejo')

    def form_valid(self, form):
        if form.is_valid():
            email = form.cleaned_data['email']
            if User.objects.filter(email=email).exists():
                User.objects.filter(email=email).update(is_staff=True)
            consejo = form.save(commit=False)
            consejo.save()
        return redirect('lista_miembros_consejo')

  
class EditaDecisionConsejoView(LoginRequiredMixin,UpdateView):
    model = DecisionConsejo
    fields = ['decision','numero_votos_favor','numero_votos_contra','numero_votos_abstencion']
    template_name = 'core/miembro_consejo_form.html'
    success_url = reverse_lazy('lista_reuniones_consejo')

    def get_success_url(self):
        idreunion = self.object.reunion_consejo_id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

class EditaCompromisoConsejoView(LoginRequiredMixin,UpdateView):
    model = CompromisoConsejo
    fields = ['compromiso','cumplido','fecha_cumplido']
    template_name = 'core/compromiso_reunion_consejo_form.html'
    success_url = reverse_lazy('lista_reuniones_consejo')

    def get_success_url(self):
        idreunion = self.object.reunion_consejo_id
        return reverse_lazy('detalle_reunion_consejo',args=[idreunion])

def ajaxSwitchEnvioConsejeroView(request):
    id = request.GET.get('id', None)
    miembro_consejo = MiembroConsejo.objects.get(id=id)
    if miembro_consejo.envio == True:
        envio = False
    else:
        envio = True
    miembro_consejo = MiembroConsejo.objects.filter(id=id).update(envio=envio)    
    data = {'envio':envio}
    return JsonResponse(data)    


def ajaxSwitchActivoConsejeroView(request):
    id = request.GET.get('id', None)
    miembro_consejo = MiembroConsejo.objects.get(id=id)
    if miembro_consejo.activo == True:
        activo = False
    else:
        activo = True
    miembro_consejo = MiembroConsejo.objects.filter(id=id).update(activo=activo)
    data = {'activo':activo}
    return JsonResponse(data)    

def ajaxSwitchPublicarConsejeroView(request):
    id = request.GET.get('id', None)
    miembro_consejo = MiembroConsejo.objects.get(id=id)
    if miembro_consejo.publicar == True:
        publicar = False
    else:
        publicar = True
    miembro_consejo = MiembroConsejo.objects.filter(id=id).update(publicar=publicar)
    data = {'publicar':publicar}
    return JsonResponse(data)    

def ajaxSwitchComunidadConsejeroView(request):
    id = request.GET.get('id', None)
    miembro_consejo = MiembroConsejo.objects.get(id=id)
    if miembro_consejo.comunidad == True:
        comunidad = False
    else:
        comunidad = True
    miembro_consejo = MiembroConsejo.objects.filter(id=id).update(comunidad=comunidad)
    data = {'comunidad':comunidad}
    return JsonResponse(data)    

############################## Vigilante ##########################

@login_required
def ListaVigilantesView(request):
    vigilantes = VigilantesTable(Vigilante.objects.all())
    return render(request, "core/vigilantes_lista.html", {"vigilantes":vigilantes})

class CreaVigilanteView(LoginRequiredMixin,CreateView):
    model = Vigilante
    template_name = 'core/vigilante_form.html'
    form_class = VigilanteForm
    success_url = reverse_lazy('lista_vigilantes')

class EditaVigilanteView(LoginRequiredMixin,UpdateView):
    model = Vigilante
    fields = ['nombre','email','publicar','comunidad','foto','orden']
    template_name = 'core/vigilante_form.html'
    success_url = reverse_lazy('lista_vigilantes')

class BorraVigilanteView(LoginRequiredMixin,DeleteView):
    model = Vigilante
    success_url = reverse_lazy('lista_vigilantes')
    template_name = 'core/confirmar_borrado_registro.html'

############################### Staff ############################

@login_required
def ListaMiembrosStaffView(request):
    miembros = MiembrosStaffTable(MiembroStaff.objects.all())
    return render(request, "core/miembros_staff_lista.html", {"miembros":miembros})

class CreaMiembroStaffView(LoginRequiredMixin,CreateView):
    model = MiembroStaff
    template_name = 'core/miembro_staff_form.html'
    form_class = MiembroStaffForm
    success_url = reverse_lazy('lista_miembros_staff')

class EditaMiembroStaffView(LoginRequiredMixin,UpdateView):
    model = MiembroStaff
    fields = ['nombre','cargo','email','envio','foto','orden']
    template_name = 'core/miembro_staff_form.html'
    success_url = reverse_lazy('lista_miembros_staff')

class BorraMiembroStaffView(LoginRequiredMixin,DeleteView):
    model = MiembroStaff
    success_url = reverse_lazy('lista_miembros_staff')
    template_name = 'core/confirmar_borrado_registro.html'

def ajaxSwitchEnvioStaffView(request):
    id = request.GET.get('id', None)
    miembro_staff = MiembroStaff.objects.get(id=id)
    if miembro_staff.envio == True:
        envio = False
    else:
        envio = True
    miembro_staff = MiembroStaff.objects.filter(id=id).update(envio=envio)    
    data = {'envio':envio}
    return JsonResponse(data)    

def ajaxSwitchPublicarStaffView(request):
    id = request.GET.get('id', None)
    miembro_staff = MiembroStaff.objects.get(id=id)
    if miembro_staff.publicar == True:
        publicar = False
    else:
        publicar = True
    miembro_staff = MiembroStaff.objects.filter(id=id).update(publicar=publicar)    
    data = {'publicar':publicar}
    return JsonResponse(data)    

def ajaxSwitchComunidadStaffView(request):
    id = request.GET.get('id', None)
    miembro_staff = MiembroStaff.objects.get(id=id)
    if miembro_staff.comunidad == True:
        comunidad = False
    else:
        comunidad = True
    miembro_staff = MiembroStaff.objects.filter(id=id).update(comunidad=comunidad)    
    data = {'comunidad':comunidad}
    return JsonResponse(data)    

def CargaFotoMiembroStaffView(request,id):
    if request.method == 'POST':
        miembro=MiembroStaff.objects.get(id=id)
        form = CargaFotoMiembroStaffForm(request.POST, request.FILES,instance=miembro)
        form.save()
        return redirect("lista_miembros_staff")
    else:
        form = CargaFotoMiembroStaffForm()

    return render(request,'core/foto_miembro_staff_form.html', locals())

############################## Revisor ####################################

@login_required
def ListaInformesRevisorView(request):
    tipo_usuario = request.session['tipo_usuario']
    request.session['idinforme'] = 0
    request.session['user'] = request.user.id
    queryset = InformeRevisor.objects.all().order_by('-fecha')
    f = InformesRevisorFilter(request.GET, queryset=queryset)
    lista = []
    informes_revisor =f.qs
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    recomendaciones = RecomendacionRevisor.objects.filter(informe_revisor_id__in=lista)
    
    return render(request, "core/informes_revisor_lista.html", {"informes_revisor":informes_revisor,'recomendaciones':recomendaciones,'filter':f,'tipo_usuario':tipo_usuario})

@login_required
def InformeRevisorDetalleView(request,id):
    request.session['idinforme'] = id
    informe_revisor = InformeRevisor.objects.filter(id=id)
    recomendacion = RecomendacionRevisor.objects.filter(informe_revisor_id=id)
    anexos = AnexoInformeRevisor.objects.filter(informe_revisor_id=id)
    return render(request, "core/informe_revisor_detalle.html", {'informe_revisor':informe_revisor,'recomendacion':recomendacion,'anexos':anexos})

def InformeRevisorContenidoView(request,id):
    informe_revisor = InformeRevisor.objects.filter(id=id)
    return render(request, "core/informe_revisor_contenido.html", {'informe_revisor':informe_revisor})

class CreaInformeRevisorView(LoginRequiredMixin,CreateView):
    model = InformeRevisor
    template_name = 'core/informe_revisor_form.html'
    form_class = InformeRevisorForm
    success_url = reverse_lazy('lista_informes_revisor')

def AjaxGuardaIdInformeView(request):
    id = request.GET.get('id', None)
    request.session["idinforme"] = int(id)
    data = {'id':id}
    return JsonResponse(data)    

class EditaInformeRevisorView(LoginRequiredMixin,UpdateView):
    model = InformeRevisor
    fields = ['fecha','contenido']
    template_name = 'core/informe_revisor_form.html'
    #success_url = reverse_lazy('informe_revisor_detalle')

    def get_success_url(self):
        idinforme = self.object.id
        return reverse_lazy('informe_revisor_detalle',args=[idinforme])

class BorraInformeRevisorView(LoginRequiredMixin,DeleteView):
    model = InformeRevisor
    success_url = reverse_lazy('lista_informes_revisor')
    template_name = 'core/confirmar_borrado_registro.html'

import re
def strip_tags(value):
    return re.compile(r'<[^<]*?/?>').sub('', value)

def cabecera_informes_revisor_pdf(pdf,pagina,titulo):
    #Canvas.setStrokeColorRGB(255, 0, 0)
    #Canvas.setFillColorRGB(0.2, 0.2, 0.2)
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))

def tabla_informes_revisor_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,tipo,idinforme,estado):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if tipo == 1:
        if idinforme > 0 :
            cuerpo = InformeRevisor.objects.filter(id=idinforme)[y1:y2]
        else:
            cuerpo = InformeRevisor.objects.filter(id__in=lista)[y1:y2]        
    else :
        if idinforme > 0 :
            if estado == 0:
                cuerpo = RecomendacionRevisor.objects.filter(informe_revisor_id=idinforme,cumplido=False)[y1:y2]
            else:
                cuerpo = RecomendacionRevisor.objects.filter(informe_revisor_id=idinforme,cumplido=True)[y1:y2]    
        else:
            if estado == 0 :    
                cuerpo = RecomendacionRevisor.objects.filter(cumplido=False)[y1:y2]
            else:    
                cuerpo = RecomendacionRevisor.objects.filter(informe_revisor_id__in=lista,cumplido=True)[y1:y2]
    if tipo == 2 or tipo ==3 :
        for i in cuerpo:
            if i.cumplido == True:
                i.cumplido ='Si'
            else:
                i.fecha_cumplido = (date.today()-i.informe_revisor.fecha).days
                i.cumplido ='No'    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    if tipo == 1:
        c = 600 - registros_cuerpo*16
    else:    
        c = 700 - registros_cuerpo*16
    y = c
    sw = 0
    if tipo == 1:
        encabezados = ['Fecha','Contenido']
    else:
        if tipo == 2:    
            encabezados = ['fecha','Recomendación','Cumplido','Fecha Cump.']
        else:
           if tipo == 3:
               encabezados = ['fecha','Recomendación','Cumplido','Días']         
    largo = 120
    if tipo == 1:
        if idinforme == 0:
                inicial = 0
                final = 120
                for j in cuerpo:
                    detalles = [(j.fecha,j.contenido[inicial:final])]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,18 *cm])
                    if sw == 0:
                        encabezados = []
                        detalle.setStyle(TableStyle(
                        [
                        #La primera fila(encabezados) va a estar centrada
                        ('ALIGN',(0,0),(3,0),'CENTER'),
                        #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        #El tamaño de las letras de cada una de las celdas será de 10
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ]
                        ))
                    else:
                        detalle.setStyle(TableStyle(
                        [
                        #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        #El tamaño de las letras de cada una de las celdas será de 10
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ]
                        ))                    
                    detalle.wrapOn(pdf, 500, 600)
                    detalle.drawOn(pdf, 15,y)
                    y -= 18
                    sw = 1
        else:        

            for j in cuerpo:
                segmentos = int(len(j.contenido)/largo)
                if segmentos<1:
                    segmentos = 0
                segmentos = segmentos+1
                x = range(segmentos)
                final = 0
                for i in x:
                    if i == 0:
                        inicial = 0
                    else:
                        inicial = (final)
                    final = (inicial + largo)
                    if i == 0:            
                        detalles = [(j.fecha,j.contenido[inicial:final])]
                        detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,18 *cm])
                    else: 
                        encabezados = []   
                        detalles = [('',j.contenido[inicial:final])]
                        detalle = Table(detalles, colWidths=[1.5 * cm,18 * cm])

                    if sw == 0:
                        detalle.setStyle(TableStyle(
                        [
                        #La primera fila(encabezados) va a estar centrada
                        ('ALIGN',(0,0),(3,0),'CENTER'),
                        #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        #El tamaño de las letras de cada una de las celdas será de 10
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ]
                        ))
                    else:
                        detalle.setStyle(TableStyle(
                        [
                        #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        #El tamaño de las letras de cada una de las celdas será de 10
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ]
                        ))                    
                    detalle.wrapOn(pdf, 500, 600)
                    detalle.drawOn(pdf, 15,y)
                    y -= 18
                    sw = 1
    else:
        for j in cuerpo:
            segmentos = int(len(j.recomendacion)/largo)
            if segmentos<1:
                segmentos = 0
            segmentos = segmentos+1    
            x = range(segmentos)
            final = 0
            for i in x:
                if i == 0:
                    inicial = 0
                else:
                    inicial = (final)
                final = (inicial + largo)
                if i == 0:            
                    detalles = [(j.informe_revisor.fecha,j.recomendacion[inicial:final],j.cumplido,j.fecha_cumplido)]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,15 *cm,1.5 *cm,1.5 *cm])
                else: 
                    encabezados = []   
                    detalles = [('',j.recomendacion[inicial:final],'','')]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 *cm,15 *cm,1.5 *cm,1.5 *cm])
                
                if sw ==0 :
                    detalle.setStyle(TableStyle(
                    [
                    #La primera fila(encabezados) va a estar centrada
                    ('ALIGN',(0,0),(3,0),'CENTER'),
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ]
                    ))
                else:
                    detalle.setStyle(TableStyle(
                    [
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ]
                    ))
                detalle.wrapOn(pdf, 500, 600)
                detalle.drawOn(pdf, 15,y)
                y -= 18
                sw = 1

class ReportesInformeRevisorPdfView(TemplateView):
    template_name = 'core/reportes_informe_revisor_pdf.html'

    """ def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        self.request.session['idinforme'] = 0
        self.request.session['lista_filtrada'] = ''
        #self.request.session['detalle_reporte_reunion_consejo'] = 0 """

def ReporteInformeRevisorPdfView(request,tipo):
    idinforme = request.session['idinforme']
    filas_hoja = 25
    lista = request.session['lista_filtrada']
    estado = request.session['estado']
    print('lista :',lista)
    print('tipo :',tipo)
    print('estado :',estado)
    print('idinforme :',id)
    if tipo == 1:
        if idinforme > 0 : 
            cantidad_reg = InformeRevisor.objects.filter(id=idinforme).count()
        else :
            cantidad_reg = InformeRevisor.objects.filter(id__in=lista).count()
    if tipo == 2:
        if idinforme > 0 : 
            cantidad_reg = RecomendacionRevisor.objects.filter(informe_revisor_id=idinforme).count()
        else :
            cantidad_reg = RecomendacionRevisor.objects.filter(informe_revisor_id__in=lista).count()
    if tipo ==3:
            cantidad_reg = RecomendacionRevisor.objects.filter(cumplido=False).count()            
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    #request.session["cantidad_reg"] =cantidad_reg
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "informes Revisor"
    if  tipo == 1:
        titulo = 'REPORTE DE INFORME REVISOR FISCAL '
        retorno = 'reporte_informe_revisor_pdf'    
    if tipo == 2:
        titulo = 'REPORTE DE INFORME REVISOR FISCAL '
        retorno = 'reporte_recomendaciones_informe_revisor_pdf'   
    if tipo == 3:
        titulo = 'REPORTE DE PENDIENTES RECOMENDACIONES REVISOR FISCAL '
        retorno = 'reporte_pendientes_recomendaciones_informe_revisor_pdf'        
    titulo = 'REPORTE DE INFORMES REVISOR FISCAL'
    request.session["titulo"] = titulo
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteInformeRevisorPdfView(request,pagina,tipo):
    if tipo ==3:
        idinforme = 0
    else:    
        idinforme = request.session['idinforme']
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"]
    estado = request.session['estado']
    #cantidad_reg = request.session["cantidad_reg"] 
    #total_paginas = request.session["total_paginas"]
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_informes_revisor_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_informes_revisor_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,tipo,idinforme,estado)
    #y -= 10
    #if pagina == total_paginas: 
    #    pdf.drawString(430, y, u"Total Activos      : "+('{:,}'.format(total_activos)+'.00'))
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

def ReporteInformeRevisorXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
        
    revisor_imprimir = InformeRevisor.objects.all().order_by('fecha')
    worksheet.write('A1','Fecha' )
    worksheet.write('B1','Contenido' )
    worksheet.write('C1','Recomendación' )
    worksheet.write('D1','Contenido' )
    worksheet.write('E1','Fecha Cumplido' )
    n = 2
    for j in revisor_imprimir :
        nn = str(n)
        sfecha= j.fecha.strftime("%m/%d/%Y")
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        #exec("worksheet.write('B"+nn+"','"+j.contenido+"' )")
        recomendacion_imprimir = RecomendacionRevisor.objects.filter(id=j.id)
        for k in recomendacion_imprimir :
            exec("worksheet.write('C"+nn+"','"+k.recomendacion+"' )")
            if k.cumplido == True:
                exec("worksheet.write('D"+nn+"','Si' )")
            else:
                exec("worksheet.write('D"+nn+"','No' )")
            sfecha_cumplido= k.fecha_cumplido.strftime("%m/%d/%Y")
            exec("worksheet.write('E"+nn+"','"+sfecha_cumplido+"' )")
        n = n + 1    

    workbook.close()
    output.seek(0)
    filename='informe_revisor.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    

class BorraAnexoInformeRevisorView(LoginRequiredMixin,DeleteView):
    model = AnexoInformeRevisor
    success_url = reverse_lazy('lista_informes_revisor')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idinforme = self.object.informe_revisor_id
        informe = InformeRevisor.objects.get(id=idinforme)
        return reverse_lazy('informe_revisor_detalle',args=[informe.id])

class BorraRecomendacionView(LoginRequiredMixin,DeleteView):
    model = RecomendacionRevisor
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idinforme = self.object.informe_revisor_id
        return reverse_lazy('detalle_reunion_consejo',args=[idinforme])

class CreaRecomendacionView(LoginRequiredMixin,CreateView):
    model = RecomendacionRevisor
    template_name = 'core/recomendacion_revisor_form.html'
    form_class = RecomendacionRevisorForm
    
    def get_success_url(self):
        idinforme = self.object.informe_revisor_id
        return reverse_lazy('informe_revisor_detalle',args=[idinforme])

    def form_valid(self, form):
        if form.is_valid():
            idinforme = self.kwargs['id']
            object = form.save(commit=False)
            object.informe_revisor_id = idinforme
            object.cumplido = False
            object.fecha_cumplido=None
            object.save()
            idinforme = self.kwargs['id']
            return super(CreaRecomendacionView, self).form_valid(form)
        else:
            mensaje = "Error"
            return render(self.request, "core/mensaje_error_usuario.html", {'mensaje':mensaje})

class EditaRecomendacionView(LoginRequiredMixin,UpdateView):
    model = RecomendacionRevisor
    fields = ['recomendacion','cumplido','fecha_cumplido']
    template_name = 'core/recomendacion_revisor_form.html'
    
    def get_success_url(self):
        idinforme = self.object.informe_revisor_id
        return reverse_lazy('informe_revisor_detalle',args=[idinforme])

class AplicaRecomendacionView(LoginRequiredMixin,UpdateView):
    model = RecomendacionRevisor
    fields = ['cumplido','fecha_cumplido']
    template_name = 'core/recomendacion_revisor_form.html'
    
    def get_success_url(self):
        tipo_usuario = self.request.session['tipo_usuario']
        idrecomendacion = self.object.id
        recomendacion = RecomendacionRevisor.objects.get(id=idrecomendacion) 
        idinforme = recomendacion.informe_revisor_id
        consejo = MiembroConsejo.objects.filter(activo=True)
        asunto = "Cumplimiento recomendación revisor fiscal"
        for i in consejo:
            mensaje = "Sr. Consejero: "+i.nombre+" se le informa que la administración ha aplicado la recomendacion del revisor fiscal : "+recomendacion.recomendacion
            EnvioMailOtros(mensaje,asunto,i.email)
        idinforme = self.object.informe_revisor_id
        return reverse_lazy('informe_revisor_detalle',args=[idinforme])


@login_required
def RecomendacionDetalleView(request,id):
    tipo_usuario = request.session['tipo_usuario']
    recomendacion = RecomendacionRevisor.objects.get(id=id)
    informe = InformeRevisor.objects.get(id=recomendacion.informe_revisor_id)
    informe_revisor = InformeRevisor.objects.filter(id=informe.id)
    recomendacion = RecomendacionRevisor.objects.filter(informe_revisor_id=informe.id)
    anexos = AnexoInformeRevisor.objects.filter(informe_revisor_id=informe.id)
    return render(request, "core/informe_revisor_detalle.html", {'informe_revisor':informe_revisor,'recomendacion':recomendacion,'anexos':anexos,'tipo_usuario':tipo_usuario})


######################################## Juridico #################################

@login_required
def ListaProcesosJuridicosView(request):
    request.session["idproceso"] = 0
    request.session['numero_proceso'] = ''
    request.session['tipo_proceso'] = ''
    queryset = ProcesoJuridico.objects.all().order_by('-fecha_inicial')
    f = ProcesosJuridicosFilter(request.GET, queryset=queryset)
    lista = []
    procesos_juridicos = f.qs
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    return render(request, "core/procesos_juridicos_lista.html",  {'procesos_juridicos':procesos_juridicos,'filter': f})


@login_required
def ProcesoJuridicoDetalleView(request,id):
    request.session["idproceso"] = id
    request.session['proceso_juridico'] = id
    queryset = ProcesoJuridico.objects.filter(id=id)
    f = ProcesosJuridicosFilter(request.GET, queryset=queryset)
    lista = []
    proceso_juridico = f.qs
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    anexos = AnexoProcesoJuridico.objects.filter(proceso_juridico_id=id)
    gestiones = GestionProcesoJuridico.objects.filter(proceso_juridico_id=id)
    lista_gestiones = []
    for i in gestiones:
        lista_gestiones.append(i.id)
    anexos_gestion = AnexoGestionProcesoJuridico.objects.filter(gestion_proceso_juridico_id__in=lista_gestiones)
    return render(request, "core/procesos_juridicos_detalle.html",  {'proceso_juridico':proceso_juridico,'anexos':anexos,'gestiones':gestiones,'anexos_gestion':anexos_gestion,'filter': f})

def DireccionaSalidaProcesosJuridicosView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
       return redirect('pagina_administrador')
    else:
        if tipo_usuario == 2 or tipo_usuario == 5:
            return redirect('pagina_consejo')    
        else:
            return redirect('user_home',tipo_usuario)
        
class CreaProcesoJuridicoView(LoginRequiredMixin,CreateView):
    model = ProcesoJuridico
    template_name = 'core/proceso_juridico_form.html'
    form_class = ProcesoJuridicoForm
    
    def get_initial(self):
        initial = super(CreaProcesoJuridicoView, self).get_initial()
        initial['numero_proceso'] = self.request.session['numero_proceso']
        initial['tipo_proceso'] = self.request.session['tipo_proceso']
        return initial

    def get_success_url(self):
        self.request.session['numero_proceso'] = ''
        self.request.session['tipo_proceso'] = ''
        return reverse_lazy('lista_procesos_juridicos')

class EditaProcesoJuridicoView(LoginRequiredMixin,UpdateView):
    model = ProcesoJuridico
    fields = ['proceso_numero','tipo_proceso','fecha_inicial','abogado','demandante','demandado','juzgado','contenido','interior','apartamento','fecha_final','valor_demanda','activo']
    template_name = 'core/proceso_juridico_form.html'
    
    def get_success_url(self):
        idproceso = self.object.id
        return reverse_lazy('proceso_juridico_detalle',args=[idproceso])

class BorraProcesoJuridicoView(LoginRequiredMixin,DeleteView):
    model = ProcesoJuridico
    success_url = reverse_lazy('lista_procesos_juridicos')
    template_name = 'core/confirmar_borrado_registro.html'

def AjaxGuardaIdProcesoView(request):
    id = request.GET.get('id', None)
    request.session["idproceso"] = id
    data = {'id':id}
    return JsonResponse(data)
      

def ReporteProcesosJuridicosPdfView(request):
    filas_hoja = 35
    lista = request.session['lista_filtrada'] 
    cantidad_reg = ProcesoJuridico.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Proceso"    
    titulo = 'REPORTE DE PROCESOS JURIDICOS'
    request.session["titulo"] = titulo
    retorno = 'reporte_procesos_juridicos'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def ReporteGestionProcesoPdfView(request,id):
    filas_hoja = 35
    lista = request.session['lista_filtrada']
    request.session['idgestion'] = id
    cantidad_reg = GestionProcesoJuridico.objects.filter(id__in=lista).count()
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = "Proceso"    
    titulo = 'REPORTE DE GESTION PROCESOS JURIDICOS'
    request.session["titulo"] = titulo
    retorno = 'reporte_gestion_procesos_juridicos'
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})


def cabecera_procesos_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    pdf.drawImage(archivo_imagen, 40, 540, 120, 90,preserveAspectRatio=True)
    pdf.drawString(250,580, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,550 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,550 , u"Página No. : "+str(pagina))

def tabla_gestion_proceso_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    
    cuerpo = GestionProcesoJuridico.objects.filter(id=id).select_related()[y1:y2]
    registros_cuerpo = cuerpo.count()
    c = 480 - registros_cuerpo*16
    y = c
    cuerpo_total = GestionProcesoJuridico.objects.filter(id=id)
    encabezados = ['Fecha','Titulo','Gestión']  
    sw = 0
    largo = 200
    for j in cuerpo:
        if sw == 0 :
            pdf.drawString(20,y , u"Proceso : "+j.proceso_juridico.proceso_numero)
            pdf.drawString(370,y , u"Tipo Proceso : "+j.proceso_juridico.tipo_proceso.descripcion)
            y -= 18
            pdf.drawString(20,y , u"Fecha Inicial : "+j.proceso_juridico.fecha_inicial.strftime("%Y-%m-%d %H:%M:%S"))
            if j.proceso_juridico.fecha_final != None :
                pdf.drawString(370,y , u"Fecha Final : "+j.proceso_juridico.fecha_final.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                pdf.drawString(370,y , u"Fecha Final : ")    
            y -= 18
            pdf.drawString(20,y , u"Abogado : "+j.proceso_juridico.abogado.nombre)
            pdf.drawString(370,y , u"Demandante : "+j.proceso_juridico.demandante)
            y -= 18
            pdf.drawString(20,y , u"Demandado : "+j.proceso_juridico.demandado)
            pdf.drawString(370,y , u"Juzgado : "+j.proceso_juridico.juzgado)
            y -= 18
            pdf.drawString(20,y , u"Apartamento : "+j.proceso_juridico.apartamento.numero)
            pdf.drawString(370,y , u"Valor Demanda : "+str(j.proceso_juridico.valor_demanda))
            y -= 40
            segmentos = int(len(j.gestion)/largo)
            if segmentos<1:
                segmentos = 0
            segmentos = segmentos+1    
            x = range(segmentos)
            final = 0

            for i in x:
                if i == 0:
                    inicial = 0
                else:
                    inicial = (final)
                final = (inicial + largo)
                if i == 0:            
                    detalles = [(j.fecha,j.titulo[0:35],j.gestion[inicial:final])]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,3 * cm,23 * cm])
                else: 
                    encabezados = []   
                    detalles = [('','',j.gestion[inicial:final])]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,3 * cm,23 * cm])
                
                if sw == 0:
                    detalle.setStyle(TableStyle(
                    [
                    #La primera fila(encabezados) va a estar centrada
                    ('ALIGN',(0,0),(3,0),'CENTER'),
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ]
                    ))
                else:
                    detalle.setStyle(TableStyle(
                    [
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ]
                    ))    

                detalle.wrapOn(pdf, 500, 600)
                detalle.drawOn(pdf, 5,y)
                y -= 18
                sw = 1

def PaginasGestionProcesoPdfView(request,pagina):
    id = request.session["idgestion"] 
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    #cantidad_reg = request.session["cantidad_reg"] 
    #total_paginas = request.session["total_paginas"]
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    from reportlab.lib.pagesizes import letter, landscape
    pdf.setPageSize(landscape(letter))
    cabecera_procesos_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_gestion_proceso_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response
    
def tabla_procesos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,idproceso):
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    
    if idproceso > 0:
        cuerpo = ProcesoJuridico.objects.filter(id=idproceso).select_related()[y1:y2]
    else:
        cuerpo = ProcesoJuridico.objects.filter(id__in=lista).select_related()[y1:y2]
    for i in cuerpo:
        if i.activo == True:
            i.activo ='✔'
        else:
            i.activo ='✘'
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 480 - registros_cuerpo*16
    y = c
    if idproceso > 0:
        cuerpo_total = ProcesoJuridico.objects.filter(id=idproceso)
        encabezados = ['Número','Tipo Proceso','Contenido']  
    else:
        cuerpo_total = ProcesoJuridico.objects.filter(id__in=lista)                
        encabezados = ['Número','Tipo Proceso','Fecha Inicial','Abogado','Demandante','Demandado','Juzgado','Apartamento','Fecha Final','Valor','Activo']  
    sw = 0
    largo = 200
    if idproceso > 0:
        for j in cuerpo:
            if sw == 0 :
                
                pdf.drawString(20,y , u"Proceso : "+j.proceso_numero)
                pdf.drawString(370,y , u"Tipo Proceso : "+j.tipo_proceso.descripcion)
                y -= 18
                pdf.drawString(20,y , u"Fecha Inicial : "+j.fecha_inicial.strftime("%Y-%m-%d %H:%M:%S"))
                if j.fecha_final != None :
                    pdf.drawString(370,y , u"Fecha Final : "+j.fecha_final.strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    pdf.drawString(370,y , u"Fecha Final : ")    
                y -= 18
                pdf.drawString(20,y , u"Abogado : "+j.abogado.nombre)
                pdf.drawString(370,y , u"Demandante : "+j.demandante)
                y -= 18
                pdf.drawString(20,y , u"Demandado : "+j.demandado)
                pdf.drawString(370,y , u"Juzgado : "+j.juzgado)
                y -= 18
                pdf.drawString(20,y , u"Apartamento : "+j.apartamento.numero)
                pdf.drawString(370,y , u"Valor Demanda : "+str(j.valor_demanda))
                y -= 40
            segmentos = int(len(j.contenido)/largo)
            if segmentos<1:
                segmentos = 0
            segmentos = segmentos+1    
            x = range(segmentos)
            final = 0

            for i in x:
                if i == 0:
                    inicial = 0
                else:
                    inicial = (final)
                final = (inicial + largo)
                if i == 0:            
                    detalles = [(j.proceso_numero,j.tipo_proceso.descripcion[0:15],j.contenido[inicial:final])]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,1.5 * cm,24 * cm])
                else: 
                    encabezados = []   
                    detalles = [('','',j.contenido[inicial:final])]
                    detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,1.5 * cm,24 * cm])
                
                if sw == 0:
                    detalle.setStyle(TableStyle(
                    [
                    #La primera fila(encabezados) va a estar centrada
                    ('ALIGN',(0,0),(3,0),'CENTER'),
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ]
                    ))
                else:
                    detalle.setStyle(TableStyle(
                    [
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ]
                    ))    

                detalle.wrapOn(pdf, 500, 600)
                detalle.drawOn(pdf, 5,y)
                y -= 18
                sw = 1
    else:
        for j in cuerpo:
            if sw > 0:
                encabezados = []
            detalles = [(j.proceso_numero,j.tipo_proceso.descripcion[0:16],j.fecha_inicial,j.abogado.nombre[0:30],j.demandante[0:30],j.demandado[0:30],j.juzgado[0:32],j.apartamento,j.fecha_final,formatNumber(j.valor_demanda,0),j.activo)]
            detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,1.5 * cm,1.5 * cm,4.2 * cm,4.3 * cm,4.3 * cm,4.3 * cm,2 *cm,2 *cm,1 *cm])
            if sw == 0:            
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]
                ))
            else:
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]
                ))    
            detalle.wrapOn(pdf, 800, 400)
            detalle.drawOn(pdf, 5,y)
            y -= 18
            sw = 1

def PaginasReporteProcesosJuridicosPdfView(request,pagina):
    idproceso = request.session["idproceso"] 
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    from reportlab.lib.pagesizes import letter, landscape
    pdf.setPageSize(landscape(letter))
    cabecera_procesos_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    tabla_procesos_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,idproceso)
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

def ReporteProcesosJuridicosXlsView(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada']
    procesos_imprimir = ProcesoJuridico.objects.all().filter(id__in=lista).order_by('proceso_numero')
    worksheet.write('A1','Proceso No.' )
    worksheet.write('B1','Tipo Proceso' )
    worksheet.write('C1','Fecha Inicial')
    worksheet.write('D1','Contenido')
    worksheet.write('E1','Abogado')
    worksheet.write('F1','Demandante')
    worksheet.write('G1','Juzgado')
    worksheet.write('H1','Aparatmento')
    worksheet.write('I1','Fecha Final')
    worksheet.write('J1','Valor')
    worksheet.write('K1','Activo')
    n = 2
    
    for j in procesos_imprimir :
        nn = str(n)
        if j.fecha_final != None:
            sfecha_final = j.fecha_final.strftime("%m/%d/%Y")
        else:
            sfecha_final = ''
        if j.fecha_inicial != None:    
            sfecha_inicial = j.fecha_inicial.strftime("%m/%d/%Y")
        else:    
            sfecha_inicial = ''
        exec("worksheet.write('A"+nn+"','"+j.proceso_numero+"' )")
        exec("worksheet.write('B"+nn+"','"+j.tipo_proceso.descripcion+"' )")
        exec("worksheet.write('C"+nn+"','"+sfecha_inicial+"' )")
        exec("worksheet.write('D"+nn+"','"+j.contenido+"' )")
        exec("worksheet.write('E"+nn+"','"+j.abogado.nombre+"' )")
        exec("worksheet.write('F"+nn+"','"+j.demandante+"' )")
        exec("worksheet.write('G"+nn+"','"+j.juzgado+"' )")
        exec("worksheet.write('H"+nn+"','"+j.apartamento.numero+"' )")
        exec("worksheet.write('I"+nn+"','"+sfecha_final+"' )")
        exec("worksheet.write('J"+nn+"','"+str(j.valor_demanda)+"' )")
        if j.activo == True:
            exec("worksheet.write('K"+nn+"','Si' )")
        else:
            exec("worksheet.write('K"+nn+"','No' )")          
        n = n + 1 
    workbook.close()
    output.seek(0)
    filename='procesos.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response

class BorraAnexoProcesoJuridicoView(LoginRequiredMixin,DeleteView):
    model = AnexoProcesoJuridico
    success_url = reverse_lazy('lista_procesos_juridicos')
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idproceso = self.object.proceso_juridico_id
        proceso = ProcesoJuridico.objects.get(id=idproceso)
        return reverse_lazy('proceso_juridico_detalle',args=[proceso.id])  

def ProcesoJuridicoContenidoView(request,id):
    proceso_juridico = ProcesoJuridico.objects.filter(id=id)
    return render(request, "core/proceso_juridico_contenido.html", {'proceso_juridico':proceso_juridico})

class CreaGestionProcesoJuridicoView(LoginRequiredMixin,CreateView):
    model = GestionProcesoJuridico
    template_name = 'core/gestion_proceso_juridico_form.html'
    form_class = GestionProcesoJuridicoForm
    
    def get_initial(self,*args,**kwargs):
        initial=super(CreaGestionProcesoJuridicoView,self).get_initial(**kwargs)
        idproceso = self.request.session['proceso_juridico']
        initial['proceso_juridico'] = idproceso
        return initial

    def get_success_url(self):
        idgestion = self.object.id
        gestion_proceso_juridico = GestionProcesoJuridico.objects.get(id=idgestion) 
        idproceso = gestion_proceso_juridico.proceso_juridico_id
        return reverse_lazy('proceso_juridico_detalle',args=[idproceso])

class EditaGestionProcesoJuridicoView(LoginRequiredMixin,UpdateView):
    model = GestionProcesoJuridico
    fields = ['fecha','titulo','proceso_juridico','gestion']
    template_name = 'core/gestion_proceso_juridico_form.html'
    
    def get_success_url(self):
        idgestion = self.object.id
        gestion_proceso_juridico = GestionProcesoJuridico.objects.get(id=idgestion) 
        idproceso = gestion_proceso_juridico.proceso_juridico_id
        return reverse_lazy('proceso_juridico_detalle',args=[idproceso])

class BorraGestionProcesoJuridicoView(LoginRequiredMixin,DeleteView):
    model = GestionProcesoJuridico
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idgestion = self.object.id
        gestion_proceso_juridico = GestionProcesoJuridico.objects.get(id=idgestion) 
        idproceso = gestion_proceso_juridico.proceso_juridico_id
        return reverse_lazy('proceso_juridico_detalle',args=[idproceso])

class BorraAnexoGestionProcesoJuridicoView(LoginRequiredMixin,DeleteView):
    model = AnexoGestionProcesoJuridico
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idgestion = self.object.gestion_proceso_juridico_id
        gestion_proceso = GestionProcesoJuridico.objects.get(id=idgestion)
        idproceso = gestion_proceso.proceso_juridico_id
        return reverse_lazy('proceso_juridico_detalle',args=[idproceso])  

class CreaTipoProcesoView(LoginRequiredMixin,CreateView):
    model = TipoProceso
    template_name = 'core/crear_tipo_proceso.html'
    form_class = TipoProcesoForm
    #success_url = reverse_lazy('lista_proveedores')

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        self.request.session['tipo_proceso'] = self.object.id
        return HttpResponseRedirect(reverse('crea_proceso_juridico'))

@csrf_exempt
def GuardaNumeroProcesoView(request):    
    numero_proceso = request.GET.get('numero_proceso', None)
    request.session['numero_proceso'] = numero_proceso
    id = 0
    data = {'id':id}
    return JsonResponse(data)   

def CargarNumeroProcesoView(request):    
    numero_proceso = request.session['numero_proceso']
    id = 0
    data = {'numero':numero_proceso}
    
    ############################ Importacion Datos  ##########################

def import_xls_apartamentos(request):
    if "GET" == request.method:
        return render(request, 'core/importar_xls_apartamentos.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Apartamentos"]
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        f = Apartamento()
        """ for row in worksheet.iter_rows(): """
        r=0
        for row in excel_data:
            if r > 0:
                interior = row[0]
                interior = Interior.objects.get(numero=interior)
                f.interior_id = interior.id
                f.apartamento = row[1]
                f.numero = row[2]
                f.coeficiente = row[3]
                f.save()
                f.id  += 1
            r = r+1    
        return render(request, 'core/ver_importar_xls_apartamentos.html', {"excel_data":excel_data})

def import_xls_residentes(request):
    if "GET" == request.method:
        return render(request, 'core/importar_xls_residentes.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Residentes"]
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        f = Residente()
        """ for row in worksheet.iter_rows(): """
        r=0
        for row in excel_data:
            if r > 0:
                f.identificacion = row[0]
                f.nombre = row[1]
                f.tipo_residente = row[2]
                f.telefono = row[3]
                f.celular = row[4]
                f.email = row[5]
                f.edad = 0
                f.genero = row[7]
                f.persona_discapacitada = row[8]
                interior = row[9]
                interior = Interior.objects.get(numero=interior)
                f.interior_id =interior.id
                apartamento = row[10] 
                apartamento = Apartamento.objects.filter(apartamento=apartamento,interior_id=interior.id).latest('id')
                f.apartamento_id = apartamento.id    
                f.save()
                f.id  += 1
            r = r+1    
        return render(request, 'core/ver_importar_xls_residentes.html', {"excel_data":excel_data})

def import_xls_parqueaderos(request):
    if "GET" == request.method:
        return render(request, 'core/importar_xls_parqueaderos.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Parqueaderos"]
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        f = Parqueadero()
        r=0
        for row in excel_data:
            if r > 0:
                f.numero = row[0]
                interior = row[1]
                numero = row[0]
                interior = Interior.objects.get(numero=interior)
                f.interior_id = interior.id
                apartamento = row[2]
                apartamento = Apartamento.objects.filter(apartamento=apartamento,interior_id=interior.id).latest('id')
                f.apartamento_id = apartamento.id
                f.coeficiente = row[3]
                f.save()
                f.id  += 1
            r = r+1    
        return render(request, 'core/ver_importar_xls_parqueaderos.html', {"excel_data":excel_data})

#################################### Tokens ############################

def SolicitudTokenView(request):
    tipo_usuario=request.session['tipo_usuario']
    form = SolicitudTokenForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            idapartamento = form.cleaned_data['apartamento']
            idinterior = form.cleaned_data['interior']
            email = form.cleaned_data['email']    
            if Residente.objects.filter(email=email,interior_id=idinterior,apartamento_id=idapartamento).exists():
                token = get_random_string(length=12)
                residente = Residente.objects.get(email=email)
                apartamento = Apartamento.objects.get(id=residente.apartamento_id)
                mensaje =" su token para accesar al sistema de conjunto-web es: "+token+' Ingrese por <login> <usuario no registrado> y con éste token complete el registro'
                asunto = "Envío Token para registro en conjunto-web"
                EnvioMailResidente(apartamento.id,mensaje,asunto,email)
                messages.add_message(request, messages.INFO,"Se le ha enviado el token a su email, para iniciar el registro")
                #messages.info(request, f"Se le ha enviado el token a su email, para iniciar el registro")
                return render(request, template_name="core/mensaje_login_token.html",context={'url':'home',})
            else:
                messages.add_message(request, messages.INFO,"El email ingresado no esta registrado en la información del conjunto, Solicite actualizar su información en la administración")
                #messages.error(request,"El email ingresado no esta registrado en la información del conjunto, Solicite actualizar su información en la administración")
                return render(request, template_name="core/mensaje_login_token.html",context={'url':'home'})
                
        else:
            messages.add_message(request, messages.ERROR,"Error Formulario")
            messages.error(request,"Error formulario")
    else:
        form = SolicitudTokenForm()        
    return render(request, template_name="core/registro_residente.html",context={'form':form})

def MarcaTokensResidentesView(request):
    lista = request.session['lista_filtrada']
    residentes = Residente.objects.filter(id__in=lista)
    for i in residentes:
        if i.envio_token == True:
            Residente.objects.filter(id=i.id).update(envio_token=False)
        else:
            Residente.objects.filter(id=i.id).update(envio_token=True)
    residentes = Residente.objects.filter(envio_token=True)        
    f = ResidentesFilter(request.GET, queryset=residentes)
    residentes = TokensResidentesTable(f.qs)
    return render(request, "core/pagina_token_residentes_result.html", {'residentes':residentes})   

def EnvioTokensResidentesView(request):
    Residente.objects.update(envio_token=False)
    queryset = Residente.objects.all()
    f = ResidentesTokenFilter(request.GET, queryset=queryset)
    lista = []
    residentes = TokensResidentesTable(f.qs)
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    return render(request, "core/pagina_tokens_residentes.html", {'residentes':residentes,'filter': f})

def ajaxSwitchTokenResidentesView(request):
    id = request.GET.get('id', None)
    residente = Residente.objects.get(id=id)
    if residente.envio_token == True:
        envio_token = False
    else:
        envio_token = True
    Residente.objects.filter(id=id).update(aprobado=envio_token)    
    data = {'envio':envio_token}
    return JsonResponse(data) 

def GenerarTokensView(request):
    residentes = Residente.objects.filter(envio_token = True)
    for i in residentes:
        token = get_random_string(length=12)
        mensaje ="Sr/Sra."+i.nombre+" su token para accesar al sistema de conjunto-web es: "+token+' Ingrese por <login> <usuario no registrado> y con éste token complete el registro'
        asunto = "Envío Token para registro en conjunto-web"
        mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [i.email])
        mail.send()
        Residente.objects.filter(id=i.id).update(token=token)
    return redirect('pagina_envio_tokens') 


############################################# ASAMBLEA  ####################################    

class PaginaAsambleaView(TemplateView):
    template_name = 'core\pagina_asamblea.html'

def DireccionaSalidaPaginaAsambleaView(request):
    tipo_usuario = request.session['tipo_usuario']
    if tipo_usuario == 1:
       return redirect('pagina_administrador')
    else:
        if tipo_usuario == 2 : 
            return redirect('pagina_consejo')
        else :
            return redirect('user_home',tipo_usuario)    

@login_required            
def ListaAsambleaView(request):
    request.session['user'] = request.user.id
    request.session['detalle_reporte_asamblea'] = 0
    request.session['tipo_asamblea'] = 1
    request.session['fecha'] = ''
    request.session["idasamblea"] = 0
    queryset = Asamblea.objects.all().order_by('-fecha')
    f = AsambleaFilter(request.GET, queryset=queryset)
    lista = []
    asambleas =f.qs
    for i in f.qs:
        lista.append(i.id)
    request.session['lista_filtrada'] = lista
    request.session["idasamblea"] = 0
    return render(request, "core/asamblea_lista.html", {"asambleas":asambleas,'filter':f})

@login_required
def DetalleAsambleaView(request,id):
    request.session["idasamblea"] = id
    request.session['detalle_reporte_asamblea'] = 1
    asamblea = Asamblea.objects.filter(id=id)
    anexos = AnexoAsamblea.objects.filter(asamblea_id=id)
    decisiones = DecisionAsamblea.objects.filter(asamblea_id=id)
    return render(request, "core/asamblea_detalle.html", {"asamblea":asamblea,'anexos':anexos,'decisiones':decisiones})

class CreaAsambleaView(LoginRequiredMixin,CreateView):
    model = Asamblea
    template_name = 'core/asamblea_form.html'
    form_class = AsambleaForm
    success_url = reverse_lazy('lista_asambleas')
    
    def get_initial(self,*args,**kwargs):
        initial=super(CreaAsambleaView,self).get_initial(**kwargs)
        initial['tipo_asamblea']=self.request.session['tipo_asamblea']
        initial['fecha']=self.request.session['fecha']
        return initial
    
class BorraAsambleaView(LoginRequiredMixin,DeleteView):
    model = Asamblea
    template_name = 'core/confirmar_borrado_registro.html'
    success_url = reverse_lazy('lista_asambleas')

class EditaAsambleaView(LoginRequiredMixin,UpdateView):
    model = Asamblea
    template_name = 'core/asamblea_form.html'
    form_class = AsambleaForm
    
    def get_success_url(self):
        idasamblea = self.object.id
        return reverse_lazy('detalle_asamblea',args=[idasamblea])

class CreaDecisionAsambleaView(LoginRequiredMixin,CreateView):
    model = DecisionAsamblea
    template_name = 'core/decision_asamblea_form.html'
    form_class = DecisionAsambleaForm
    
    def get_initial(self,*args,**kwargs):
        initial=super(CreaDecisionAsambleaView,self).get_initial(**kwargs)
        asamblea = Asamblea.objects.get(id=self.kwargs['id']) 
        initial['asamblea']=asamblea.id
        return initial

    def get_success_url(self):
        idasamblea = self.object.asamblea_id
        return reverse_lazy('detalle_asamblea',args=[idasamblea])

class EditaDecisionAsambleaView(LoginRequiredMixin,UpdateView):
    model = DecisionAsamblea
    fields = ['decision','numero_votos_favor','numero_votos_contra']
    template_name = 'core/decision_asamblea_form.html'
    
    def get_success_url(self):
        idasamblea = self.object.asamblea_id
        return reverse_lazy('detalle_asamblea',args=[idasamblea])
    
class BorraDecisionAsambleaView(LoginRequiredMixin,DeleteView):
    model = DecisionAsamblea
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idasamblea = self.object.asamblea_id
        return reverse_lazy('detalle_asamblea',args=[idasamblea])

   
def cabecera_asamblea_pdf(pdf,pagina,titulo):
    pdf.setFont('Helvetica', 12)
    archivo_imagen = settings.STATIC_ROOT+'/img/logo_conjunto.PNG'
    #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    pdf.drawImage(archivo_imagen, 40, 740, 120, 90,preserveAspectRatio=True)
    pdf.drawString(240,780, titulo)
    pdf.setFont('Helvetica', 9)
    pdf.drawString(30,750 , u"Fecha : "+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf.drawString(450,750 , u"Página No. : "+str(pagina))

from html.parser import HTMLParser

def tabla_asamblea_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,detalle,id):
    id = int(id)
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if id > 0:
        cuerpo = Asamblea.objects.filter(id=id)[y1:y2]
    else:        
        cuerpo = Asamblea.objects.filter(id__in=lista)[y1:y2]
    registros_cuerpo = cuerpo.count()
    print('total Registro :'+str(registros_cuerpo))
    #Uso el factor 16 que seria los pixel por cada registro que resto
    if detalle == 0:
        c = 700 - registros_cuerpo*16
    else:        
        c = 520 - registros_cuerpo*16
    y = c
    if id > 0:
        cuerpo_total = Asamblea.objects.filter(id=id)
    else:    
        cuerpo_total = Asamblea.objects.filter(id__in=lista)
    if detalle == 0:
        encabezados = ['Fecha','Número Acta','Hora Inicial','Hora Final','Total Horas']
    else :
        encabezados = ['Fecha','Contenido','Acta','Hora.Inic.','Hora.Fin.','Total Hor.']    
    largo = 120
    tiempo_total = 0
    sw = 0
    for j in cuerpo:
        linea =j.contenido.strip()
        segmentos = int(len(linea)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        final = 0
        tiempo_asamblea = float((datetime.strptime(str(j.hora_final), '%H:%M:%S')-datetime.strptime(str(j.hora_inicio), '%H:%M:%S')).seconds/3600.0)
        if detalle == 1:
            for i in x:
                if i == 0:
                    inicial = 0
                else:
                    inicial = (final)
                final = (inicial + largo)
                if sw == 0 :
                    detalles = [(j.fecha,linea[inicial:final],j.numero_acta,j.hora_inicio,j.hora_final,str(round(tiempo_asamblea,2)))]
                    sdetalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,14 * cm,1.0 * cm,1.3 * cm,1.3 *cm,1.3 *cm])
                else:
                    encabezados=[]
                    detalles = [('',linea[inicial:final],'','','','')]    
                    sdetalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,14 * cm,1.0 * cm,1.3 * cm,1.3 *cm,1.3 *cm])
                if sw == 0:
                    sdetalle.setStyle(TableStyle(
                    [
                    #La primera fila(encabezados) va a estar centrada
                    ('ALIGN',(0,0),(3,0),'CENTER'),
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ]
                    ))
                else:
                    sdetalle.setStyle(TableStyle(
                    [
                    #La primera fila(encabezados) va a estar centrada
                    #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    #El tamaño de las letras de cada una de las celdas será de 10
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ]
                    ))        
                sdetalle.wrapOn(pdf, 500, 600)
                if detalle == 0:
                    sdetalle.drawOn(pdf, 30,y)
                    y -= 18
                else:    
                    sdetalle.drawOn(pdf, 10,y)
                    y -= 18
                sw = 1    
        else:
            print('sw :'+str(sw))
            print('linea : '+linea[0:100])
            if sw == 0 :
                detalles = [(j.fecha,j.numero_acta,j.hora_inicio,j.hora_final,str(round(tiempo_asamblea,2)))]
                sdetalle = Table([encabezados] + detalles, colWidths=[3 * cm,5 * cm,2 * cm,2 * cm,2 *cm])
                sdetalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                ))
            else:
                encabezados= []
                detalles = [(j.fecha,j.numero_acta,j.hora_inicio,j.hora_final,str(round(tiempo_asamblea,2)))]
                sdetalle = Table([encabezados] + detalles, colWidths=[3 * cm,5 * cm,2 * cm,2 * cm,2 *cm])
                sdetalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                ))        
            sdetalle.wrapOn(pdf, 500, 600)
            if detalle == 0:
                sdetalle.drawOn(pdf, 30,y)
                y -= 18
            else:    
                sdetalle.drawOn(pdf, 10,y)
                y -= 18
            sw = 1        
        tiempo_total  += tiempo_asamblea
                    
    if detalle == 0:
        pdf.drawString(450,y, u"Tiempo Total ==> "+str(round(tiempo_total,2)))
    else:    
        pdf.drawString(490,y, u"Tiempo Total ==> "+str(round(tiempo_total,2)))

def tabla_decisiones_asamblea_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id):
    id = int(id)
    if pagina == 1:
        y1 = pagina-1
        y2 = (filas_hoja*pagina) - 1
    else:
        y1 = (pagina*filas_hoja)-1
        y2 = ((pagina*filas_hoja*pagina)+filas_hoja) - 1
    if id > 0 :
        cuerpo = DecisionAsamblea.objects.filter(asamblea_id=id)[y1:y2]
    else:
        cuerpo = DecisionAsamblea.objects.filter(asamblea_id__in=lista)[y1:y2]
    
    registros_cuerpo = cuerpo.count()
    #Uso el factor 16 que seria los pixel por cada registro que resto
    c = 700 - registros_cuerpo*16
    y = c
    """ if id > 0 :
        cuerpo_total = DecisionAsamblea.objects.filter(asamblea_id=id).select_related
    else:
        cuerpo_total = DecisionAsamblea.objects.filter(asamblea_id__in=lista).select_related     """
    sw = 0
    encabezados = ['Fecha','Decisión','Votos Favor','Votos Contra']
    largo = 120
    for j in cuerpo:
        segmentos = int(len(j.decision)/largo)
        if segmentos<1:
            segmentos = 0
        segmentos = segmentos+1    
        x = range(segmentos)
        final = 0
        if sw >= 1 :
            encabezados = [] 
        for i in x:
            if i == 0:
                inicial = 0
            else:
                inicial = (final)
            final = (inicial + largo)
            if i == 0:            
                detalles = [(j.asamblea.fecha,j.decision[inicial:final],j.numero_votos_favor,j.numero_votos_contra)]
                detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,15 * cm,1.5 * cm,1.8 * cm])
            else: 
                detalles = [('',j.decision[inicial:final],'','')]
                detalle = Table([encabezados] + detalles, colWidths=[1.5 * cm,15 * cm,1.5 * cm,1.8 * cm])
            if i==0 :
                detalle.setStyle(TableStyle(
                [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ]
                ))
            
            detalle.wrapOn(pdf, 500, 600)
            detalle.drawOn(pdf, 15,y)
            y -= 18
            sw = 1

def ReporteAsambleasPdfView(request,tipo):
    id = request.session["idasamblea"]
    lista = request.session['lista_filtrada']
    filas_hoja = 25
    if  tipo == 2:
        if id > 0 :
            cantidad_reg = DecisionAsamblea.objects.filter(asamblea_id=id).count()
        else:    
            cantidad_reg = DecisionAsamblea.objects.filter(asamblea_id__in=lista).count()
    elif tipo == 1 :
        if id > 0 :
            cantidad_reg = Asamblea.objects.filter(id=id).count()
        else:        
            cantidad_reg = Asamblea.objects.filter(id__in=lista).count()    
    total_paginas = int(cantidad_reg/filas_hoja)
    if total_paginas<=0:
        total_paginas=1
    request.session["filas_hoja"] =filas_hoja
    request.session["total_paginas"] = total_paginas
    paginas=[]
    n=0
    for i in range(total_paginas):
        n = n+1
        paginas.append(n)
    request.session["tabla"] = ""    
    if  tipo == 1:
        titulo = 'REPORTE DE ASAMBLEAS '
        retorno = 'reporte_asambleas_pdf'    
    if tipo == 2:
        titulo = 'REPORTE DE DECISIONES ASAMBLEAS '
        retorno = 'reporte_decisiones_asamblea_pdf'
    request.session["titulo"] = titulo
    request.session["idreunion"] = id
    return render(request, "core/paginas.html", {"paginas":paginas,"titulo":titulo,"retorno":retorno})

def PaginasReporteAsambleasPdfView(request,pagina,tipo):
    id = request.session["idreunion"]
    detalle = request.session['detalle_reporte_asamblea']
    titulo = request.session["titulo"] 
    filas_hoja = request.session["filas_hoja"]
    ultima_pagina = request.session["total_paginas"] 
    lista = request.session['lista_filtrada'] 
    #Indicamos el tipo de contenido a devolver, en este caso un pdf
    response = HttpResponse(content_type='application/pdf')
    #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    buffer = BytesIO()
    #Canvas nos permite hacer el reporte con coordenadas X y Y
    pdf = Canvas(buffer)
    #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    cabecera_asamblea_pdf(pdf,pagina,titulo)
    #Con show page hacemos un corte de página para pasar a la siguiente
    if tipo == 2:
        tabla_decisiones_asamblea_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,id)
    if tipo == 1:
        tabla_asamblea_pdf(pdf,pagina,filas_hoja,lista,ultima_pagina,detalle,id)        
    pdf.showPage()
    pdf.save()
    pdf = buffer.getvalue()
    response.write(pdf)
    return response

class TipoReportesAsambleasPdfView(TemplateView):
    template_name = 'core/reportes_asambleas_pdf.html'

def AjaxGuardaIdAsambleaView(request):
    id = request.GET.get('id', None)
    request.session["idasamblea"] = int(id)
    data = {'id':id}
    return JsonResponse(data)    

def ReporteAsambleasXlsView(request):
    detalle = request.session['detalle_reporte_asamblea']
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    lista = request.session['lista_filtrada']
    asamblea = Asamblea.objects.filter(id__in=lista)
    n = 2
    p = 1
    asamblea = Asamblea.objects.filter(id__in=lista)
    for j in asamblea :
        worksheet.write('A1','Fecha' )
        worksheet.write('B1','Acta No.' )
        worksheet.write('C1','Hora Inicio' )
        worksheet.write('D1','Hora Final' )
        worksheet.write('E1','Tiempo Asamblea' )
        nn = str(n)
        if j.fecha != None :
            sfecha= j.fecha.strftime("%m/%d/%Y")
        else:
            sfecha=''    
        exec("worksheet.write('A"+nn+"','"+sfecha+"' )")
        exec("worksheet.write('B"+nn+"','"+j.numero_acta+"' )")
        tiempo_asamblea = float((datetime.strptime(str(j.hora_final), '%H:%M:%S')-datetime.strptime(str(j.hora_inicio), '%H:%M:%S')).seconds/3600.0)
        exec("worksheet.write('C"+nn+"','"+str(j.hora_inicio)+"' )")
        exec("worksheet.write('D"+nn+"','"+str(j.hora_final)+"' )")
        exec("worksheet.write('E"+nn+"','"+str(tiempo_asamblea)+"' )")      
        n += 1
        nn = str(n)
        asamblea = Asamblea.objects.get(id=j.id)
        decisiones_asamblea = DecisionAsamblea.objects.filter(asamblea_id=asamblea.id)
        exec("worksheet.write('A"+nn+"','Decisión')")
        exec("worksheet.write('B"+nn+"','Votos a Favor')")
        exec("worksheet.write('C"+nn+"','Votos Contra')")
        exec("worksheet.write('D"+nn+"','Total Decisiones')")
        n += 1
        nn = str(n)
        tdecision = 0
        for k in decisiones_asamblea:
            exec("worksheet.write('A"+nn+"','"+k.decision+"' )")
            exec("worksheet.write('B"+nn+"','"+str(k.numero_votos_favor)+"' )")
            exec("worksheet.write('C"+nn+"','"+str(k.numero_votos_contra)+"' )")
            n += 1
            nn = str(n)
            tdecision += 1
        d = n-1
        dd = str(d)
        exec("worksheet.write('D"+dd+"','"+str(tdecision)+"' )")    
                    
    workbook.close()
    output.seek(0)
    filename='informe_asamblea.xlsx'
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename="+filename
    return response    


class BorraAnexoAsambleaView(LoginRequiredMixin,DeleteView):
    model = AnexoAsamblea
    template_name = 'core/confirmar_borrado_registro.html'

    def get_success_url(self):
        idasamblea = self.object.asamblea_id
        asamblea = Asamblea.objects.get(id=idasamblea)
        return reverse_lazy('detalle_asamblea',args=[asamblea.id])

def AsambleaContenidoView(request,id):
    asamblea = Asamblea.objects.filter(id=id)
    return render(request, "core/asamblea_contenido.html", {'asamblea':asamblea})

@csrf_exempt
def GuardaFechaAsambleaView(request):    
    fecha_asamblea = datetime.strptime(request.GET.get('fecha'), "%d/%m/%Y").date()
    print(fecha_asamblea)
    request.session['fecha_asamblea'] = fecha_asamblea
    id = 0
    data = {'id':id}
    return JsonResponse(data)  

class CreaTipoAsambleaView(LoginRequiredMixin,CreateView):
    model = TipoAsamblea
    template_name = 'core/tipo_asamblea_form.html'
    form_class = TipoAsambleaForm
    #success_url = reverse_lazy('lista_proveedores')

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        self.request.session['tipo_asamblea'] = self.object.id
        return HttpResponseRedirect(reverse('crea_asamblea'))

def AjaxGuardaIdAsambleaView(request):
    id = request.GET.get('id', None)
    request.session["idasamblea"] = int(id)
    data = {'id':id}
    return JsonResponse(data)        

def PublicaAsambleaView(request,id):
    asamblea = Asamblea.objects.get(id=id)
    idasamblea = id
    anexos = AnexoAsamblea.objects.filter(reunion_id=id,publicar=True)
    envia = request.user.username
    email = request.user.email
    subject ="De: "+envia+" Email:  "+email+" Hora: "+strftime("%Y-%m-%d %H:%M:%S", gmtime())
    message = "Acta Asamblea No."+asamblea.numero_acta
    from_email = settings.EMAIL_HOST_USER
    message = "Acta Asamblea No."+asamblea.numero_acta
    currentDateTime = datetime.datetime.now()
    consejeros = MiembroConsejo.objects.filter(envio_comunicado=True)
    con_lis = list(consejeros.email.items()) 
    listamail= con_lis
    destinatarios = listamail
    mail = EmailMessage(
            subject,
            message,
            from_email,
            to=None,
            bcc=destinatarios,
    )
    for anexo in anexos:
        ruta_archivo = anexo.archivo
        filename = anexo.descripcion
        mail.attach_file(ruta_archivo)
        mail.send()
    return reverse_lazy('detalle_asamblea',args=[idasamblea])


########################## COMITE CONVIVENCIA #####################

@login_required
def ListaComiteConvivenciaView(request):
    comite = ComiteConvivenciaTable(MiembroComiteConvivencia.objects.all().order_by('orden'))
    return render(request, "core/comite_convivencia_lista.html", {"comite":comite})    

class EditaComiteConvivenciaView(UpdateView):
    model = MiembroComiteConvivencia
    fields = ['interior','apartamento','nombre','cargo','email','envio','publicar','comunidad','activo','orden']
    template_name = 'core/comite_convivencia_form.html'
    success_url = reverse_lazy('lista_comite_convivencia')

def CargaFotoComiteConvivenciaView(request,id):
    if request.method == 'POST':
        comite= MiembroComiteConvivencia.objects.get(id=id)
        form = CargaFotoMiembroComiteForm(request.POST, request.FILES,instance=comite)
        form.save()
        return redirect("lista_miembros_comite")
    else:
        form = CargaFotoMiembroComiteForm()

    return render(request,'core/foto_miembro_comite_form.html', locals())

def BorraComiteConvivenciaView(request,pk):
    comite = MiembroComiteConvivencia.objects.get(pk=pk)
    email = comite.email
    MiembroComiteConvivencia.objects.filter(pk=pk).delete()
    return redirect('lista_comite_convivencia')    

class CreaComiteConvivenciaView(LoginRequiredMixin,CreateView):
    model = MiembroComiteConvivencia
    template_name = 'core/comite_convivencia_form.html'
    form_class = ComiteConvivenciaForm
    success_url = reverse_lazy('lista_comite_convivencia')

    def form_valid(self, form):
        if form.is_valid():
            email = form.cleaned_data['email']
            comite = form.save(commit=False)
            comite.save()
        return redirect('lista_comite_convivencia')

def ajaxSwitchEnvioConvivenciaView(request):
    id = request.GET.get('id', None)
    miembro_comite = MiembroComiteConvivencia.objects.get(id=id)
    if miembro_comite.envio == True:
        envio = False
    else:
        envio = True
    miembro_comite = MiembroComiteConvivencia.objects.filter(id=id).update(envio=envio)    
    data = {'envio':envio}
    return JsonResponse(data)    


def ajaxSwitchActivoConvivenciaView(request):
    id = request.GET.get('id', None)
    miembro_comite = MiembroComiteConvivencia.objects.get(id=id)
    if miembro_comite.activo == True:
        activo = False
    else:
        activo = True
    miembro_comite = MiembroComiteConvivencia.objects.filter(id=id).update(activo=activo)
    data = {'activo':activo}
    return JsonResponse(data)    

def ajaxSwitchPublicarConvivenciaView(request):
    id = request.GET.get('id', None)
    miembro_comite = MiembroComiteConvivencia.objects.get(id=id)
    if miembro_comite.publicar == True:
        publicar = False
    else:
        publicar = True
    miembro_comite = MiembroComiteConvivencia.objects.filter(id=id).update(publicar=publicar)
    data = {'publicar':publicar}
    return JsonResponse(data)    

def ajaxSwitchComunidadConvivenciaView(request):
    id = request.GET.get('id', None)
    miembro_comite = MiembroComiteConvivencia.objects.get(id=id)
    if miembro_comite.comunidad == True:
        comunidad = False
    else:
        comunidad = True
    miembro_comite = MiembroComiteConvivencia.objects.filter(id=id).update(comunidad=comunidad)
    data = {'comunidad':comunidad}
    return JsonResponse(data)
    
############################ CONTACTENOS ###########################
 
def ContactView(request):
    #tipo_usuario=request.session['tipo_usuario']
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            subject = "Contacto Web"
            body = {
			'Nombre': form.cleaned_data['Nombre'], 
			'Apellido': form.cleaned_data['Apellido'], 
			'email': form.cleaned_data['email'], 
			'mensaje':form.cleaned_data['mensaje'], 
			}
            conjunto = Conjunto.objects.latest('id')
            message = "\n".join(body.values())
            try:
                send_mail(subject, message, settings.EMAIL_HOST_USER, [conjunto.email])
            except BadHeaderError:
                return HttpResponse('Invalid header found.')
            return redirect('user_home',0)    
    form = ContactForm()
    return render(request, "core/contact_form.html", {'form':form})

####################### EMAIL ################################


class EnvioMailResidentesView(CreateView):

    def get(self,request):
        residentes = Residente.objects.filter(enviar_mail=True)
        context = {
            'residentes':residentes
        }
        return render(request, 'email_form.html',context)

    def post(self,request,*args,**kwargs):
        if request.method == 'POST':
            residentes_ids = request.POST.getlist('ids[]')

            for id in residentes_ids:
                residente = Residente.objects.get(pk=id)
                send_mail(
                    'Mail From Tuts-Station.com',
                    'Welcome to Tuts-Station.com',
                    'bhaveshs.aatmaninfo@gmail.com',
                    [residente.email],
                    fail_silently=False,
                )
            return redirect('pagina_administrador')


class EnvioEmailUnResidenteView(View):
    form_class = EmailForm
    template_name = 'core/email_form.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'email_form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            idresidente = self.request.session['idresidente']
            residente = Residente.objects.get(id=idresidente)
            asunto = form.cleaned_data['asunto']
            mensaje = form.cleaned_data['mensaje']
            email = residente.email
            if request.FILES:
                files = request.FILES.getlist('anexo')
                try:    
                    mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [email])
                    for f in files:
                        mail.attach(f.name, f.read(), f.content_type)
                    mail.send()
                    return render(request, self.template_name, {'email_form': form, 'error_message': 'Correo enviado a %s'%email})
                except:
                    return render(request, self.template_name, {'email_form': form, 'error_message': 'El anexo es muy grande o está corrupto'})
            else:
                mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [email])
                mail.send()
                return render(request, self.template_name, {'email_form': form, 'error_message': 'Correo enviado a %s'%email})
            
        else:
            return redirect('informacion_residentes')
        return render(request, self.template_name, {'email_form': form, 'error_message': 'Imposible de enviar correo. Pruebe más tarde'})

class EnvioEmailAnexoUnResidenteView(View):
    form_class = EmailAnexoForm
    template_name = 'core/email_form.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'email_form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            idresidente = self.request.session['idresidente']
            residente = Residente.objects.get(id=idresidente)
            asunto = form.cleaned_data['asunto']
            mensaje = form.cleaned_data['mensaje']
            email = residente.email
            if request.FILES:
                files = request.FILES.getlist('anexo')
                try:    
                    mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [email])
                    for f in files:
                        mail.attach(f.name, f.read(), f.content_type)
                    mail.send()
                    return render(request, self.template_name, {'email_form': form, 'error_message': 'Correo enviado a %s'%email})
                except:
                    return render(request, self.template_name, {'email_form': form, 'error_message': 'El anexo es muy grande o está corrupto'})
            else:
                mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [email])
                mail.send()
                return render(request, self.template_name, {'email_form': form, 'error_message': 'Correo enviado a %s'%email})
            
        else:
            return redirect('informacion_residentes')
        return render(request, self.template_name, {'email_form': form, 'error_message': 'Imposible de enviar correo. Pruebe más tarde'})

class EnvioEmailAnexoResidentesFiltroView(View):
    form_class = EmailAnexoForm
    template_name = 'core/email_form.html'

    def get(self, request, *args, **kwargs):
        id = request.GET.get('id')
        form = self.form_class()
        return render(request, self.template_name, {'email_form': form})
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            asunto = form.cleaned_data['asunto']
            mensaje = form.cleaned_data['mensaje']
            files = request.FILES.getlist('anexo')
            residentes = Residente.objects.filter(envio_email = True)
            for i in residentes:
                mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [i.email])
                for f in files:
                    mail.attach(f.name, f.read(), f.content_type)
                    mail.send()
                return render(request, self.template_name, {'email_form': form, 'error_message': 'Correo enviado a %s'%residentes})
        else:
            return render(request, self.template_name, {'email_form': form, 'error_message': 'El anexo es muy grande o está corrupto'})
        
        return render(request, self.template_name, {'email_form': form, 'error_message': 'Imposible de enviar correo. Pruebe más tarde'})

class EnvioEmailResidentesFiltroView(View):
    form_class = EmailForm
    template_name = 'core/email_form.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'email_form': form})
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            asunto = form.cleaned_data['asunto']
            mensaje = form.cleaned_data['mensaje']
            residentes = Residente.objects.filter(envio_email = True)
            for i in residentes:
                mail = EmailMessage(asunto, mensaje, settings.EMAIL_HOST_USER, [i.email])
                mail.send()
            return render(request, self.template_name, {'email_form': form, 'error_message': 'Correo enviado a %s'%residentes})
        else:
            return render(request, self.template_name, {'email_form': form, 'error_message': 'Imposible de enviar correo. Pruebe más tarde'})
        
  
def EnvioMailResidente(idapartamento,mensaje,asunto,email):
    apartamento = Apartamento.objects.get(id=idapartamento)
    if email == '':
        residente = Residente.objects.filter(apartamento_id=idapartamento).exclude(email=None)
    else:
        residente = Residente.objects.filter(email=email)
    for i in residente:
        nombre = i.nombre
        email = i.email
        if i.email != None:
            message ="Sres. "+"Apartamento <"+apartamento.numero+"> "+i.nombre+" "+mensaje
            subject = asunto
            from_email = settings.EMAIL_HOST_USER
            mail = EmailMessage(
            subject,
            message,
            from_email,
            to=[email],
            bcc=None,
            )
            mail.send()
    return email

def EnvioMailOtros(mensaje,asunto,email):
    if email != None:
        subject = asunto
        from_email = settings.EMAIL_HOST_USER
        mail = EmailMessage(
        subject,
        mensaje,
        from_email,
        to=[email],
        bcc=None,
        )
        mail.send()
    return email

""" def EnvioMailResidentes(mensaje,asunto):
    residente = Residente.objects.all().exclude(email=None)

    for i in residente:
        time.sleep(.5)
        nombre = i.nombre
        email = i.email
        apartamento = Apartamento.objects.get(id=i.apartamento_id)
        if i.email != None:
            message ="Sres. "+"Apartamento <"+apartamento.numero+"> "+i.nombre+" "+mensaje
            subject = asunto
            from_email = settings.EMAIL_HOST_USER
            mail = EmailMessage(
            subject,
            message,
            from_email,
            to=[email],
            bcc=None,
            )
            mail.send()
    return email """

def ajaxEmailBuscarApartamentoView(request):
    email = request.GET.get('email', None)
    if Residente.objects.filter(email=email).exists():
        residente = Residente.objects.get(email=email)
        idinterior = residente.interior_id
        idpartamento = residente.apartamento_id
        data = { 'idinterior':idinterior,'idapartamento':idpartamento}
    else:
        data = { 'idinterior':None,'idapartamento':None}
    return JsonResponse(data)


def EnvioMailResidentes(mensaje,asunto):
    destinatarios = []
    residentes = Residente.objects.all()
    for residente in residentes:
        destinatario = {}
        apartamento = Apartamento.objects.get(id=residente.apartamento_id)
        destinatario['apartamento'] = apartamento.numero
        destinatario['nombre'] = residente.nombre
        destinatario['email'] = residente.email
        destinatario['message'] =  message ="Sres. "+"Apartamento <"+apartamento.numero+"> "+residente.nombre+" "+mensaje
        destinatarios.append(destinatario)
    # Recorre los destinatarios
    conjunto = Conjunto.objects.latest('id')
    for destinatario in destinatarios:
        subject = asunto
        from_email = settings.EMAIL_HOST_USER
        """ mail = EmailMessage(
        subject,
        message,
        from_email,
        to=[destinatario['email']],
        bcc=None,
        )
        mail.send() """
        send_mail(subject,destinatario['message'],from_email,[destinatario['email']], fail_silently=False,)
        #time.sleep(.3)
    return 0

